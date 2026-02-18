# SPDX-License-Identifier: Apache-2.0
# SPDX-FileCopyrightText: 2025 - 2026 BMO Soluciones, S.A.
"""Unit tests for ExportService class."""

import importlib.util
from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO
from pathlib import Path

import pytest

from coati_payroll.enums import NominaEstado
from coati_payroll.model import (
    db,
    Empresa,
    Empleado,
    Moneda,
    Nomina,
    NominaDetalle,
    NominaEmpleado,
    Planilla,
    TipoPlanilla,
)


def _import_export_service():
    """Import ExportService directly from file to avoid queue initialization."""
    file_path = (
        Path(__file__).parent.parent.parent / "coati_payroll" / "vistas" / "planilla" / "services" / "export_service.py"
    )
    spec = importlib.util.spec_from_file_location("export_service", file_path)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module.ExportService


# Lazy import to avoid queue initialization issues
ExportService = None


def _prepare_objects_for_export(planilla, nomina):
    """Prepare planilla and nomina objects for export by merging into current session."""
    planilla = db.session.merge(planilla)
    nomina = db.session.merge(nomina)
    # Eagerly load relationships to avoid lazy loading issues
    if planilla.empresa_id:
        _ = planilla.empresa  # Load empresa relationship
    return planilla, nomina


@pytest.fixture
def moneda(app, db_session):
    """Create a Moneda for testing."""
    with app.app_context():
        moneda = Moneda(codigo="USD", nombre="Dollar", simbolo="$", activo=True)
        db_session.add(moneda)
        db_session.commit()
        db_session.refresh(moneda)
        return moneda


@pytest.fixture
def tipo_planilla(app, db_session):
    """Create a TipoPlanilla for testing."""
    with app.app_context():
        tipo_planilla = TipoPlanilla(
            codigo="MONTHLY",
            descripcion="Planilla mensual",
            periodicidad="monthly",
            dias=30,
            periodos_por_anio=12,
            mes_inicio_fiscal=1,
            dia_inicio_fiscal=1,
        )
        db_session.add(tipo_planilla)
        db_session.commit()
        db_session.refresh(tipo_planilla)
        return tipo_planilla


@pytest.fixture
def empresa(app, db_session):
    """Create an Empresa for testing."""
    with app.app_context():
        empresa = Empresa(
            codigo="TEST001",
            razon_social="Test Company S.A.",
            ruc="123456789",
            activo=True,
        )
        db_session.add(empresa)
        db_session.commit()
        db_session.refresh(empresa)
        return empresa


@pytest.fixture
def planilla(app, db_session, tipo_planilla, moneda, empresa):
    """Create a Planilla for testing."""
    with app.app_context():
        planilla = Planilla(
            nombre="Test Planilla",
            descripcion="Planilla de prueba",
            tipo_planilla_id=tipo_planilla.id,
            moneda_id=moneda.id,
            empresa_id=empresa.id,
            activo=True,
        )
        db_session.add(planilla)
        db_session.commit()
        db_session.refresh(planilla)
        return planilla


@pytest.fixture
def planilla_sin_empresa(app, db_session, tipo_planilla, moneda):
    """Create a Planilla without empresa for testing."""
    with app.app_context():
        planilla = Planilla(
            nombre="Test Planilla Sin Empresa",
            descripcion="Planilla sin empresa",
            tipo_planilla_id=tipo_planilla.id,
            moneda_id=moneda.id,
            empresa_id=None,
            activo=True,
        )
        db_session.add(planilla)
        db_session.commit()
        db_session.refresh(planilla)
        return planilla


@pytest.fixture
def empleado(app, db_session, empresa, moneda):
    """Create an Empleado for testing."""
    with app.app_context():
        empleado = Empleado(
            empresa_id=empresa.id,
            codigo_empleado="EMP001",
            primer_nombre="Juan",
            segundo_nombre="Carlos",
            primer_apellido="Perez",
            segundo_apellido="Garcia",
            identificacion_personal="123456789",
            id_seguridad_social="SS001",
            id_fiscal="FISCAL001",
            salario_base=Decimal("1000.00"),
            cargo="Desarrollador",
            moneda_id=moneda.id,
            fecha_alta=date.today() - timedelta(days=365),
            activo=True,
        )
        db_session.add(empleado)
        db_session.commit()
        db_session.refresh(empleado)
        return empleado


@pytest.fixture
def empleado_minimo(app, db_session, empresa, moneda):
    """Create a minimal Empleado for testing (no optional fields)."""
    with app.app_context():
        empleado = Empleado(
            empresa_id=empresa.id,
            codigo_empleado="EMP002",
            primer_nombre="Maria",
            primer_apellido="Lopez",
            identificacion_personal="987654321",
            salario_base=Decimal("800.00"),
            moneda_id=moneda.id,
            fecha_alta=date.today() - timedelta(days=180),
            activo=True,
        )
        db_session.add(empleado)
        db_session.commit()
        db_session.refresh(empleado)
        return empleado


@pytest.fixture
def nomina(app, db_session, planilla):
    """Create a Nomina for testing."""
    with app.app_context():
        nomina = Nomina(
            planilla_id=planilla.id,
            periodo_inicio=date(2025, 1, 1),
            periodo_fin=date(2025, 1, 31),
            generado_por="test_user",
            estado="generated",
        )
        db_session.add(nomina)
        db_session.commit()
        db_session.refresh(nomina)
        return nomina


@pytest.fixture
def nomina_empleado(app, db_session, nomina, empleado):
    """Create a NominaEmpleado for testing."""
    with app.app_context():
        nomina_empleado = NominaEmpleado(
            nomina_id=nomina.id,
            empleado_id=empleado.id,
            salario_bruto=Decimal("1000.00"),
            total_ingresos=Decimal("1200.00"),
            total_deducciones=Decimal("200.00"),
            salario_neto=Decimal("1000.00"),
            sueldo_base_historico=Decimal("1000.00"),
            cargo_snapshot="Desarrollador Senior",
        )
        db_session.add(nomina_empleado)
        db_session.commit()
        db_session.refresh(nomina_empleado)
        return nomina_empleado


@pytest.fixture
def nomina_empleado_minimo(app, db_session, nomina, empleado_minimo):
    """Create a minimal NominaEmpleado for testing."""
    with app.app_context():
        nomina_empleado = NominaEmpleado(
            nomina_id=nomina.id,
            empleado_id=empleado_minimo.id,
            salario_bruto=Decimal("800.00"),
            total_ingresos=Decimal("800.00"),
            total_deducciones=Decimal("0.00"),
            salario_neto=Decimal("800.00"),
            sueldo_base_historico=Decimal("800.00"),
        )
        db_session.add(nomina_empleado)
        db_session.commit()
        db_session.refresh(nomina_empleado)
        return nomina_empleado


@pytest.fixture
def nomina_detalle_prestacion(app, db_session, nomina_empleado):
    """Create a NominaDetalle for prestacion testing."""
    with app.app_context():
        detalle = NominaDetalle(
            nomina_empleado_id=nomina_empleado.id,
            tipo="benefit",
            codigo="VAC001",
            descripcion="Vacaciones",
            monto=Decimal("100.00"),
            orden=1,
        )
        db_session.add(detalle)
        db_session.commit()
        db_session.refresh(detalle)
        return detalle


class TestExportarNominaExcel:
    """Tests for exportar_nomina_excel method."""

    def test_exportar_nomina_excel_success(self, app, db_session, planilla, nomina, nomina_empleado):
        """
        Test that exportar_nomina_excel successfully exports nomina to Excel.

        Setup:
            - Create planilla, nomina, and nomina_empleado

        Action:
            - Call exportar_nomina_excel

        Verification:
            - Returns BytesIO object and filename
            - File is valid Excel format
            - Filename contains planilla name and nomina info
        """
        global ExportService
        if ExportService is None:
            ExportService = _import_export_service()

        with app.app_context():
            planilla, nomina = _prepare_objects_for_export(planilla, nomina)

            from coati_payroll.vistas.planilla.services.export_service import ExportService

            output, filename = ExportService.exportar_nomina_excel(planilla, nomina)

            assert isinstance(output, BytesIO)
            assert isinstance(filename, str)
            assert filename.startswith("nomina_")
            assert planilla.nombre in filename
            assert ".xlsx" in filename

            # Verify file is not empty
            output.seek(0)
            content = output.read()
            assert len(content) > 0

            # Verify it's a valid Excel file (starts with ZIP signature)
            assert content.startswith(b"PK")

    def test_exportar_nomina_excel_with_empresa(self, app, db_session, planilla, nomina, nomina_empleado):
        """
        Test that export includes empresa information when available.

        Setup:
            - Create planilla with empresa, nomina, and nomina_empleado

        Action:
            - Call exportar_nomina_excel

        Verification:
            - Export succeeds
            - File contains empresa data
        """
        global ExportService
        if ExportService is None:
            ExportService = _import_export_service()

        with app.app_context():
            planilla, nomina = _prepare_objects_for_export(planilla, nomina)

            from coati_payroll.vistas.planilla.services.export_service import ExportService

            output, filename = ExportService.exportar_nomina_excel(planilla, nomina)

            assert isinstance(output, BytesIO)
            assert filename is not None

    def test_exportar_nomina_excel_without_empresa(self, app, db_session, planilla_sin_empresa, nomina):
        """
        Test that export works without empresa.

        Setup:
            - Create planilla without empresa, nomina

        Action:
            - Call exportar_nomina_excel

        Verification:
            - Export succeeds even without empresa
        """
        with app.app_context():
            # Create nomina_empleado for this test
            from tests.factories.employee_factory import create_employee

            empleado = create_employee(
                db_session,
                empresa_id=None,
                codigo="EMP003",
                primer_nombre="Test",
                primer_apellido="User",
            )

            nomina_empleado = NominaEmpleado(
                nomina_id=nomina.id,
                empleado_id=empleado.id,
                salario_bruto=Decimal("500.00"),
                total_ingresos=Decimal("500.00"),
                total_deducciones=Decimal("0.00"),
                salario_neto=Decimal("500.00"),
                sueldo_base_historico=Decimal("500.00"),
            )
            db_session.add(nomina_empleado)
            db_session.commit()

            planilla_sin_empresa, nomina = _prepare_objects_for_export(planilla_sin_empresa, nomina)

            from coati_payroll.vistas.planilla.services.export_service import ExportService

            output, filename = ExportService.exportar_nomina_excel(planilla_sin_empresa, nomina)

            assert isinstance(output, BytesIO)
            assert filename is not None

    def test_exportar_nomina_excel_multiple_employees(
        self, app, db_session, planilla, nomina, nomina_empleado, empleado_minimo
    ):
        """
        Test that export handles multiple employees correctly.

        Setup:
            - Create planilla, nomina with multiple nomina_empleados

        Action:
            - Call exportar_nomina_excel

        Verification:
            - Export succeeds with multiple employees
        """
        with app.app_context():
            # Add second employee
            nomina_empleado2 = NominaEmpleado(
                nomina_id=nomina.id,
                empleado_id=empleado_minimo.id,
                salario_bruto=Decimal("800.00"),
                total_ingresos=Decimal("800.00"),
                total_deducciones=Decimal("0.00"),
                salario_neto=Decimal("800.00"),
                sueldo_base_historico=Decimal("800.00"),
            )
            db_session.add(nomina_empleado2)
            db_session.commit()

            planilla, nomina = _prepare_objects_for_export(planilla, nomina)

            from coati_payroll.vistas.planilla.services.export_service import ExportService

            output, filename = ExportService.exportar_nomina_excel(planilla, nomina)

            assert isinstance(output, BytesIO)
            assert filename is not None

    def test_exportar_nomina_excel_empty_nomina(self, app, db_session, planilla, nomina):
        """
        Test that export works with empty nomina (no employees).

        Setup:
            - Create planilla and nomina without nomina_empleados

        Action:
            - Call exportar_nomina_excel

        Verification:
            - Export succeeds even with no employees
        """
        with app.app_context():
            planilla, nomina = _prepare_objects_for_export(planilla, nomina)

            from coati_payroll.vistas.planilla.services.export_service import ExportService

            output, filename = ExportService.exportar_nomina_excel(planilla, nomina)

            assert isinstance(output, BytesIO)
            assert filename is not None

    def test_exportar_nomina_excel_missing_openpyxl(self, app, db_session, planilla, nomina, monkeypatch):
        """
        Test that export raises ImportError when openpyxl is not available.

        Setup:
            - Mock check_openpyxl_available to return None

        Action:
            - Call exportar_nomina_excel

        Verification:
            - Raises ImportError with appropriate message
        """
        with app.app_context():
            # Import the module to patch the function where it's used
            from coati_payroll.vistas.planilla.services import export_service

            def mock_check_openpyxl():
                return None

            # Patch the function in the module where it's imported and used
            monkeypatch.setattr(export_service, "check_openpyxl_available", mock_check_openpyxl)

            planilla, nomina = _prepare_objects_for_export(planilla, nomina)

            from coati_payroll.vistas.planilla.services.export_service import ExportService

            with pytest.raises(ImportError, match="openpyxl no está disponible"):
                ExportService.exportar_nomina_excel(planilla, nomina)

    def test_exportar_nomina_excel_incluye_encabezado_extendido_y_auditoria(
        self, app, db_session, planilla, nomina, nomina_empleado
    ):
        """Header must include planilla metadata and user traceability."""
        from openpyxl import load_workbook
        from coati_payroll.vistas.planilla.services.export_service import ExportService

        with app.app_context():
            planilla.estado_aprobacion = "approved"
            nomina.estado = "applied"
            nomina.generado_por = "creator_user"
            nomina.aprobado_por = "approver_user"
            nomina.aplicado_por = "applier_user"
            db_session.commit()

            planilla, nomina = _prepare_objects_for_export(planilla, nomina)
            output, _filename = ExportService.exportar_nomina_excel(planilla, nomina)

            wb = load_workbook(output)
            ws = wb.active

            labels = {}
            for row in range(1, 80):
                key = ws.cell(row=row, column=1).value
                if key:
                    labels[str(key)] = ws.cell(row=row, column=2).value

            assert labels.get("ID Empresa:") == planilla.empresa_id
            assert labels.get("ID Planilla:") == planilla.id
            assert labels.get("Status Planilla:") == "approved"
            assert labels.get("Estado Nomina:") == "applied"
            assert labels.get("Creado por:") == "creator_user"
            assert labels.get("Aprobado por:") == "approver_user"
            assert labels.get("Aplicado por:") == "applier_user"

    def test_exportar_nomina_excel_estructura_secciones_dinamicas(
        self, app, db_session, planilla, nomina, nomina_empleado
    ):
        """Export must render the 5 sections and dynamic concept columns."""
        from openpyxl import load_workbook
        from coati_payroll.model import (
            Percepcion,
            Deduccion,
            Prestacion,
            PlanillaIngreso,
            PlanillaDeduccion,
            PlanillaPrestacion,
            NominaDetalle,
        )
        from coati_payroll.vistas.planilla.services.export_service import ExportService

        with app.app_context():
            percepcion = Percepcion(codigo="P_DYN", nombre="Ingreso Dinamico", formula_tipo="fixed", activo=True)
            deduccion = Deduccion(codigo="D_DYN", nombre="Deduccion Dinamica", formula_tipo="fixed", activo=True)
            prestacion = Prestacion(codigo="B_DYN", nombre="Prestacion Dinamica", formula_tipo="fixed", activo=True)
            db_session.add_all([percepcion, deduccion, prestacion])
            db_session.flush()

            db_session.add(PlanillaIngreso(planilla_id=planilla.id, percepcion_id=percepcion.id, orden=1, activo=True))
            db_session.add(PlanillaDeduccion(planilla_id=planilla.id, deduccion_id=deduccion.id, prioridad=1, activo=True))
            db_session.add(PlanillaPrestacion(planilla_id=planilla.id, prestacion_id=prestacion.id, orden=1, activo=True))
            db_session.add_all(
                [
                    NominaDetalle(
                        nomina_empleado_id=nomina_empleado.id,
                        tipo="income",
                        codigo=percepcion.codigo,
                        descripcion=percepcion.nombre,
                        monto=Decimal("100.00"),
                        orden=1,
                        percepcion_id=percepcion.id,
                    ),
                    NominaDetalle(
                        nomina_empleado_id=nomina_empleado.id,
                        tipo="deduction",
                        codigo=deduccion.codigo,
                        descripcion=deduccion.nombre,
                        monto=Decimal("10.00"),
                        orden=1,
                        deduccion_id=deduccion.id,
                    ),
                    NominaDetalle(
                        nomina_empleado_id=nomina_empleado.id,
                        tipo="benefit",
                        codigo=prestacion.codigo,
                        descripcion=prestacion.nombre,
                        monto=Decimal("25.00"),
                        orden=1,
                        prestacion_id=prestacion.id,
                    ),
                ]
            )
            db_session.commit()

            planilla, nomina = _prepare_objects_for_export(planilla, nomina)
            output, _filename = ExportService.exportar_nomina_excel(planilla, nomina)

            wb = load_workbook(output)
            ws = wb.active

            section_row = None
            for row in range(1, 120):
                values = [ws.cell(row=row, column=col).value for col in range(1, ws.max_column + 1)]
                if "Informacion del Empleado" in values and "Ingresos" in values:
                    section_row = row
                    break
            assert section_row is not None

            section_values = [ws.cell(row=section_row, column=col).value for col in range(1, ws.max_column + 1)]
            non_empty_sections = [v for v in section_values if v]
            assert "Informacion del Empleado" in non_empty_sections
            assert "Ingresos" in non_empty_sections
            assert "Deducciones" in non_empty_sections
            assert "Total a Pagar" in non_empty_sections
            assert "Prestaciones Laborales" in non_empty_sections
            assert sum(1 for value in section_values if value in (None, "")) >= 4

            header_row = section_row + 1
            headers = [ws.cell(row=header_row, column=col).value for col in range(1, ws.max_column + 1)]
            assert "Codigo" in headers
            assert "Nombre Completo" in headers
            assert "Salario Bruto" in headers
            assert "Ingreso Dinamico" in headers
            assert "Deduccion Dinamica" in headers
            assert "Prestacion Dinamica" in headers
            assert "Total Ingresos" in headers
            assert "Total Deducciones" in headers
            assert "Salario Neto" in headers
            assert "Total Prestaciones" in headers

    def test_exportar_nomina_excel_ajuste_reclasificacion_salario_bruto(
        self, app, db_session, planilla, nomina, nomina_empleado
    ):
        """Gross salary column must be visually adjusted by reclassification incomes."""
        from openpyxl import load_workbook
        from coati_payroll.model import Percepcion, PlanillaIngreso, NominaDetalle
        from coati_payroll.vistas.planilla.services.export_service import ExportService

        with app.app_context():
            percepcion = Percepcion(
                codigo="VAC_RECLAS",
                nombre="Vacaciones Descansadas",
                formula_tipo="fixed",
                activo=True,
                mostrar_como_ingreso_reportes=False,
            )
            db_session.add(percepcion)
            db_session.flush()

            nomina_empleado.salario_bruto = Decimal("1000.00")
            db_session.add(PlanillaIngreso(planilla_id=planilla.id, percepcion_id=percepcion.id, orden=1, activo=True))
            db_session.add(
                NominaDetalle(
                    nomina_empleado_id=nomina_empleado.id,
                    tipo="income",
                    codigo=percepcion.codigo,
                    descripcion=percepcion.nombre,
                    monto=Decimal("125.00"),
                    orden=1,
                    percepcion_id=percepcion.id,
                )
            )
            db_session.commit()

            planilla, nomina = _prepare_objects_for_export(planilla, nomina)
            output, _filename = ExportService.exportar_nomina_excel(planilla, nomina)

            wb = load_workbook(output)
            ws = wb.active

            header_row = None
            for row in range(1, 120):
                row_values = [ws.cell(row=row, column=col).value for col in range(1, ws.max_column + 1)]
                if "Salario Bruto" in row_values and "Total Ingresos" in row_values:
                    header_row = row
                    break
            assert header_row is not None

            headers = [ws.cell(row=header_row, column=col).value for col in range(1, ws.max_column + 1)]
            salario_bruto_col = headers.index("Salario Bruto") + 1
            reclas_col = headers.index("Vacaciones Descansadas") + 1

            data_row = header_row + 1
            assert float(ws.cell(row=data_row, column=salario_bruto_col).value) == pytest.approx(875.00)
            assert float(ws.cell(row=data_row, column=reclas_col).value) == pytest.approx(125.00)

    def test_exportar_nomina_excel_incluye_provision_vacaciones(
        self, app, db_session, planilla, nomina, nomina_empleado
    ):
        """Benefits section must include vacation liability from persisted voucher lines."""
        from openpyxl import load_workbook
        from coati_payroll.model import (
            Prestacion,
            PlanillaPrestacion,
            NominaDetalle,
            ComprobanteContable,
            ComprobanteContableLinea,
        )
        from coati_payroll.vistas.planilla.services.export_service import ExportService

        with app.app_context():
            prestacion = Prestacion(codigo="BEN_A", nombre="Prestacion Base", formula_tipo="fixed", activo=True)
            db_session.add(prestacion)
            db_session.flush()
            db_session.add(PlanillaPrestacion(planilla_id=planilla.id, prestacion_id=prestacion.id, orden=1, activo=True))
            db_session.add(
                NominaDetalle(
                    nomina_empleado_id=nomina_empleado.id,
                    tipo="benefit",
                    codigo=prestacion.codigo,
                    descripcion=prestacion.nombre,
                    monto=Decimal("50.00"),
                    orden=1,
                    prestacion_id=prestacion.id,
                )
            )

            comprobante = ComprobanteContable(
                nomina_id=nomina.id,
                fecha_calculo=nomina.periodo_fin,
                concepto="Voucher test",
                moneda_id=planilla.moneda_id,
            )
            db_session.add(comprobante)
            db_session.flush()
            db_session.add(
                ComprobanteContableLinea(
                    comprobante_id=comprobante.id,
                    nomina_empleado_id=nomina_empleado.id,
                    empleado_id=nomina_empleado.empleado_id,
                    empleado_codigo=nomina_empleado.empleado.codigo_empleado,
                    empleado_nombre=f"{nomina_empleado.empleado.primer_nombre} {nomina_empleado.empleado.primer_apellido}",
                    codigo_cuenta="2199",
                    descripcion_cuenta="Pasivo vacaciones",
                    centro_costos="CC-001",
                    tipo_debito_credito="credito",
                    debito=Decimal("0.00"),
                    credito=Decimal("123.45"),
                    monto_calculado=Decimal("123.45"),
                    concepto="Vacaciones pagadas",
                    tipo_concepto="vacation_liability",
                    concepto_codigo="VAC_PAID_LIAB",
                    orden=1,
                )
            )
            db_session.commit()

            planilla, nomina = _prepare_objects_for_export(planilla, nomina)
            output, _filename = ExportService.exportar_nomina_excel(planilla, nomina)

            wb = load_workbook(output)
            ws = wb.active

            header_row = None
            for row in range(1, 120):
                row_values = [ws.cell(row=row, column=col).value for col in range(1, ws.max_column + 1)]
                if "Total Prestaciones" in row_values:
                    header_row = row
                    break
            assert header_row is not None

            headers = [ws.cell(row=header_row, column=col).value for col in range(1, ws.max_column + 1)]
            vac_col = headers.index("Provision de Vacaciones") + 1
            total_prest_col = headers.index("Total Prestaciones") + 1

            data_row = header_row + 1
            assert float(ws.cell(row=data_row, column=vac_col).value) == pytest.approx(123.45)
            assert float(ws.cell(row=data_row, column=total_prest_col).value) == pytest.approx(173.45)


class TestExportarPrestacionesExcel:
    """Tests for exportar_prestaciones_excel method."""

    def test_exportar_prestaciones_excel_success(
        self, app, db_session, planilla, nomina, nomina_empleado, nomina_detalle_prestacion
    ):
        """
        Test that exportar_prestaciones_excel successfully exports prestaciones to Excel.

        Setup:
            - Create planilla, nomina, nomina_empleado, and nomina_detalle with prestacion

        Action:
            - Call exportar_prestaciones_excel

        Verification:
            - Returns BytesIO object and filename
            - File is valid Excel format
            - Filename contains planilla name and nomina info
        """
        with app.app_context():
            planilla, nomina = _prepare_objects_for_export(planilla, nomina)

            from coati_payroll.vistas.planilla.services.export_service import ExportService

            output, filename = ExportService.exportar_prestaciones_excel(planilla, nomina)

            assert isinstance(output, BytesIO)
            assert isinstance(filename, str)
            assert filename.startswith("prestaciones_")
            assert planilla.nombre in filename
            assert ".xlsx" in filename

            # Verify file is not empty
            output.seek(0)
            content = output.read()
            assert len(content) > 0

            # Verify it's a valid Excel file
            assert content.startswith(b"PK")

    def test_exportar_prestaciones_excel_with_empresa(
        self, app, db_session, planilla, nomina, nomina_empleado, nomina_detalle_prestacion
    ):
        """
        Test that export includes empresa information when available.

        Setup:
            - Create planilla with empresa, nomina, nomina_empleado, and prestacion

        Action:
            - Call exportar_prestaciones_excel

        Verification:
            - Export succeeds
        """
        with app.app_context():
            planilla, nomina = _prepare_objects_for_export(planilla, nomina)

            from coati_payroll.vistas.planilla.services.export_service import ExportService

            output, filename = ExportService.exportar_prestaciones_excel(planilla, nomina)

            assert isinstance(output, BytesIO)
            assert filename is not None

    def test_exportar_prestaciones_excel_multiple_prestaciones(
        self, app, db_session, planilla, nomina, nomina_empleado
    ):
        """
        Test that export handles multiple prestaciones correctly.

        Setup:
            - Create nomina with multiple prestaciones

        Action:
            - Call exportar_prestaciones_excel

        Verification:
            - Export succeeds with multiple prestaciones
        """
        with app.app_context():
            # Add multiple prestaciones
            detalle1 = NominaDetalle(
                nomina_empleado_id=nomina_empleado.id,
                tipo="benefit",
                codigo="VAC001",
                descripcion="Vacaciones",
                monto=Decimal("100.00"),
                orden=1,
            )
            detalle2 = NominaDetalle(
                nomina_empleado_id=nomina_empleado.id,
                tipo="benefit",
                codigo="BON001",
                descripcion="Bono",
                monto=Decimal("50.00"),
                orden=2,
            )
            db_session.add(detalle1)
            db_session.add(detalle2)
            db_session.commit()

            planilla, nomina = _prepare_objects_for_export(planilla, nomina)

            from coati_payroll.vistas.planilla.services.export_service import ExportService

            output, filename = ExportService.exportar_prestaciones_excel(planilla, nomina)

            assert isinstance(output, BytesIO)
            assert filename is not None

    def test_exportar_prestaciones_excel_no_prestaciones(self, app, db_session, planilla, nomina, nomina_empleado):
        """
        Test that export works when there are no prestaciones.

        Setup:
            - Create nomina with nomina_empleado but no prestaciones

        Action:
            - Call exportar_prestaciones_excel

        Verification:
            - Export succeeds even with no prestaciones
        """
        with app.app_context():
            planilla, nomina = _prepare_objects_for_export(planilla, nomina)

            from coati_payroll.vistas.planilla.services.export_service import ExportService

            output, filename = ExportService.exportar_prestaciones_excel(planilla, nomina)

            assert isinstance(output, BytesIO)
            assert filename is not None

    def test_exportar_prestaciones_excel_empty_nomina(self, app, db_session, planilla, nomina):
        """
        Test that export works with empty nomina (no employees).

        Setup:
            - Create planilla and nomina without nomina_empleados

        Action:
            - Call exportar_prestaciones_excel

        Verification:
            - Export succeeds even with no employees
        """
        with app.app_context():
            planilla, nomina = _prepare_objects_for_export(planilla, nomina)

            from coati_payroll.vistas.planilla.services.export_service import ExportService

            output, filename = ExportService.exportar_prestaciones_excel(planilla, nomina)

            assert isinstance(output, BytesIO)
            assert filename is not None

    def test_exportar_prestaciones_excel_incluye_encabezado_y_auditoria(
        self, app, db_session, planilla, nomina, nomina_empleado, nomina_detalle_prestacion
    ):
        """Export must include planilla id, nomina status and audit marks."""
        from openpyxl import load_workbook
        from coati_payroll.vistas.planilla.services.export_service import ExportService

        with app.app_context():
            nomina.generado_por = "creator_user"
            nomina.aprobado_por = "approver_user"
            nomina.aplicado_por = "applier_user"
            nomina.estado = "applied"
            db_session.commit()

            planilla, nomina = _prepare_objects_for_export(planilla, nomina)
            output, _filename = ExportService.exportar_prestaciones_excel(planilla, nomina)

            wb = load_workbook(output)
            ws = wb.active

            labels = {}
            for row in range(1, 50):
                key = ws[f"A{row}"].value
                if key:
                    labels[key] = ws[f"B{row}"].value

            assert labels.get("ID Planilla:") == planilla.id
            assert labels.get("Estado Nómina (Generado, Aprobado, Aplicado):") == "applied"
            assert labels.get("Creado por:") == "creator_user"
            assert labels.get("Aprobado por:") == "approver_user"
            assert labels.get("Aplicado por:") == "applier_user"

    def test_exportar_prestaciones_excel_incluye_provision_vacaciones_desde_comprobante(
        self, app, db_session, planilla, nomina, nomina_empleado, nomina_detalle_prestacion
    ):
        """Export must include vacation liability column and amount from voucher lines."""
        from openpyxl import load_workbook
        from coati_payroll.model import ComprobanteContable, ComprobanteContableLinea, VacationPolicy
        from coati_payroll.vistas.planilla.services.export_service import ExportService

        with app.app_context():
            policy = VacationPolicy(
                codigo="VAC-POL-EXP-001",
                nombre="Vacation Export Policy",
                planilla_id=planilla.id,
                son_vacaciones_pagadas=True,
            )
            db_session.add(policy)
            db_session.flush()

            planilla.vacation_policy_id = policy.id

            comprobante = ComprobanteContable(
                nomina_id=nomina.id,
                fecha_calculo=nomina.periodo_fin,
                concepto="Comprobante test",
                moneda_id=planilla.moneda_id,
            )
            db_session.add(comprobante)
            db_session.flush()

            linea = ComprobanteContableLinea(
                comprobante_id=comprobante.id,
                nomina_empleado_id=nomina_empleado.id,
                empleado_id=nomina_empleado.empleado_id,
                empleado_codigo=nomina_empleado.empleado.codigo_empleado,
                empleado_nombre=f"{nomina_empleado.empleado.primer_nombre} {nomina_empleado.empleado.primer_apellido}",
                codigo_cuenta="2199",
                descripcion_cuenta="Pasivo vacaciones",
                centro_costos="CC-001",
                tipo_debito_credito="credito",
                debito=Decimal("0.00"),
                credito=Decimal("123.45"),
                monto_calculado=Decimal("123.45"),
                concepto="Vacaciones pagadas",
                tipo_concepto="vacation_liability",
                concepto_codigo="VAC_PAID_LIAB",
                orden=1,
            )
            db_session.add(linea)
            db_session.commit()

            planilla, nomina = _prepare_objects_for_export(planilla, nomina)
            output, _filename = ExportService.exportar_prestaciones_excel(planilla, nomina)

            wb = load_workbook(output)
            ws = wb.active

            header_row = None
            for row in range(1, 60):
                if ws.cell(row=row, column=1).value in ("Cód. Empleado", "Cod. Empleado"):
                    header_row = row
                    break
            assert header_row is not None

            headers = []
            col = 1
            while True:
                value = ws.cell(row=header_row, column=col).value
                if value is None:
                    break
                headers.append(value)
                col += 1

            assert "Provisión de Vacaciones" in headers
            vac_col = headers.index("Provisión de Vacaciones") + 1
            assert float(ws.cell(row=header_row + 1, column=vac_col).value) == pytest.approx(123.45)

    def test_exportar_prestaciones_excel_missing_openpyxl(self, app, db_session, planilla, nomina, monkeypatch):
        """
        Test that export raises ImportError when openpyxl is not available.

        Setup:
            - Mock check_openpyxl_available to return None

        Action:
            - Call exportar_prestaciones_excel

        Verification:
            - Raises ImportError with appropriate message
        """
        with app.app_context():
            # Import the module to patch the function where it's used
            from coati_payroll.vistas.planilla.services import export_service

            def mock_check_openpyxl():
                return None

            # Patch the function in the module where it's imported and used
            monkeypatch.setattr(export_service, "check_openpyxl_available", mock_check_openpyxl)

            # Prepare objects to avoid detached instance errors
            planilla, nomina = _prepare_objects_for_export(planilla, nomina)

            from coati_payroll.vistas.planilla.services.export_service import ExportService

            with pytest.raises(ImportError, match="openpyxl no está disponible"):
                ExportService.exportar_prestaciones_excel(planilla, nomina)


class TestExportarComprobanteExcel:
    """Tests for exportar_comprobante_excel method."""

    def test_exportar_comprobante_excel_requires_comprobante(self, app, db_session, planilla, nomina):
        """
        Test that exportar_comprobante_excel requires an existing comprobante.

        Setup:
            - Create planilla and nomina without comprobante

        Action:
            - Call exportar_comprobante_excel

        Verification:
            - Raises ValueError when comprobante doesn't exist
        """
        global ExportService
        if ExportService is None:
            ExportService = _import_export_service()

        from coati_payroll.vistas.planilla.services.export_service import ExportService

        with app.app_context():
            planilla, nomina = _prepare_objects_for_export(planilla, nomina)

            with pytest.raises(ValueError, match="No existe comprobante contable"):
                ExportService.exportar_comprobante_excel(planilla, nomina)

    def test_exportar_comprobante_excel_success(self, app, db_session, planilla, nomina, moneda, empleado):
        """
        Test successful export of comprobante with complete accounting configuration.

        Setup:
            - Create planilla, nomina, empleado with complete accounting setup
            - Create comprobante contable with lines

        Action:
            - Call exportar_comprobante_excel

        Verification:
            - Returns BytesIO object and filename
            - File is valid Excel format
            - Filename contains planilla name and nomina info
        """
        from coati_payroll.model import ComprobanteContable, ComprobanteContableLinea, NominaEmpleado
        from coati_payroll.vistas.planilla.services.export_service import ExportService
        from datetime import datetime, timezone

        with app.app_context():
            # Create nomina empleado
            ne = NominaEmpleado(
                nomina_id=nomina.id,
                empleado_id=empleado.id,
                salario_bruto=Decimal("1000.00"),
                total_ingresos=Decimal("1000.00"),
                total_deducciones=Decimal("0.00"),
                salario_neto=Decimal("1000.00"),
                sueldo_base_historico=Decimal("1000.00"),
            )
            db_session.add(ne)
            db_session.flush()

            # Create comprobante
            comprobante = ComprobanteContable(
                nomina_id=nomina.id,
                fecha_calculo=date(2025, 1, 31),
                concepto="Nómina Enero 2025",
                moneda_id=moneda.id,
                total_debitos=Decimal("1000.00"),
                total_creditos=Decimal("1000.00"),
                balance=Decimal("0.00"),
                aplicado_por="test_user",
                fecha_aplicacion=datetime.now(timezone.utc),
            )
            db_session.add(comprobante)
            db_session.flush()

            # Create comprobante lines
            linea1 = ComprobanteContableLinea(
                comprobante_id=comprobante.id,
                nomina_empleado_id=ne.id,
                empleado_id=empleado.id,
                empleado_codigo=empleado.codigo_empleado,
                empleado_nombre=f"{empleado.primer_nombre} {empleado.primer_apellido}",
                codigo_cuenta="5101",
                descripcion_cuenta="Gasto por Salario",
                centro_costos="CC-001",
                tipo_debito_credito="debito",
                debito=Decimal("1000.00"),
                credito=Decimal("0.00"),
                monto_calculado=Decimal("1000.00"),
                concepto="Salario Base",
                tipo_concepto="salario_base",
                concepto_codigo="SALARIO_BASE",
                orden=1,
            )
            linea2 = ComprobanteContableLinea(
                comprobante_id=comprobante.id,
                nomina_empleado_id=ne.id,
                empleado_id=empleado.id,
                empleado_codigo=empleado.codigo_empleado,
                empleado_nombre=f"{empleado.primer_nombre} {empleado.primer_apellido}",
                codigo_cuenta="2101",
                descripcion_cuenta="Salario por Pagar",
                centro_costos="CC-001",
                tipo_debito_credito="credito",
                debito=Decimal("0.00"),
                credito=Decimal("1000.00"),
                monto_calculado=Decimal("1000.00"),
                concepto="Salario Base",
                tipo_concepto="salario_base",
                concepto_codigo="SALARIO_BASE",
                orden=2,
            )
            db_session.add(linea1)
            db_session.add(linea2)
            db_session.commit()

            planilla, nomina = _prepare_objects_for_export(planilla, nomina)

            output, filename = ExportService.exportar_comprobante_excel(planilla, nomina)

            assert isinstance(output, BytesIO)
            assert isinstance(filename, str)
            assert filename.startswith("comprobante_")
            assert planilla.nombre in filename
            assert ".xlsx" in filename

            # Verify file is not empty
            output.seek(0)
            content = output.read()
            assert len(content) > 0

            # Verify it's a valid Excel file (starts with ZIP signature)
            assert content.startswith(b"PK")


    def test_exportar_comprobante_excel_incluye_estado_id_y_trazabilidad(
        self, app, db_session, planilla, nomina, moneda, empleado
    ):
        """Test that summarized export includes planilla status/id and user traceability."""
        from openpyxl import load_workbook

        from coati_payroll.model import ComprobanteContable, ComprobanteContableLinea, NominaEmpleado
        from coati_payroll.vistas.planilla.services.export_service import ExportService

        with app.app_context():
            nomina.aprobado_por = "approver_user"
            nomina.aplicado_por = "applier_user"
            db_session.flush()

            ne = NominaEmpleado(
                nomina_id=nomina.id,
                empleado_id=empleado.id,
                salario_bruto=Decimal("1000.00"),
                total_ingresos=Decimal("1000.00"),
                total_deducciones=Decimal("0.00"),
                salario_neto=Decimal("1000.00"),
                sueldo_base_historico=Decimal("1000.00"),
            )
            db_session.add(ne)
            db_session.flush()

            comprobante = ComprobanteContable(
                nomina_id=nomina.id,
                fecha_calculo=date(2025, 1, 31),
                concepto="Nómina Enero 2025",
                moneda_id=moneda.id,
                total_debitos=Decimal("1000.00"),
                total_creditos=Decimal("1000.00"),
                balance=Decimal("0.00"),
            )
            db_session.add(comprobante)
            db_session.flush()

            linea = ComprobanteContableLinea(
                comprobante_id=comprobante.id,
                nomina_empleado_id=ne.id,
                empleado_id=empleado.id,
                empleado_codigo=empleado.codigo_empleado,
                empleado_nombre=f"{empleado.primer_nombre} {empleado.primer_apellido}",
                codigo_cuenta="5101",
                descripcion_cuenta="Gasto por Salario",
                tipo_debito_credito="debito",
                debito=Decimal("1000.00"),
                credito=Decimal("0.00"),
                monto_calculado=Decimal("1000.00"),
                concepto="Salario Base",
                tipo_concepto="salario_base",
                concepto_codigo="SALARIO_BASE",
                orden=1,
            )
            db_session.add(linea)
            db_session.commit()

            planilla, nomina = _prepare_objects_for_export(planilla, nomina)

            output, _filename = ExportService.exportar_comprobante_excel(planilla, nomina)

            wb = load_workbook(output)
            ws = wb.active
            values = {ws[f"A{row}"].value: ws[f"B{row}"].value for row in range(1, ws.max_row + 1)}

            assert values["ID Planilla:"] == planilla.id
            assert values["Estatus Planilla:"] == nomina.estado
            assert values["Creado por:"] == nomina.generado_por
            assert values["Aprobado por:"] == "approver_user"
            assert values["Aplicado por:"] == "applier_user"

    def test_exportar_comprobante_excel_with_empresa(self, app, db_session, planilla, nomina, moneda, empleado):
        """
        Test that export includes empresa information when available.

        Setup:
            - Create planilla with empresa, comprobante with lines

        Action:
            - Call exportar_comprobante_excel

        Verification:
            - Export succeeds with empresa info
        """
        from coati_payroll.model import ComprobanteContable, ComprobanteContableLinea, NominaEmpleado
        from coati_payroll.vistas.planilla.services.export_service import ExportService

        with app.app_context():
            # Create nomina empleado
            ne = NominaEmpleado(
                nomina_id=nomina.id,
                empleado_id=empleado.id,
                salario_bruto=Decimal("1000.00"),
                total_ingresos=Decimal("1000.00"),
                total_deducciones=Decimal("0.00"),
                salario_neto=Decimal("1000.00"),
                sueldo_base_historico=Decimal("1000.00"),
            )
            db_session.add(ne)
            db_session.flush()

            # Create comprobante
            comprobante = ComprobanteContable(
                nomina_id=nomina.id,
                fecha_calculo=date(2025, 1, 31),
                concepto="Nómina Enero 2025",
                moneda_id=moneda.id,
                total_debitos=Decimal("500.00"),
                total_creditos=Decimal("500.00"),
                balance=Decimal("0.00"),
            )
            db_session.add(comprobante)
            db_session.flush()

            # Create comprobante lines
            linea = ComprobanteContableLinea(
                comprobante_id=comprobante.id,
                nomina_empleado_id=ne.id,
                empleado_id=empleado.id,
                empleado_codigo=empleado.codigo_empleado,
                empleado_nombre=f"{empleado.primer_nombre} {empleado.primer_apellido}",
                codigo_cuenta="5101",
                descripcion_cuenta="Gasto por Salario",
                tipo_debito_credito="debito",
                debito=Decimal("500.00"),
                credito=Decimal("0.00"),
                monto_calculado=Decimal("500.00"),
                concepto="Salario Base",
                tipo_concepto="salario_base",
                concepto_codigo="SALARIO_BASE",
                orden=1,
            )
            db_session.add(linea)
            db_session.commit()

            planilla, nomina = _prepare_objects_for_export(planilla, nomina)

            output, filename = ExportService.exportar_comprobante_excel(planilla, nomina)

            assert isinstance(output, BytesIO)
            assert filename is not None

    def test_exportar_comprobante_excel_without_empresa(
        self, app, db_session, planilla_sin_empresa, nomina, moneda, empleado
    ):
        """
        Test that export works without empresa.

        Setup:
            - Create planilla without empresa, comprobante with lines

        Action:
            - Call exportar_comprobante_excel

        Verification:
            - Export succeeds even without empresa
        """
        from coati_payroll.model import ComprobanteContable, ComprobanteContableLinea, NominaEmpleado
        from coati_payroll.vistas.planilla.services.export_service import ExportService

        with app.app_context():
            # Create nomina empleado
            ne = NominaEmpleado(
                nomina_id=nomina.id,
                empleado_id=empleado.id,
                salario_bruto=Decimal("800.00"),
                total_ingresos=Decimal("800.00"),
                total_deducciones=Decimal("0.00"),
                salario_neto=Decimal("800.00"),
                sueldo_base_historico=Decimal("800.00"),
            )
            db_session.add(ne)
            db_session.flush()

            # Create comprobante
            comprobante = ComprobanteContable(
                nomina_id=nomina.id,
                fecha_calculo=date(2025, 1, 31),
                concepto="Nómina Enero 2025",
                moneda_id=moneda.id,
                total_debitos=Decimal("800.00"),
                total_creditos=Decimal("800.00"),
                balance=Decimal("0.00"),
            )
            db_session.add(comprobante)
            db_session.flush()

            # Create comprobante line
            linea = ComprobanteContableLinea(
                comprobante_id=comprobante.id,
                nomina_empleado_id=ne.id,
                empleado_id=empleado.id,
                empleado_codigo=empleado.codigo_empleado,
                empleado_nombre=f"{empleado.primer_nombre} {empleado.primer_apellido}",
                codigo_cuenta="5101",
                descripcion_cuenta="Gasto por Salario",
                tipo_debito_credito="debito",
                debito=Decimal("800.00"),
                credito=Decimal("0.00"),
                monto_calculado=Decimal("800.00"),
                concepto="Salario Base",
                tipo_concepto="salario_base",
                concepto_codigo="SALARIO_BASE",
                orden=1,
            )
            db_session.add(linea)
            db_session.commit()

            planilla_sin_empresa, nomina = _prepare_objects_for_export(planilla_sin_empresa, nomina)

            output, filename = ExportService.exportar_comprobante_excel(planilla_sin_empresa, nomina)

            assert isinstance(output, BytesIO)
            assert filename is not None

    def test_exportar_comprobante_excel_with_warnings(self, app, db_session, planilla, nomina, moneda, empleado):
        """
        Test that export includes warnings section when advertencias exist.

        Setup:
            - Create comprobante with advertencias

        Action:
            - Call exportar_comprobante_excel

        Verification:
            - Export succeeds and includes warnings
        """
        from coati_payroll.model import ComprobanteContable, ComprobanteContableLinea, NominaEmpleado
        from coati_payroll.vistas.planilla.services.export_service import ExportService

        with app.app_context():
            # Create nomina empleado
            ne = NominaEmpleado(
                nomina_id=nomina.id,
                empleado_id=empleado.id,
                salario_bruto=Decimal("1000.00"),
                total_ingresos=Decimal("1000.00"),
                total_deducciones=Decimal("0.00"),
                salario_neto=Decimal("1000.00"),
                sueldo_base_historico=Decimal("1000.00"),
            )
            db_session.add(ne)
            db_session.flush()

            # Create comprobante with warnings
            comprobante = ComprobanteContable(
                nomina_id=nomina.id,
                fecha_calculo=date(2025, 1, 31),
                concepto="Nómina Enero 2025",
                moneda_id=moneda.id,
                total_debitos=Decimal("1000.00"),
                total_creditos=Decimal("1000.00"),
                balance=Decimal("0.00"),
                advertencias=["Cuenta contable no configurada", "Centro de costos faltante"],
            )
            db_session.add(comprobante)
            db_session.flush()

            # Create comprobante line
            linea = ComprobanteContableLinea(
                comprobante_id=comprobante.id,
                nomina_empleado_id=ne.id,
                empleado_id=empleado.id,
                empleado_codigo=empleado.codigo_empleado,
                empleado_nombre=f"{empleado.primer_nombre} {empleado.primer_apellido}",
                codigo_cuenta="5101",
                descripcion_cuenta="Gasto por Salario",
                tipo_debito_credito="debito",
                debito=Decimal("1000.00"),
                credito=Decimal("0.00"),
                monto_calculado=Decimal("1000.00"),
                concepto="Salario Base",
                tipo_concepto="salario_base",
                concepto_codigo="SALARIO_BASE",
                orden=1,
            )
            db_session.add(linea)
            db_session.commit()

            planilla, nomina = _prepare_objects_for_export(planilla, nomina)

            output, filename = ExportService.exportar_comprobante_excel(planilla, nomina)

            assert isinstance(output, BytesIO)
            assert filename is not None

    def test_exportar_comprobante_excel_with_audit_trail(self, app, db_session, planilla, nomina, moneda, empleado):
        """
        Test that export includes audit trail information.

        Setup:
            - Create comprobante with aplicado_por and fecha_aplicacion

        Action:
            - Call exportar_comprobante_excel

        Verification:
            - Export succeeds with audit trail info
        """
        from coati_payroll.model import ComprobanteContable, ComprobanteContableLinea, NominaEmpleado
        from coati_payroll.vistas.planilla.services.export_service import ExportService
        from datetime import datetime, timezone

        with app.app_context():
            # Create nomina empleado
            ne = NominaEmpleado(
                nomina_id=nomina.id,
                empleado_id=empleado.id,
                salario_bruto=Decimal("1000.00"),
                total_ingresos=Decimal("1000.00"),
                total_deducciones=Decimal("0.00"),
                salario_neto=Decimal("1000.00"),
                sueldo_base_historico=Decimal("1000.00"),
            )
            db_session.add(ne)
            db_session.flush()

            # Create comprobante with audit trail
            comprobante = ComprobanteContable(
                nomina_id=nomina.id,
                fecha_calculo=date(2025, 1, 31),
                concepto="Nómina Enero 2025",
                moneda_id=moneda.id,
                total_debitos=Decimal("1000.00"),
                total_creditos=Decimal("1000.00"),
                balance=Decimal("0.00"),
                aplicado_por="admin_user",
                fecha_aplicacion=datetime(2025, 1, 31, 10, 30, 0, tzinfo=timezone.utc),
            )
            db_session.add(comprobante)
            db_session.flush()

            # Create comprobante line
            linea = ComprobanteContableLinea(
                comprobante_id=comprobante.id,
                nomina_empleado_id=ne.id,
                empleado_id=empleado.id,
                empleado_codigo=empleado.codigo_empleado,
                empleado_nombre=f"{empleado.primer_nombre} {empleado.primer_apellido}",
                codigo_cuenta="5101",
                descripcion_cuenta="Gasto por Salario",
                tipo_debito_credito="debito",
                debito=Decimal("1000.00"),
                credito=Decimal("0.00"),
                monto_calculado=Decimal("1000.00"),
                concepto="Salario Base",
                tipo_concepto="salario_base",
                concepto_codigo="SALARIO_BASE",
                orden=1,
            )
            db_session.add(linea)
            db_session.commit()

            planilla, nomina = _prepare_objects_for_export(planilla, nomina)

            output, filename = ExportService.exportar_comprobante_excel(planilla, nomina)

            assert isinstance(output, BytesIO)
            assert filename is not None

    def test_exportar_comprobante_excel_with_modifications(self, app, db_session, planilla, nomina, moneda, empleado):
        """
        Test that export includes modification tracking information.

        Setup:
            - Create comprobante with veces_modificado > 0

        Action:
            - Call exportar_comprobante_excel

        Verification:
            - Export succeeds with modification info
        """
        from coati_payroll.model import ComprobanteContable, ComprobanteContableLinea, NominaEmpleado
        from coati_payroll.vistas.planilla.services.export_service import ExportService
        from datetime import datetime, timezone

        with app.app_context():
            # Create nomina empleado
            ne = NominaEmpleado(
                nomina_id=nomina.id,
                empleado_id=empleado.id,
                salario_bruto=Decimal("1000.00"),
                total_ingresos=Decimal("1000.00"),
                total_deducciones=Decimal("0.00"),
                salario_neto=Decimal("1000.00"),
                sueldo_base_historico=Decimal("1000.00"),
            )
            db_session.add(ne)
            db_session.flush()

            # Create comprobante with modification tracking
            comprobante = ComprobanteContable(
                nomina_id=nomina.id,
                fecha_calculo=date(2025, 1, 31),
                concepto="Nómina Enero 2025",
                moneda_id=moneda.id,
                total_debitos=Decimal("1000.00"),
                total_creditos=Decimal("1000.00"),
                balance=Decimal("0.00"),
                veces_modificado=2,
                modificado_por="supervisor_user",
                fecha_modificacion=datetime(2025, 2, 1, 14, 20, 0, tzinfo=timezone.utc),
            )
            db_session.add(comprobante)
            db_session.flush()

            # Create comprobante line
            linea = ComprobanteContableLinea(
                comprobante_id=comprobante.id,
                nomina_empleado_id=ne.id,
                empleado_id=empleado.id,
                empleado_codigo=empleado.codigo_empleado,
                empleado_nombre=f"{empleado.primer_nombre} {empleado.primer_apellido}",
                codigo_cuenta="5101",
                descripcion_cuenta="Gasto por Salario",
                tipo_debito_credito="debito",
                debito=Decimal("1000.00"),
                credito=Decimal("0.00"),
                monto_calculado=Decimal("1000.00"),
                concepto="Salario Base",
                tipo_concepto="salario_base",
                concepto_codigo="SALARIO_BASE",
                orden=1,
            )
            db_session.add(linea)
            db_session.commit()

            planilla, nomina = _prepare_objects_for_export(planilla, nomina)

            output, filename = ExportService.exportar_comprobante_excel(planilla, nomina)

            assert isinstance(output, BytesIO)
            assert filename is not None

    def test_exportar_comprobante_excel_with_non_zero_balance(
        self, app, db_session, planilla, nomina, moneda, empleado
    ):
        """
        Test that export handles non-zero balance (unbalanced voucher).

        Setup:
            - Create comprobante with balance != 0

        Action:
            - Call exportar_comprobante_excel

        Verification:
            - Raises ValueError for unbalanced voucher
        """
        from coati_payroll.model import ComprobanteContable, ComprobanteContableLinea, NominaEmpleado
        from coati_payroll.vistas.planilla.services.export_service import ExportService

        with app.app_context():
            # Create nomina empleado
            ne = NominaEmpleado(
                nomina_id=nomina.id,
                empleado_id=empleado.id,
                salario_bruto=Decimal("1000.00"),
                total_ingresos=Decimal("1000.00"),
                total_deducciones=Decimal("0.00"),
                salario_neto=Decimal("1000.00"),
                sueldo_base_historico=Decimal("1000.00"),
            )
            db_session.add(ne)
            db_session.flush()

            # Create unbalanced comprobante
            comprobante = ComprobanteContable(
                nomina_id=nomina.id,
                fecha_calculo=date(2025, 1, 31),
                concepto="Nómina Enero 2025",
                moneda_id=moneda.id,
                total_debitos=Decimal("1000.00"),
                total_creditos=Decimal("950.00"),
                balance=Decimal("50.00"),  # Non-zero balance
            )
            db_session.add(comprobante)
            db_session.flush()

            # Create comprobante line
            linea = ComprobanteContableLinea(
                comprobante_id=comprobante.id,
                nomina_empleado_id=ne.id,
                empleado_id=empleado.id,
                empleado_codigo=empleado.codigo_empleado,
                empleado_nombre=f"{empleado.primer_nombre} {empleado.primer_apellido}",
                codigo_cuenta="5101",
                descripcion_cuenta="Gasto por Salario",
                tipo_debito_credito="debito",
                debito=Decimal("1000.00"),
                credito=Decimal("0.00"),
                monto_calculado=Decimal("1000.00"),
                concepto="Salario Base",
                tipo_concepto="salario_base",
                concepto_codigo="SALARIO_BASE",
                orden=1,
            )
            db_session.add(linea)
            db_session.commit()

            planilla, nomina = _prepare_objects_for_export(planilla, nomina)

            with pytest.raises(ValueError, match="balanceado"):
                ExportService.exportar_comprobante_excel(planilla, nomina)

    def test_exportar_comprobante_excel_blocks_generated_with_errors(
        self, app, db_session, planilla, nomina, moneda, empleado
    ):
        """Test summarized export fails when nomina is generated with errors."""
        from coati_payroll.model import ComprobanteContable, ComprobanteContableLinea, NominaEmpleado
        from coati_payroll.vistas.planilla.services.export_service import ExportService

        with app.app_context():
            nomina.estado = NominaEstado.GENERADO_CON_ERRORES
            db_session.commit()

            ne = NominaEmpleado(
                nomina_id=nomina.id,
                empleado_id=empleado.id,
                salario_bruto=Decimal("1000.00"),
                total_ingresos=Decimal("1000.00"),
                total_deducciones=Decimal("0.00"),
                salario_neto=Decimal("1000.00"),
                sueldo_base_historico=Decimal("1000.00"),
            )
            db_session.add(ne)
            db_session.flush()

            comprobante = ComprobanteContable(
                nomina_id=nomina.id,
                fecha_calculo=date(2025, 1, 31),
                concepto="Nómina Enero 2025",
                moneda_id=moneda.id,
                total_debitos=Decimal("1000.00"),
                total_creditos=Decimal("1000.00"),
                balance=Decimal("0.00"),
            )
            db_session.add(comprobante)
            db_session.flush()

            linea = ComprobanteContableLinea(
                comprobante_id=comprobante.id,
                nomina_empleado_id=ne.id,
                empleado_id=empleado.id,
                empleado_codigo=empleado.codigo_empleado,
                empleado_nombre=f"{empleado.primer_nombre} {empleado.primer_apellido}",
                codigo_cuenta="5101",
                descripcion_cuenta="Gasto por Salario",
                tipo_debito_credito="debito",
                debito=Decimal("1000.00"),
                credito=Decimal("0.00"),
                monto_calculado=Decimal("1000.00"),
                concepto="Salario Base",
                tipo_concepto="salario_base",
                concepto_codigo="SALARIO_BASE",
                orden=1,
            )
            db_session.add(linea)
            db_session.commit()

            planilla, nomina = _prepare_objects_for_export(planilla, nomina)

            with pytest.raises(ValueError, match="calculada con errores"):
                ExportService.exportar_comprobante_excel(planilla, nomina)

            output, filename = ExportService.exportar_comprobante_detallado_excel(planilla, nomina)
            assert isinstance(output, BytesIO)
            assert filename is not None

    def test_exportar_comprobante_excel_incomplete_accounting_config(
        self, app, db_session, planilla, nomina, moneda, empleado
    ):
        """
        Test that export raises ValueError when summarize_voucher fails.

        Setup:
            - Create comprobante with NULL account codes (incomplete config)

        Action:
            - Call exportar_comprobante_excel

        Verification:
            - Raises ValueError with message about incomplete configuration
        """
        from coati_payroll.model import ComprobanteContable, ComprobanteContableLinea, NominaEmpleado
        from coati_payroll.vistas.planilla.services.export_service import ExportService

        with app.app_context():
            # Create nomina empleado
            ne = NominaEmpleado(
                nomina_id=nomina.id,
                empleado_id=empleado.id,
                salario_bruto=Decimal("1000.00"),
                total_ingresos=Decimal("1000.00"),
                total_deducciones=Decimal("0.00"),
                salario_neto=Decimal("1000.00"),
                sueldo_base_historico=Decimal("1000.00"),
            )
            db_session.add(ne)
            db_session.flush()

            # Create comprobante
            comprobante = ComprobanteContable(
                nomina_id=nomina.id,
                fecha_calculo=date(2025, 1, 31),
                concepto="Nómina Enero 2025",
                moneda_id=moneda.id,
                total_debitos=Decimal("1000.00"),
                total_creditos=Decimal("1000.00"),
                balance=Decimal("0.00"),
            )
            db_session.add(comprobante)
            db_session.flush()

            # Create comprobante line with NULL codigo_cuenta (incomplete config)
            linea = ComprobanteContableLinea(
                comprobante_id=comprobante.id,
                nomina_empleado_id=ne.id,
                empleado_id=empleado.id,
                empleado_codigo=empleado.codigo_empleado,
                empleado_nombre=f"{empleado.primer_nombre} {empleado.primer_apellido}",
                codigo_cuenta=None,  # NULL account - incomplete config
                descripcion_cuenta=None,
                tipo_debito_credito="debito",
                debito=Decimal("1000.00"),
                credito=Decimal("0.00"),
                monto_calculado=Decimal("1000.00"),
                concepto="Salario Base",
                tipo_concepto="salario_base",
                concepto_codigo="SALARIO_BASE",
                orden=1,
            )
            db_session.add(linea)
            db_session.commit()

            planilla, nomina = _prepare_objects_for_export(planilla, nomina)

            with pytest.raises(ValueError, match="No se puede exportar comprobante sumarizado"):
                ExportService.exportar_comprobante_excel(planilla, nomina)

    def test_exportar_comprobante_excel_missing_openpyxl(self, app, db_session, planilla, nomina, monkeypatch):
        """
        Test that export raises ImportError when openpyxl is not available.

        Setup:
            - Mock check_openpyxl_available to return None

        Action:
            - Call exportar_comprobante_excel

        Verification:
            - Raises ImportError with appropriate message
        """
        from coati_payroll.vistas.planilla.services import export_service

        with app.app_context():

            def mock_check_openpyxl():
                return None

            # Patch the function in the module where it's imported and used
            monkeypatch.setattr(export_service, "check_openpyxl_available", mock_check_openpyxl)

            planilla, nomina = _prepare_objects_for_export(planilla, nomina)

            from coati_payroll.vistas.planilla.services.export_service import ExportService

            with pytest.raises(ImportError, match="openpyxl no está disponible"):
                ExportService.exportar_comprobante_excel(planilla, nomina)


class TestExportarComprobanteDetalladoExcel:
    """Tests for exportar_comprobante_detallado_excel method."""

    def test_exportar_comprobante_detallado_excel_requires_comprobante(self, app, db_session, planilla, nomina):
        """
        Test that exportar_comprobante_detallado_excel requires an existing comprobante.

        Setup:
            - Create planilla and nomina without comprobante

        Action:
            - Call exportar_comprobante_detallado_excel

        Verification:
            - Raises ValueError when comprobante doesn't exist
        """
        global ExportService
        if ExportService is None:
            ExportService = _import_export_service()

        from coati_payroll.vistas.planilla.services.export_service import ExportService

        with app.app_context():
            planilla, nomina = _prepare_objects_for_export(planilla, nomina)

            with pytest.raises(ValueError, match="No existe comprobante contable"):
                ExportService.exportar_comprobante_detallado_excel(planilla, nomina)

    def test_exportar_comprobante_detallado_excel_success(self, app, db_session, planilla, nomina, moneda, empleado):
        """
        Test successful export of detailed comprobante.

        Setup:
            - Create planilla, nomina, empleado with complete accounting setup
            - Create comprobante contable with detailed lines

        Action:
            - Call exportar_comprobante_detallado_excel

        Verification:
            - Returns BytesIO object and filename
            - File is valid Excel format
            - Filename contains planilla name and nomina info
        """
        from coati_payroll.model import ComprobanteContable, ComprobanteContableLinea, NominaEmpleado
        from coati_payroll.vistas.planilla.services.export_service import ExportService

        with app.app_context():
            # Create nomina empleado
            ne = NominaEmpleado(
                nomina_id=nomina.id,
                empleado_id=empleado.id,
                salario_bruto=Decimal("1000.00"),
                total_ingresos=Decimal("1000.00"),
                total_deducciones=Decimal("0.00"),
                salario_neto=Decimal("1000.00"),
                sueldo_base_historico=Decimal("1000.00"),
            )
            db_session.add(ne)
            db_session.flush()

            # Create comprobante
            comprobante = ComprobanteContable(
                nomina_id=nomina.id,
                fecha_calculo=date(2025, 1, 31),
                concepto="Nómina Enero 2025",
                moneda_id=moneda.id,
                total_debitos=Decimal("1000.00"),
                total_creditos=Decimal("1000.00"),
                balance=Decimal("0.00"),
            )
            db_session.add(comprobante)
            db_session.flush()

            # Create comprobante lines
            linea1 = ComprobanteContableLinea(
                comprobante_id=comprobante.id,
                nomina_empleado_id=ne.id,
                empleado_id=empleado.id,
                empleado_codigo=empleado.codigo_empleado,
                empleado_nombre=f"{empleado.primer_nombre} {empleado.primer_apellido}",
                codigo_cuenta="5101",
                descripcion_cuenta="Gasto por Salario",
                centro_costos="CC-001",
                tipo_debito_credito="debito",
                debito=Decimal("1000.00"),
                credito=Decimal("0.00"),
                monto_calculado=Decimal("1000.00"),
                concepto="Salario Base",
                tipo_concepto="salario_base",
                concepto_codigo="SALARIO_BASE",
                orden=1,
            )
            linea2 = ComprobanteContableLinea(
                comprobante_id=comprobante.id,
                nomina_empleado_id=ne.id,
                empleado_id=empleado.id,
                empleado_codigo=empleado.codigo_empleado,
                empleado_nombre=f"{empleado.primer_nombre} {empleado.primer_apellido}",
                codigo_cuenta="2101",
                descripcion_cuenta="Salario por Pagar",
                centro_costos="CC-001",
                tipo_debito_credito="credito",
                debito=Decimal("0.00"),
                credito=Decimal("1000.00"),
                monto_calculado=Decimal("1000.00"),
                concepto="Salario Base",
                tipo_concepto="salario_base",
                concepto_codigo="SALARIO_BASE",
                orden=2,
            )
            db_session.add(linea1)
            db_session.add(linea2)
            db_session.commit()

            planilla, nomina = _prepare_objects_for_export(planilla, nomina)

            output, filename = ExportService.exportar_comprobante_detallado_excel(planilla, nomina)

            assert isinstance(output, BytesIO)
            assert isinstance(filename, str)
            assert filename.startswith("comprobante_detallado_")
            assert planilla.nombre in filename
            assert ".xlsx" in filename

            # Verify file is not empty
            output.seek(0)
            content = output.read()
            assert len(content) > 0

            # Verify it's a valid Excel file (starts with ZIP signature)
            assert content.startswith(b"PK")


    def test_exportar_comprobante_detallado_excel_incluye_estado_id_y_trazabilidad(
        self, app, db_session, planilla, nomina, moneda, empleado
    ):
        """Test that detailed export includes planilla status/id and user traceability."""
        from openpyxl import load_workbook

        from coati_payroll.model import ComprobanteContable, ComprobanteContableLinea, NominaEmpleado
        from coati_payroll.vistas.planilla.services.export_service import ExportService

        with app.app_context():
            nomina.aprobado_por = "approver_user"
            nomina.aplicado_por = "applier_user"
            db_session.flush()

            ne = NominaEmpleado(
                nomina_id=nomina.id,
                empleado_id=empleado.id,
                salario_bruto=Decimal("1000.00"),
                total_ingresos=Decimal("1000.00"),
                total_deducciones=Decimal("0.00"),
                salario_neto=Decimal("1000.00"),
                sueldo_base_historico=Decimal("1000.00"),
            )
            db_session.add(ne)
            db_session.flush()

            comprobante = ComprobanteContable(
                nomina_id=nomina.id,
                fecha_calculo=date(2025, 1, 31),
                concepto="Nómina Enero 2025",
                moneda_id=moneda.id,
                total_debitos=Decimal("1000.00"),
                total_creditos=Decimal("1000.00"),
                balance=Decimal("0.00"),
            )
            db_session.add(comprobante)
            db_session.flush()

            linea = ComprobanteContableLinea(
                comprobante_id=comprobante.id,
                nomina_empleado_id=ne.id,
                empleado_id=empleado.id,
                empleado_codigo=empleado.codigo_empleado,
                empleado_nombre=f"{empleado.primer_nombre} {empleado.primer_apellido}",
                codigo_cuenta="5101",
                descripcion_cuenta="Gasto por Salario",
                tipo_debito_credito="debito",
                debito=Decimal("1000.00"),
                credito=Decimal("0.00"),
                monto_calculado=Decimal("1000.00"),
                concepto="Salario Base",
                tipo_concepto="salario_base",
                concepto_codigo="SALARIO_BASE",
                orden=1,
            )
            db_session.add(linea)
            db_session.commit()

            planilla, nomina = _prepare_objects_for_export(planilla, nomina)

            output, _filename = ExportService.exportar_comprobante_detallado_excel(planilla, nomina)

            wb = load_workbook(output)
            ws = wb.active
            values = {ws[f"A{row}"].value: ws[f"B{row}"].value for row in range(1, ws.max_row + 1)}

            assert values["ID Planilla:"] == planilla.id
            assert values["Estatus Planilla:"] == nomina.estado
            assert values["Creado por:"] == nomina.generado_por
            assert values["Aprobado por:"] == "approver_user"
            assert values["Aplicado por:"] == "applier_user"

    def test_exportar_comprobante_detallado_excel_with_empresa(
        self, app, db_session, planilla, nomina, moneda, empleado
    ):
        """
        Test that export includes empresa information when available.

        Setup:
            - Create planilla with empresa, comprobante with detailed lines

        Action:
            - Call exportar_comprobante_detallado_excel

        Verification:
            - Export succeeds with empresa info
        """
        from coati_payroll.model import ComprobanteContable, ComprobanteContableLinea, NominaEmpleado
        from coati_payroll.vistas.planilla.services.export_service import ExportService

        with app.app_context():
            # Create nomina empleado
            ne = NominaEmpleado(
                nomina_id=nomina.id,
                empleado_id=empleado.id,
                salario_bruto=Decimal("1000.00"),
                total_ingresos=Decimal("1000.00"),
                total_deducciones=Decimal("0.00"),
                salario_neto=Decimal("1000.00"),
                sueldo_base_historico=Decimal("1000.00"),
            )
            db_session.add(ne)
            db_session.flush()

            # Create comprobante
            comprobante = ComprobanteContable(
                nomina_id=nomina.id,
                fecha_calculo=date(2025, 1, 31),
                concepto="Nómina Enero 2025",
                moneda_id=moneda.id,
                total_debitos=Decimal("500.00"),
                total_creditos=Decimal("500.00"),
                balance=Decimal("0.00"),
            )
            db_session.add(comprobante)
            db_session.flush()

            # Create comprobante lines
            linea = ComprobanteContableLinea(
                comprobante_id=comprobante.id,
                nomina_empleado_id=ne.id,
                empleado_id=empleado.id,
                empleado_codigo=empleado.codigo_empleado,
                empleado_nombre=f"{empleado.primer_nombre} {empleado.primer_apellido}",
                codigo_cuenta="5101",
                descripcion_cuenta="Gasto por Salario",
                tipo_debito_credito="debito",
                debito=Decimal("500.00"),
                credito=Decimal("0.00"),
                monto_calculado=Decimal("500.00"),
                concepto="Salario Base",
                tipo_concepto="salario_base",
                concepto_codigo="SALARIO_BASE",
                orden=1,
            )
            db_session.add(linea)
            db_session.commit()

            planilla, nomina = _prepare_objects_for_export(planilla, nomina)

            output, filename = ExportService.exportar_comprobante_detallado_excel(planilla, nomina)

            assert isinstance(output, BytesIO)
            assert filename is not None

    def test_exportar_comprobante_detallado_excel_without_empresa(
        self, app, db_session, planilla_sin_empresa, nomina, moneda, empleado
    ):
        """
        Test that export works without empresa.

        Setup:
            - Create planilla without empresa, comprobante with detailed lines

        Action:
            - Call exportar_comprobante_detallado_excel

        Verification:
            - Export succeeds even without empresa
        """
        from coati_payroll.model import ComprobanteContable, ComprobanteContableLinea, NominaEmpleado
        from coati_payroll.vistas.planilla.services.export_service import ExportService

        with app.app_context():
            # Create nomina empleado
            ne = NominaEmpleado(
                nomina_id=nomina.id,
                empleado_id=empleado.id,
                salario_bruto=Decimal("800.00"),
                total_ingresos=Decimal("800.00"),
                total_deducciones=Decimal("0.00"),
                salario_neto=Decimal("800.00"),
                sueldo_base_historico=Decimal("800.00"),
            )
            db_session.add(ne)
            db_session.flush()

            # Create comprobante
            comprobante = ComprobanteContable(
                nomina_id=nomina.id,
                fecha_calculo=date(2025, 1, 31),
                concepto="Nómina Enero 2025",
                moneda_id=moneda.id,
                total_debitos=Decimal("800.00"),
                total_creditos=Decimal("800.00"),
                balance=Decimal("0.00"),
            )
            db_session.add(comprobante)
            db_session.flush()

            # Create comprobante line
            linea = ComprobanteContableLinea(
                comprobante_id=comprobante.id,
                nomina_empleado_id=ne.id,
                empleado_id=empleado.id,
                empleado_codigo=empleado.codigo_empleado,
                empleado_nombre=f"{empleado.primer_nombre} {empleado.primer_apellido}",
                codigo_cuenta="5101",
                descripcion_cuenta="Gasto por Salario",
                tipo_debito_credito="debito",
                debito=Decimal("800.00"),
                credito=Decimal("0.00"),
                monto_calculado=Decimal("800.00"),
                concepto="Salario Base",
                tipo_concepto="salario_base",
                concepto_codigo="SALARIO_BASE",
                orden=1,
            )
            db_session.add(linea)
            db_session.commit()

            planilla_sin_empresa, nomina = _prepare_objects_for_export(planilla_sin_empresa, nomina)

            output, filename = ExportService.exportar_comprobante_detallado_excel(planilla_sin_empresa, nomina)

            assert isinstance(output, BytesIO)
            assert filename is not None

    def test_exportar_comprobante_detallado_excel_multiple_employees(
        self, app, db_session, planilla, nomina, moneda, empleado, empleado_minimo
    ):
        """
        Test that export handles multiple employees correctly.

        Setup:
            - Create planilla, nomina with multiple employees
            - Create comprobante with lines for multiple employees

        Action:
            - Call exportar_comprobante_detallado_excel

        Verification:
            - Export succeeds with multiple employees
        """
        from coati_payroll.model import ComprobanteContable, ComprobanteContableLinea, NominaEmpleado
        from coati_payroll.vistas.planilla.services.export_service import ExportService

        with app.app_context():
            # Create nomina empleados
            ne1 = NominaEmpleado(
                nomina_id=nomina.id,
                empleado_id=empleado.id,
                salario_bruto=Decimal("1000.00"),
                total_ingresos=Decimal("1000.00"),
                total_deducciones=Decimal("0.00"),
                salario_neto=Decimal("1000.00"),
                sueldo_base_historico=Decimal("1000.00"),
            )
            ne2 = NominaEmpleado(
                nomina_id=nomina.id,
                empleado_id=empleado_minimo.id,
                salario_bruto=Decimal("800.00"),
                total_ingresos=Decimal("800.00"),
                total_deducciones=Decimal("0.00"),
                salario_neto=Decimal("800.00"),
                sueldo_base_historico=Decimal("800.00"),
            )
            db_session.add(ne1)
            db_session.add(ne2)
            db_session.flush()

            # Create comprobante
            comprobante = ComprobanteContable(
                nomina_id=nomina.id,
                fecha_calculo=date(2025, 1, 31),
                concepto="Nómina Enero 2025",
                moneda_id=moneda.id,
                total_debitos=Decimal("1800.00"),
                total_creditos=Decimal("1800.00"),
                balance=Decimal("0.00"),
            )
            db_session.add(comprobante)
            db_session.flush()

            # Create comprobante lines for first employee
            linea1 = ComprobanteContableLinea(
                comprobante_id=comprobante.id,
                nomina_empleado_id=ne1.id,
                empleado_id=empleado.id,
                empleado_codigo=empleado.codigo_empleado,
                empleado_nombre=f"{empleado.primer_nombre} {empleado.primer_apellido}",
                codigo_cuenta="5101",
                descripcion_cuenta="Gasto por Salario",
                tipo_debito_credito="debito",
                debito=Decimal("1000.00"),
                credito=Decimal("0.00"),
                monto_calculado=Decimal("1000.00"),
                concepto="Salario Base",
                tipo_concepto="salario_base",
                concepto_codigo="SALARIO_BASE",
                orden=1,
            )
            linea2 = ComprobanteContableLinea(
                comprobante_id=comprobante.id,
                nomina_empleado_id=ne1.id,
                empleado_id=empleado.id,
                empleado_codigo=empleado.codigo_empleado,
                empleado_nombre=f"{empleado.primer_nombre} {empleado.primer_apellido}",
                codigo_cuenta="2101",
                descripcion_cuenta="Salario por Pagar",
                tipo_debito_credito="credito",
                debito=Decimal("0.00"),
                credito=Decimal("1000.00"),
                monto_calculado=Decimal("1000.00"),
                concepto="Salario Base",
                tipo_concepto="salario_base",
                concepto_codigo="SALARIO_BASE",
                orden=2,
            )
            # Create comprobante lines for second employee
            linea3 = ComprobanteContableLinea(
                comprobante_id=comprobante.id,
                nomina_empleado_id=ne2.id,
                empleado_id=empleado_minimo.id,
                empleado_codigo=empleado_minimo.codigo_empleado,
                empleado_nombre=f"{empleado_minimo.primer_nombre} {empleado_minimo.primer_apellido}",
                codigo_cuenta="5101",
                descripcion_cuenta="Gasto por Salario",
                tipo_debito_credito="debito",
                debito=Decimal("800.00"),
                credito=Decimal("0.00"),
                monto_calculado=Decimal("800.00"),
                concepto="Salario Base",
                tipo_concepto="salario_base",
                concepto_codigo="SALARIO_BASE",
                orden=3,
            )
            linea4 = ComprobanteContableLinea(
                comprobante_id=comprobante.id,
                nomina_empleado_id=ne2.id,
                empleado_id=empleado_minimo.id,
                empleado_codigo=empleado_minimo.codigo_empleado,
                empleado_nombre=f"{empleado_minimo.primer_nombre} {empleado_minimo.primer_apellido}",
                codigo_cuenta="2101",
                descripcion_cuenta="Salario por Pagar",
                tipo_debito_credito="credito",
                debito=Decimal("0.00"),
                credito=Decimal("800.00"),
                monto_calculado=Decimal("800.00"),
                concepto="Salario Base",
                tipo_concepto="salario_base",
                concepto_codigo="SALARIO_BASE",
                orden=4,
            )
            db_session.add(linea1)
            db_session.add(linea2)
            db_session.add(linea3)
            db_session.add(linea4)
            db_session.commit()

            planilla, nomina = _prepare_objects_for_export(planilla, nomina)

            output, filename = ExportService.exportar_comprobante_detallado_excel(planilla, nomina)

            assert isinstance(output, BytesIO)
            assert filename is not None

    def test_exportar_comprobante_detallado_excel_with_null_accounts(
        self, app, db_session, planilla, nomina, moneda, empleado
    ):
        """
        Test that detailed export works even with NULL account codes (for audit).

        Setup:
            - Create comprobante with NULL codigo_cuenta (incomplete config)

        Action:
            - Call exportar_comprobante_detallado_excel

        Verification:
            - Export succeeds even with incomplete accounting configuration
            - NULL values are handled properly
        """
        from coati_payroll.model import ComprobanteContable, ComprobanteContableLinea, NominaEmpleado
        from coati_payroll.vistas.planilla.services.export_service import ExportService

        with app.app_context():
            # Create nomina empleado
            ne = NominaEmpleado(
                nomina_id=nomina.id,
                empleado_id=empleado.id,
                salario_bruto=Decimal("1000.00"),
                total_ingresos=Decimal("1000.00"),
                total_deducciones=Decimal("0.00"),
                salario_neto=Decimal("1000.00"),
                sueldo_base_historico=Decimal("1000.00"),
            )
            db_session.add(ne)
            db_session.flush()

            # Create comprobante
            comprobante = ComprobanteContable(
                nomina_id=nomina.id,
                fecha_calculo=date(2025, 1, 31),
                concepto="Nómina Enero 2025",
                moneda_id=moneda.id,
                total_debitos=Decimal("1000.00"),
                total_creditos=Decimal("1000.00"),
                balance=Decimal("0.00"),
            )
            db_session.add(comprobante)
            db_session.flush()

            # Create comprobante line with NULL codigo_cuenta (for audit)
            linea = ComprobanteContableLinea(
                comprobante_id=comprobante.id,
                nomina_empleado_id=ne.id,
                empleado_id=empleado.id,
                empleado_codigo=empleado.codigo_empleado,
                empleado_nombre=f"{empleado.primer_nombre} {empleado.primer_apellido}",
                codigo_cuenta=None,  # NULL account - incomplete config
                descripcion_cuenta=None,
                tipo_debito_credito="debito",
                debito=Decimal("1000.00"),
                credito=Decimal("0.00"),
                monto_calculado=Decimal("1000.00"),
                concepto="Salario Base",
                tipo_concepto="salario_base",
                concepto_codigo="SALARIO_BASE",
                orden=1,
            )
            db_session.add(linea)
            db_session.commit()

            planilla, nomina = _prepare_objects_for_export(planilla, nomina)

            output, filename = ExportService.exportar_comprobante_detallado_excel(planilla, nomina)

            assert isinstance(output, BytesIO)
            assert filename is not None
            # Verify file is not empty
            output.seek(0)
            content = output.read()
            assert len(content) > 0

    def test_exportar_comprobante_detallado_excel_missing_openpyxl(
        self, app, db_session, planilla, nomina, monkeypatch
    ):
        """
        Test that export raises ImportError when openpyxl is not available.

        Setup:
            - Mock check_openpyxl_available to return None

        Action:
            - Call exportar_comprobante_detallado_excel

        Verification:
            - Raises ImportError with appropriate message
        """
        from coati_payroll.vistas.planilla.services import export_service

        with app.app_context():

            def mock_check_openpyxl():
                return None

            # Patch the function in the module where it's imported and used
            monkeypatch.setattr(export_service, "check_openpyxl_available", mock_check_openpyxl)

            planilla, nomina = _prepare_objects_for_export(planilla, nomina)

            from coati_payroll.vistas.planilla.services.export_service import ExportService

            with pytest.raises(ImportError, match="openpyxl no está disponible"):
                ExportService.exportar_comprobante_detallado_excel(planilla, nomina)
