# Copyright 2025 BMO Soluciones, S.A.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#    http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
"""Tests for report export functionality."""

import os
from pathlib import Path
import pytest

from coati_payroll.report_export import (
    ReportExporter,
    export_report_to_csv,
    export_report_to_excel,
    OPENPYXL_AVAILABLE,
)


def test_report_exporter_initialization():
    """
    Test ReportExporter initialization.

    Setup:
        - None

    Action:
        - Create ReportExporter instance

    Verification:
        - Instance created with correct attributes
    """
    results = [
        {"Name": "John", "Age": 30},
        {"Name": "Jane", "Age": 25},
    ]

    exporter = ReportExporter("Test Report", results)

    assert exporter.report_name == "Test Report"
    assert len(exporter.results) == 2


def test_export_to_csv(tmpdir):
    """
    Test CSV export functionality.

    Setup:
        - Create test results

    Action:
        - Export to CSV

    Verification:
        - File is created
        - Contains correct data
    """
    results = [
        {"Name": "John", "Age": 30, "City": "NYC"},
        {"Name": "Jane", "Age": 25, "City": "LA"},
    ]

    output_path = str(tmpdir.join("test_report.csv"))

    exporter = ReportExporter("Test Report", results)
    file_path = exporter.to_csv(output_path)

    assert os.path.exists(file_path)

    # Read and verify contents
    with open(file_path, "r") as f:
        content = f.read()
        assert "Name" in content
        assert "John" in content
        assert "Jane" in content


def test_export_to_csv_with_empty_results(tmpdir):
    """
    Test CSV export with empty results.

    Setup:
        - Empty results list

    Action:
        - Export to CSV

    Verification:
        - File path is returned (even if empty)
    """
    results = []

    output_path = str(tmpdir.join("empty_report.csv"))

    exporter = ReportExporter("Empty Report", results)
    file_path = exporter.to_csv(output_path)

    # Empty results may not create file in some implementations
    # Just verify path is returned
    assert file_path == output_path


def test_export_report_to_csv_convenience_function(tmpdir):
    """
    Test export_report_to_csv convenience function.

    Setup:
        - Create test results

    Action:
        - Use convenience function

    Verification:
        - File is created
    """
    results = [
        {"Product": "Widget", "Price": 10.99},
    ]

    output_path = str(tmpdir.join("convenience_test.csv"))

    file_path = export_report_to_csv("Convenience Test", results, output_path)

    assert os.path.exists(file_path)


def test_csv_export_handles_special_characters(tmpdir):
    """
    Test CSV export handles special characters.

    Setup:
        - Results with special characters

    Action:
        - Export to CSV

    Verification:
        - File contains special characters correctly
    """
    results = [
        {"Name": "José García", "Description": "Test, with comma"},
    ]

    output_path = str(tmpdir.join("special_chars.csv"))

    exporter = ReportExporter("Special Chars", results)
    file_path = exporter.to_csv(output_path)

    assert os.path.exists(file_path)

    with open(file_path, "r", encoding="utf-8") as f:
        content = f.read()
        assert "José García" in content


def test_csv_export_generates_filename():
    """
    Test CSV export generates filename when not provided.

    Setup:
        - Results data

    Action:
        - Export without specifying path

    Verification:
        - File is created with auto-generated name
    """
    results = [{"Col1": "Value1"}]

    exporter = ReportExporter("Auto Name Test", results)
    file_path = exporter.to_csv()

    try:
        assert os.path.exists(file_path)
        assert "Auto_Name_Test" in file_path or "Auto Name Test" in file_path
        assert file_path.endswith(".csv")
    finally:
        # Cleanup
        if os.path.exists(file_path):
            os.remove(file_path)
            # Try to remove parent directories if empty
            try:
                parent_dir = Path(file_path).parent
                parent_dir.rmdir()
                parent_dir.parent.rmdir()
            except Exception:
                pass


def test_report_exporter_with_numeric_data(tmpdir):
    """
    Test exporter with numeric data types.

    Setup:
        - Results with integers and floats

    Action:
        - Export to CSV

    Verification:
        - Numeric values are exported correctly
    """
    results = [
        {"ID": 1, "Value": 123.45, "Count": 10},
        {"ID": 2, "Value": 678.90, "Count": 20},
    ]

    output_path = str(tmpdir.join("numeric_data.csv"))

    exporter = ReportExporter("Numeric Test", results)
    file_path = exporter.to_csv(output_path)

    with open(file_path, "r") as f:
        content = f.read()
        assert "123.45" in content
        # Python may write 678.9 instead of 678.90
        assert "678.9" in content


def test_report_exporter_with_none_values(tmpdir):
    """
    Test exporter handles None values.

    Setup:
        - Results with None values

    Action:
        - Export to CSV

    Verification:
        - None values are handled gracefully
    """
    results = [
        {"Name": "John", "Email": None},
        {"Name": "Jane", "Email": "jane@example.com"},
    ]

    output_path = str(tmpdir.join("with_none.csv"))

    exporter = ReportExporter("None Test", results)
    file_path = exporter.to_csv(output_path)

    assert os.path.exists(file_path)


def test_safe_filename_generation():
    """
    Test that unsafe characters in report name are handled.

    Setup:
        - Report name with special characters

    Action:
        - Create exporter

    Verification:
        - Exporter handles name correctly
    """
    unsafe_name = "Report/With\\Special:Chars*"
    results = [{"A": 1}]

    exporter = ReportExporter(unsafe_name, results)

    # Should not raise error
    assert exporter.report_name == unsafe_name


def test_csv_export_with_boolean_values(tmpdir):
    """
    Test CSV export with boolean values.

    Setup:
        - Results with boolean values

    Action:
        - Export to CSV

    Verification:
        - Booleans are exported
    """
    results = [
        {"Name": "Item1", "Active": True},
        {"Name": "Item2", "Active": False},
    ]

    output_path = str(tmpdir.join("boolean_test.csv"))

    exporter = ReportExporter("Boolean Test", results)
    file_path = exporter.to_csv(output_path)

    with open(file_path, "r") as f:
        content = f.read()
        assert "True" in content or "true" in content.lower()
        assert "False" in content or "false" in content.lower()


def test_export_creates_directory_structure():
    """
    Test that export creates necessary directory structure.

    Setup:
        - None

    Action:
        - Export without providing path (auto-generate)

    Verification:
        - Directory structure is created
    """
    results = [{"Test": "Data"}]

    exporter = ReportExporter("Dir Test", results)
    file_path = exporter.to_csv()

    try:
        # Should create exports/reports/ directory
        assert os.path.exists(file_path)
        parent_dir = Path(file_path).parent
        assert parent_dir.name == "reports"
    finally:
        # Cleanup
        if os.path.exists(file_path):
            os.remove(file_path)
            try:
                parent_dir = Path(file_path).parent
                parent_dir.rmdir()
                parent_dir.parent.rmdir()
            except Exception:
                pass


def test_multiple_exports_different_names():
    """
    Test multiple exports with different names don't conflict.

    Setup:
        - Same results, different report names

    Action:
        - Export twice

    Verification:
        - Two different files created
    """
    results = [{"Data": "Value"}]

    exporter1 = ReportExporter("Report A", results)
    exporter2 = ReportExporter("Report B", results)

    file1 = exporter1.to_csv()
    file2 = exporter2.to_csv()

    try:
        assert os.path.exists(file1)
        assert os.path.exists(file2)
        assert file1 != file2
    finally:
        # Cleanup
        for f in [file1, file2]:
            if os.path.exists(f):
                os.remove(f)
        try:
            parent_dir = Path(file1).parent
            parent_dir.rmdir()
            parent_dir.parent.rmdir()
        except Exception:
            pass


@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
def test_export_to_excel(tmpdir):
    """
    Test Excel export functionality.

    Setup:
        - Create test results

    Action:
        - Export to Excel

    Verification:
        - File is created
        - Contains correct data
    """
    results = [
        {"Name": "John", "Age": 30, "City": "NYC"},
        {"Name": "Jane", "Age": 25, "City": "LA"},
    ]

    output_path = str(tmpdir.join("test_report.xlsx"))

    exporter = ReportExporter("Test Report", results)
    file_path = exporter.to_excel(output_path)

    assert os.path.exists(file_path)
    assert file_path.endswith(".xlsx")

    # Verify it's a valid Excel file
    if OPENPYXL_AVAILABLE:
        from openpyxl import load_workbook

        wb = load_workbook(file_path)
        ws = wb.active
        assert ws["B1"].value == "Test Report"
        assert ws["B3"].value == 2  # Total records


@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
def test_export_to_excel_with_empty_results(tmpdir):
    """
    Test Excel export with empty results.

    Setup:
        - Empty results list

    Action:
        - Export to Excel

    Verification:
        - File is created without errors
    """
    results = []

    output_path = str(tmpdir.join("empty_report.xlsx"))

    exporter = ReportExporter("Empty Report", results)
    file_path = exporter.to_excel(output_path)

    assert os.path.exists(file_path)


@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
def test_export_report_to_excel_convenience_function(tmpdir):
    """
    Test export_report_to_excel convenience function.

    Setup:
        - Create test results

    Action:
        - Use convenience function

    Verification:
        - File is created
    """
    results = [
        {"Product": "Widget", "Price": 10.99},
    ]

    output_path = str(tmpdir.join("convenience_test.xlsx"))

    file_path = export_report_to_excel("Convenience Test", results, output_path)

    assert os.path.exists(file_path)


@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
def test_excel_export_handles_special_characters(tmpdir):
    """
    Test Excel export handles special characters.

    Setup:
        - Results with special characters

    Action:
        - Export to Excel

    Verification:
        - File contains special characters correctly
    """
    results = [
        {"Name": "José García", "Description": "Test, with comma"},
    ]

    output_path = str(tmpdir.join("special_chars.xlsx"))

    exporter = ReportExporter("Special Chars", results)
    file_path = exporter.to_excel(output_path)

    assert os.path.exists(file_path)

    if OPENPYXL_AVAILABLE:
        from openpyxl import load_workbook

        wb = load_workbook(file_path)
        ws = wb.active
        # Find the row with data (after headers)
        found = False
        for row in ws.iter_rows(min_row=6, values_only=True):
            if row[0] == "José García":
                found = True
                break
        assert found


@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
def test_excel_export_generates_filename():
    """
    Test Excel export generates filename when not provided.

    Setup:
        - Results data

    Action:
        - Export without specifying path

    Verification:
        - File is created with auto-generated name
    """
    results = [{"Col1": "Value1"}]

    exporter = ReportExporter("Auto Name Test", results)
    file_path = exporter.to_excel()

    try:
        assert os.path.exists(file_path)
        assert "Auto_Name_Test" in file_path or "Auto Name Test" in file_path
        assert file_path.endswith(".xlsx")
    finally:
        # Cleanup
        if os.path.exists(file_path):
            os.remove(file_path)
            try:
                parent_dir = Path(file_path).parent
                parent_dir.rmdir()
                parent_dir.parent.rmdir()
            except Exception:
                pass


@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
def test_excel_export_with_numeric_data(tmpdir):
    """
    Test Excel exporter with numeric data types.

    Setup:
        - Results with integers and floats

    Action:
        - Export to Excel

    Verification:
        - Numeric values are exported correctly
    """
    results = [
        {"ID": 1, "Value": 123.45, "Count": 10},
        {"ID": 2, "Value": 678.90, "Count": 20},
    ]

    output_path = str(tmpdir.join("numeric_data.xlsx"))

    exporter = ReportExporter("Numeric Test", results)
    file_path = exporter.to_excel(output_path)

    assert os.path.exists(file_path)

    if OPENPYXL_AVAILABLE:
        from openpyxl import load_workbook

        wb = load_workbook(file_path)
        ws = wb.active
        # Verify numeric data is present
        found_value = False
        for row in ws.iter_rows(min_row=6, values_only=True):
            if row[1] == 123.45:
                found_value = True
                break
        assert found_value


@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
def test_excel_export_with_none_values(tmpdir):
    """
    Test Excel exporter handles None values.

    Setup:
        - Results with None values

    Action:
        - Export to Excel

    Verification:
        - None values are handled gracefully
    """
    results = [
        {"Name": "John", "Email": None},
        {"Name": "Jane", "Email": "jane@example.com"},
    ]

    output_path = str(tmpdir.join("with_none.xlsx"))

    exporter = ReportExporter("None Test", results)
    file_path = exporter.to_excel(output_path)

    assert os.path.exists(file_path)


@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
def test_excel_export_sheet_name_truncation(tmpdir):
    """
    Test that Excel sheet names are truncated to 31 characters.

    Setup:
        - Report name longer than 31 characters

    Action:
        - Export to Excel

    Verification:
        - Sheet name is truncated properly
    """
    long_name = "This is a very long report name that exceeds the maximum limit"
    results = [{"Data": "Value"}]

    output_path = str(tmpdir.join("long_name.xlsx"))

    exporter = ReportExporter(long_name, results)
    file_path = exporter.to_excel(output_path)

    if OPENPYXL_AVAILABLE:
        from openpyxl import load_workbook

        wb = load_workbook(file_path)
        ws = wb.active
        # Excel sheet names are limited to 31 characters
        assert len(ws.title) <= 31


@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
def test_excel_export_with_boolean_values(tmpdir):
    """
    Test Excel export with boolean values.

    Setup:
        - Results with boolean values

    Action:
        - Export to Excel

    Verification:
        - Booleans are exported
    """
    results = [
        {"Name": "Item1", "Active": True},
        {"Name": "Item2", "Active": False},
    ]

    output_path = str(tmpdir.join("boolean_test.xlsx"))

    exporter = ReportExporter("Boolean Test", results)
    file_path = exporter.to_excel(output_path)

    assert os.path.exists(file_path)


@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
def test_excel_column_width_adjustment(tmpdir):
    """
    Test that Excel columns are auto-adjusted for content width.

    Setup:
        - Results with varying content lengths

    Action:
        - Export to Excel

    Verification:
        - Column widths are adjusted
    """
    results = [
        {
            "Short": "A",
            "MediumLength": "Medium content here",
            "VeryLongContent": "This is a very long content that should adjust the column width",
        },
    ]

    output_path = str(tmpdir.join("width_test.xlsx"))

    exporter = ReportExporter("Width Test", results)
    file_path = exporter.to_excel(output_path)

    if OPENPYXL_AVAILABLE:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter

        wb = load_workbook(file_path)
        ws = wb.active

        # Check that columns have different widths
        col_a_width = ws.column_dimensions[get_column_letter(1)].width
        col_c_width = ws.column_dimensions[get_column_letter(3)].width

        # Column C should be wider (but capped at 50)
        assert col_c_width > col_a_width
        assert col_c_width <= 50  # Max width cap


def test_excel_export_without_openpyxl():
    """
    Test that Excel export raises ImportError when openpyxl is not available.

    Setup:
        - Mock OPENPYXL_AVAILABLE as False

    Action:
        - Attempt to export to Excel

    Verification:
        - ImportError is raised
    """
    # This test is tricky because we can't easily mock the module-level constant
    # Instead, we'll test the behavior when openpyxl import fails
    # Skip this test if openpyxl is actually installed
    if OPENPYXL_AVAILABLE:
        pytest.skip("openpyxl is installed, cannot test ImportError path")

    results = [{"Test": "Data"}]
    exporter = ReportExporter("Test", results)

    with pytest.raises(ImportError, match="openpyxl is required"):
        exporter.to_excel()


@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
def test_excel_export_creates_directory_structure():
    """
    Test that Excel export creates necessary directory structure.

    Setup:
        - None

    Action:
        - Export without providing path (auto-generate)

    Verification:
        - Directory structure is created
    """
    results = [{"Test": "Data"}]

    exporter = ReportExporter("Dir Test Excel", results)
    file_path = exporter.to_excel()

    try:
        # Should create exports/reports/ directory
        assert os.path.exists(file_path)
        parent_dir = Path(file_path).parent
        assert parent_dir.name == "reports"
    finally:
        # Cleanup
        if os.path.exists(file_path):
            os.remove(file_path)
            try:
                parent_dir = Path(file_path).parent
                parent_dir.rmdir()
                parent_dir.parent.rmdir()
            except Exception:
                pass


@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
def test_excel_metadata_formatting(tmpdir):
    """
    Test that Excel metadata is properly formatted.

    Setup:
        - Create test results

    Action:
        - Export to Excel

    Verification:
        - Metadata cells are bold
        - Headers have proper styling
    """
    results = [{"Name": "Test", "Value": 123}]

    output_path = str(tmpdir.join("metadata_test.xlsx"))

    exporter = ReportExporter("Metadata Test", results)
    file_path = exporter.to_excel(output_path)

    if OPENPYXL_AVAILABLE:
        from openpyxl import load_workbook

        wb = load_workbook(file_path)
        ws = wb.active

        # Check metadata is bold
        assert ws["A1"].font.bold is True
        assert ws["A2"].font.bold is True
        assert ws["A3"].font.bold is True

        # Check headers have styling (row 5)
        header_cell = ws.cell(row=5, column=1)
        assert header_cell.font.bold is True
        assert header_cell.fill.start_color.rgb is not None
