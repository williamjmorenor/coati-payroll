# Copyright 2025 BMO Soluciones, S.A.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#    http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
"""Initial benefit balance loading views."""

from __future__ import annotations

from decimal import Decimal
from datetime import datetime

from flask import Blueprint, flash, redirect, render_template, request, url_for, Response
from flask_login import current_user
from sqlalchemy import and_

from coati_payroll.forms import CargaInicialPrestacionForm
from coati_payroll.i18n import _
from coati_payroll.model import (
    CargaInicialPrestacion,
    Empleado,
    Moneda,
    Prestacion,
    PrestacionAcumulada,
    db,
)
from coati_payroll.rbac import require_read_access, require_write_access
from coati_payroll.vistas.constants import PER_PAGE

carga_inicial_prestacion_bp = Blueprint("carga_inicial_prestacion", __name__, url_prefix="/carga-inicial-prestaciones")


@carga_inicial_prestacion_bp.route("/")
@require_read_access()
def index():
    """List all initial benefit balance loads."""
    page = request.args.get("page", 1, type=int)
    estado_filter = request.args.get("estado", "")

    query = CargaInicialPrestacion.query

    # Apply filters
    if estado_filter:
        query = query.filter(CargaInicialPrestacion.estado == estado_filter)

    # Order by creation date descending
    query = query.order_by(CargaInicialPrestacion.creado.desc())

    # Paginate
    pagination = query.paginate(page=page, per_page=PER_PAGE, error_out=False)
    cargas = pagination.items

    return render_template(
        "modules/carga_inicial_prestacion/index.html",
        cargas=cargas,
        pagination=pagination,
        estado_filter=estado_filter,
    )


@carga_inicial_prestacion_bp.route("/nueva", methods=["GET", "POST"])
@require_write_access()
def nueva():
    """Create a new initial benefit balance load."""
    form = CargaInicialPrestacionForm()

    # Populate select field choices
    form.empleado_id.choices = [("", _("-- Seleccionar --"))] + [
        (emp.id, f"{emp.codigo_empleado} - {emp.primer_nombre} {emp.primer_apellido}")
        for emp in Empleado.query.filter_by(activo=True).order_by(Empleado.codigo_empleado).all()
    ]

    form.prestacion_id.choices = [("", _("-- Seleccionar --"))] + [
        (prest.id, f"{prest.codigo} - {prest.nombre}")
        for prest in Prestacion.query.filter_by(activo=True).order_by(Prestacion.codigo).all()
    ]

    form.moneda_id.choices = [("", _("-- Seleccionar --"))] + [
        (mon.id, f"{mon.codigo} - {mon.nombre}")
        for mon in Moneda.query.filter_by(activo=True).order_by(Moneda.codigo).all()
    ]

    if form.validate_on_submit():
        # Check for duplicate
        existing = CargaInicialPrestacion.query.filter(
            and_(
                CargaInicialPrestacion.empleado_id == form.empleado_id.data,
                CargaInicialPrestacion.prestacion_id == form.prestacion_id.data,
                CargaInicialPrestacion.anio_corte == form.anio_corte.data,
                CargaInicialPrestacion.mes_corte == form.mes_corte.data,
            )
        ).first()

        if existing:
            flash(
                _("Ya existe una carga inicial para este empleado, prestación y periodo."),
                "warning",
            )
            return render_template("modules/carga_inicial_prestacion/form.html", form=form)

        carga = CargaInicialPrestacion(
            empleado_id=form.empleado_id.data,
            prestacion_id=form.prestacion_id.data,
            anio_corte=form.anio_corte.data,
            mes_corte=form.mes_corte.data,
            moneda_id=form.moneda_id.data,
            saldo_acumulado=form.saldo_acumulado.data if form.saldo_acumulado.data is not None else Decimal("0.00"),
            tipo_cambio=form.tipo_cambio.data if form.tipo_cambio.data is not None else Decimal("1.0"),
            saldo_convertido=form.saldo_convertido.data if form.saldo_convertido.data is not None else Decimal("0.00"),
            observaciones=form.observaciones.data,
            estado="borrador",
            creado_por=current_user.usuario if current_user.is_authenticated else None,
        )

        db.session.add(carga)
        db.session.commit()

        flash(_("Carga inicial creada exitosamente en estado borrador."), "success")
        return redirect(url_for("carga_inicial_prestacion.index"))

    return render_template("modules/carga_inicial_prestacion/form.html", form=form)


@carga_inicial_prestacion_bp.route("/<carga_id>/editar", methods=["GET", "POST"])
@require_write_access()
def editar(carga_id):
    """Edit an initial benefit balance load (only if in draft status)."""
    carga = CargaInicialPrestacion.query.get_or_404(carga_id)

    if carga.estado == "aplicado":
        flash(_("No se puede editar una carga inicial ya aplicada."), "warning")
        return redirect(url_for("carga_inicial_prestacion.index"))

    form = CargaInicialPrestacionForm(obj=carga)

    # Populate select field choices
    form.empleado_id.choices = [("", _("-- Seleccionar --"))] + [
        (emp.id, f"{emp.codigo_empleado} - {emp.primer_nombre} {emp.primer_apellido}")
        for emp in Empleado.query.filter_by(activo=True).order_by(Empleado.codigo_empleado).all()
    ]

    form.prestacion_id.choices = [("", _("-- Seleccionar --"))] + [
        (prest.id, f"{prest.codigo} - {prest.nombre}")
        for prest in Prestacion.query.filter_by(activo=True).order_by(Prestacion.codigo).all()
    ]

    form.moneda_id.choices = [("", _("-- Seleccionar --"))] + [
        (mon.id, f"{mon.codigo} - {mon.nombre}")
        for mon in Moneda.query.filter_by(activo=True).order_by(Moneda.codigo).all()
    ]

    if form.validate_on_submit():
        carga.empleado_id = form.empleado_id.data
        carga.prestacion_id = form.prestacion_id.data
        carga.anio_corte = form.anio_corte.data
        carga.mes_corte = form.mes_corte.data
        carga.moneda_id = form.moneda_id.data
        carga.saldo_acumulado = form.saldo_acumulado.data if form.saldo_acumulado.data is not None else Decimal("0.00")
        carga.tipo_cambio = form.tipo_cambio.data if form.tipo_cambio.data is not None else Decimal("1.0")
        carga.saldo_convertido = (
            form.saldo_convertido.data if form.saldo_convertido.data is not None else Decimal("0.00")
        )
        carga.observaciones = form.observaciones.data
        carga.modificado_por = current_user.usuario if current_user.is_authenticated else None

        db.session.commit()

        flash(_("Carga inicial actualizada exitosamente."), "success")
        return redirect(url_for("carga_inicial_prestacion.index"))

    return render_template("modules/carga_inicial_prestacion/form.html", form=form, carga=carga)


@carga_inicial_prestacion_bp.route("/<carga_id>/aplicar", methods=["POST"])
@require_write_access()
def aplicar(carga_id):
    """Apply an initial balance load - creates transaction in prestacion_acumulada."""
    carga = CargaInicialPrestacion.query.get_or_404(carga_id)

    if carga.estado == "aplicado":
        flash(_("Esta carga inicial ya ha sido aplicada."), "warning")
        return redirect(url_for("carga_inicial_prestacion.index"))

    try:
        # Create transaction in prestacion_acumulada
        transaccion = PrestacionAcumulada(
            empleado_id=carga.empleado_id,
            prestacion_id=carga.prestacion_id,
            fecha_transaccion=datetime.now().date(),
            tipo_transaccion="saldo_inicial",
            anio=carga.anio_corte,
            mes=carga.mes_corte,
            moneda_id=carga.moneda_id,
            monto_transaccion=carga.saldo_convertido,
            saldo_anterior=Decimal("0.00"),
            saldo_nuevo=carga.saldo_convertido,
            carga_inicial_id=carga.id,
            observaciones=f"Carga inicial - {carga.observaciones or ''}",
            procesado_por=current_user.usuario if current_user.is_authenticated else None,
            creado_por=current_user.usuario if current_user.is_authenticated else None,
        )

        db.session.add(transaccion)

        # Update carga status
        carga.estado = "aplicado"
        carga.fecha_aplicacion = datetime.now()
        carga.aplicado_por = current_user.usuario if current_user.is_authenticated else None
        carga.modificado_por = current_user.usuario if current_user.is_authenticated else None

        db.session.commit()

        flash(_("Carga inicial aplicada exitosamente."), "success")

    except Exception as e:
        db.session.rollback()
        flash(_("Error al aplicar la carga inicial: %(error)s", error=str(e)), "danger")

    return redirect(url_for("carga_inicial_prestacion.index"))


@carga_inicial_prestacion_bp.route("/<carga_id>/eliminar", methods=["POST"])
@require_write_access()
def eliminar(carga_id):
    """Delete an initial balance load (only if in draft status)."""
    carga = CargaInicialPrestacion.query.get_or_404(carga_id)

    if carga.estado == "aplicado":
        flash(_("No se puede eliminar una carga inicial ya aplicada."), "warning")
        return redirect(url_for("carga_inicial_prestacion.index"))

    try:
        db.session.delete(carga)
        db.session.commit()
        flash(_("Carga inicial eliminada exitosamente."), "success")
    except Exception as e:
        db.session.rollback()
        flash(_("Error al eliminar la carga inicial: %(error)s", error=str(e)), "danger")

    return redirect(url_for("carga_inicial_prestacion.index"))


@carga_inicial_prestacion_bp.route("/reporte")
@require_read_access()
def reporte():
    """Generate accumulated benefits report."""
    # Get filter parameters
    empleado_id = request.args.get("empleado_id")
    prestacion_id = request.args.get("prestacion_id")
    fecha_desde = request.args.get("fecha_desde")
    fecha_hasta = request.args.get("fecha_hasta")

    # Build query
    query = PrestacionAcumulada.query

    if empleado_id:
        query = query.filter(PrestacionAcumulada.empleado_id == empleado_id)

    if prestacion_id:
        query = query.filter(PrestacionAcumulada.prestacion_id == prestacion_id)

    if fecha_desde:
        query = query.filter(PrestacionAcumulada.fecha_transaccion >= fecha_desde)

    if fecha_hasta:
        query = query.filter(PrestacionAcumulada.fecha_transaccion <= fecha_hasta)

    # Order by date
    transacciones = query.order_by(
        PrestacionAcumulada.empleado_id, PrestacionAcumulada.prestacion_id, PrestacionAcumulada.fecha_transaccion
    ).all()

    # Get choices for filters
    empleados = Empleado.query.filter_by(activo=True).order_by(Empleado.codigo_empleado).all()
    prestaciones = Prestacion.query.filter_by(activo=True).order_by(Prestacion.codigo).all()

    return render_template(
        "modules/carga_inicial_prestacion/reporte.html",
        transacciones=transacciones,
        empleados=empleados,
        prestaciones=prestaciones,
        empleado_id=empleado_id,
        prestacion_id=prestacion_id,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
    )


@carga_inicial_prestacion_bp.route("/reporte/excel")
@require_read_access()
def reporte_excel():
    """Export accumulated benefits report to Excel."""
    import io

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        flash(_("La librería openpyxl no está instalada. No se puede generar el reporte Excel."), "danger")
        return redirect(url_for("carga_inicial_prestacion.reporte"))

    # Get filter parameters (same as reporte)
    empleado_id = request.args.get("empleado_id")
    prestacion_id = request.args.get("prestacion_id")
    fecha_desde = request.args.get("fecha_desde")
    fecha_hasta = request.args.get("fecha_hasta")

    # Build query
    query = PrestacionAcumulada.query

    if empleado_id:
        query = query.filter(PrestacionAcumulada.empleado_id == empleado_id)

    if prestacion_id:
        query = query.filter(PrestacionAcumulada.prestacion_id == prestacion_id)

    if fecha_desde:
        query = query.filter(PrestacionAcumulada.fecha_transaccion >= fecha_desde)

    if fecha_hasta:
        query = query.filter(PrestacionAcumulada.fecha_transaccion <= fecha_hasta)

    transacciones = query.order_by(
        PrestacionAcumulada.empleado_id, PrestacionAcumulada.prestacion_id, PrestacionAcumulada.fecha_transaccion
    ).all()

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Prestaciones Acumuladas"

    # Header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Headers - Enhanced for audit purposes
    headers = [
        "ID Transacción",
        "Fecha Transacción",
        "Código Empleado",
        "Empleado",
        "Código Prestación",
        "Prestación",
        "Tipo Acumulación",
        "Tipo Transacción",
        "Año",
        "Mes",
        "Monto Transacción",
        "Saldo Anterior",
        "Saldo Nuevo",
        "Moneda",
        "Nómina ID",
        "Carga Inicial ID",
        "Procesado Por",
        "Fecha Creación",
        "Creado Por",
        "Observaciones",
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font

    # Data rows - Enhanced with all audit fields
    for row_num, trans in enumerate(transacciones, 2):
        ws.cell(row=row_num, column=1, value=trans.id)
        ws.cell(row=row_num, column=2, value=trans.fecha_transaccion.strftime("%Y-%m-%d"))
        ws.cell(row=row_num, column=3, value=trans.empleado.codigo_empleado)
        ws.cell(
            row=row_num,
            column=4,
            value=f"{trans.empleado.primer_nombre} {trans.empleado.primer_apellido}",
        )
        ws.cell(row=row_num, column=5, value=trans.prestacion.codigo)
        ws.cell(row=row_num, column=6, value=trans.prestacion.nombre)
        ws.cell(row=row_num, column=7, value=trans.prestacion.tipo_acumulacion)
        ws.cell(row=row_num, column=8, value=trans.tipo_transaccion)
        ws.cell(row=row_num, column=9, value=trans.anio)
        ws.cell(row=row_num, column=10, value=trans.mes)
        ws.cell(row=row_num, column=11, value=float(trans.monto_transaccion))
        ws.cell(row=row_num, column=12, value=float(trans.saldo_anterior))
        ws.cell(row=row_num, column=13, value=float(trans.saldo_nuevo))
        ws.cell(row=row_num, column=14, value=trans.moneda.codigo)
        ws.cell(row=row_num, column=15, value=trans.nomina_id or "")
        ws.cell(row=row_num, column=16, value=trans.carga_inicial_id or "")
        ws.cell(row=row_num, column=17, value=trans.procesado_por or "")
        ws.cell(row=row_num, column=18, value=trans.creado.strftime("%Y-%m-%d"))
        ws.cell(row=row_num, column=19, value=trans.creado_por or "")
        ws.cell(row=row_num, column=20, value=trans.observaciones or "")

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value is not None and len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except (TypeError, AttributeError):
                # Skip cells with values that can't be converted to string
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        output.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment;filename=prestaciones_acumuladas.xlsx"},
    )
