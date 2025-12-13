# Copyright 2025 BMO Soluciones, S.A.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#    http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
"""Views for managing Planilla (master payroll) and its associations.

A Planilla is the central hub that connects:
- Employees (via PlanillaEmpleado)
- Perceptions (via PlanillaIngreso)
- Deductions (via PlanillaDeduccion) - with priority ordering
- Benefits/Prestaciones (via PlanillaPrestacion)
- Calculation Rules (via PlanillaReglaCalculo)
"""

from __future__ import annotations

import re
from datetime import date, datetime, timezone

from flask import Blueprint, flash, redirect, render_template, request, url_for
from flask_login import login_required, current_user

from coati_payroll.model import (
    db,
    Planilla,
    TipoPlanilla,
    Moneda,
    Empleado,
    Percepcion,
    Deduccion,
    Prestacion,
    ReglaCalculo,
    PlanillaEmpleado,
    PlanillaIngreso,
    PlanillaDeduccion,
    PlanillaPrestacion,
    PlanillaReglaCalculo,
)
from coati_payroll.forms import PlanillaForm
from coati_payroll.i18n import _
from coati_payroll.rbac import require_read_access, require_write_access

planilla_bp = Blueprint("planilla", __name__, url_prefix="/planilla")


@planilla_bp.route("/")
@require_read_access()
def index():
    """List all planillas."""
    planillas = db.session.execute(db.select(Planilla).order_by(Planilla.nombre)).scalars().all()
    return render_template("modules/planilla/index.html", planillas=planillas)


@planilla_bp.route("/new", methods=["GET", "POST"])
@require_write_access()
def new():
    """Create a new planilla. Admin and HR can create planillas."""
    form = PlanillaForm()
    _populate_form_choices(form)

    if form.validate_on_submit():
        planilla = Planilla(
            nombre=form.nombre.data,
            descripcion=form.descripcion.data,
            tipo_planilla_id=form.tipo_planilla_id.data,
            moneda_id=form.moneda_id.data,
            empresa_id=form.empresa_id.data or None,
            periodo_fiscal_inicio=form.periodo_fiscal_inicio.data,
            periodo_fiscal_fin=form.periodo_fiscal_fin.data,
            prioridad_prestamos=form.prioridad_prestamos.data or 250,
            prioridad_adelantos=form.prioridad_adelantos.data or 251,
            aplicar_prestamos_automatico=form.aplicar_prestamos_automatico.data,
            aplicar_adelantos_automatico=form.aplicar_adelantos_automatico.data,
            codigo_cuenta_debe_salario=form.codigo_cuenta_debe_salario.data,
            descripcion_cuenta_debe_salario=form.descripcion_cuenta_debe_salario.data,
            codigo_cuenta_haber_salario=form.codigo_cuenta_haber_salario.data,
            descripcion_cuenta_haber_salario=form.descripcion_cuenta_haber_salario.data,
            activo=form.activo.data,
            creado_por=current_user.usuario,
        )
        db.session.add(planilla)
        db.session.commit()
        flash(_("Planilla creada exitosamente."), "success")
        return redirect(url_for("planilla.edit", planilla_id=planilla.id))

    return render_template("modules/planilla/form.html", form=form, is_edit=False)


@planilla_bp.route("/<planilla_id>/edit", methods=["GET", "POST"])
@require_write_access()
def edit(planilla_id: str):
    """Edit basic planilla configuration."""
    planilla = db.get_or_404(Planilla, planilla_id)
    form = PlanillaForm(obj=planilla)
    _populate_form_choices(form)

    if form.validate_on_submit():
        planilla.nombre = form.nombre.data
        planilla.descripcion = form.descripcion.data
        planilla.tipo_planilla_id = form.tipo_planilla_id.data
        planilla.moneda_id = form.moneda_id.data
        planilla.empresa_id = form.empresa_id.data or None
        planilla.periodo_fiscal_inicio = form.periodo_fiscal_inicio.data
        planilla.periodo_fiscal_fin = form.periodo_fiscal_fin.data
        planilla.prioridad_prestamos = form.prioridad_prestamos.data or 250
        planilla.prioridad_adelantos = form.prioridad_adelantos.data or 251
        planilla.aplicar_prestamos_automatico = form.aplicar_prestamos_automatico.data
        planilla.aplicar_adelantos_automatico = form.aplicar_adelantos_automatico.data
        planilla.codigo_cuenta_debe_salario = form.codigo_cuenta_debe_salario.data
        planilla.descripcion_cuenta_debe_salario = form.descripcion_cuenta_debe_salario.data
        planilla.codigo_cuenta_haber_salario = form.codigo_cuenta_haber_salario.data
        planilla.descripcion_cuenta_haber_salario = form.descripcion_cuenta_haber_salario.data
        planilla.activo = form.activo.data
        planilla.modificado_por = current_user.usuario
        db.session.commit()
        flash(_("Planilla actualizada exitosamente."), "success")
        return redirect(url_for("planilla.config", planilla_id=planilla.id))

    # Get association counts for the summary
    counts = _get_planilla_component_counts(planilla_id)

    return render_template(
        "modules/planilla/form.html",
        form=form,
        planilla=planilla,
        is_edit=True,
        **counts,
    )


@planilla_bp.route("/<planilla_id>/config")
@require_read_access()
def config(planilla_id: str):
    """Configuration overview page for a planilla."""
    planilla = db.get_or_404(Planilla, planilla_id)

    # Get association counts for the summary
    counts = _get_planilla_component_counts(planilla_id)

    return render_template(
        "modules/planilla/config.html",
        planilla=planilla,
        **counts,
    )


# ============================================================================
# CONFIGURATION PAGES FOR ASSOCIATIONS
# ============================================================================


@planilla_bp.route("/<planilla_id>/config/empleados")
@require_read_access()
def config_empleados(planilla_id: str):
    """View employees associated with a planilla."""
    planilla = db.get_or_404(Planilla, planilla_id)

    empleados_asignados = (
        db.session.execute(db.select(PlanillaEmpleado).filter_by(planilla_id=planilla_id)).scalars().all()
    )

    # Filter employees to only show those from the same company as the planilla
    query = db.select(Empleado).filter_by(activo=True)
    if planilla.empresa_id:
        # If planilla has a company, only show employees from that company or without company
        query = query.filter((Empleado.empresa_id == planilla.empresa_id) | Empleado.empresa_id.is_(None))
    empleados_disponibles = db.session.execute(query.order_by(Empleado.primer_apellido)).scalars().all()

    return render_template(
        "modules/planilla/config_empleados.html",
        planilla=planilla,
        empleados_asignados=empleados_asignados,
        empleados_disponibles=empleados_disponibles,
    )


@planilla_bp.route("/<planilla_id>/config/percepciones")
@require_read_access()
def config_percepciones(planilla_id: str):
    """View perceptions associated with a planilla."""
    planilla = db.get_or_404(Planilla, planilla_id)

    percepciones_asignadas = (
        db.session.execute(db.select(PlanillaIngreso).filter_by(planilla_id=planilla_id)).scalars().all()
    )

    percepciones_disponibles = (
        db.session.execute(db.select(Percepcion).filter_by(activo=True).order_by(Percepcion.nombre)).scalars().all()
    )

    return render_template(
        "modules/planilla/config_percepciones.html",
        planilla=planilla,
        percepciones_asignadas=percepciones_asignadas,
        percepciones_disponibles=percepciones_disponibles,
    )


@planilla_bp.route("/<planilla_id>/config/deducciones")
@require_read_access()
def config_deducciones(planilla_id: str):
    """View deductions associated with a planilla."""
    planilla = db.get_or_404(Planilla, planilla_id)

    deducciones_asignadas = (
        db.session.execute(
            db.select(PlanillaDeduccion).filter_by(planilla_id=planilla_id).order_by(PlanillaDeduccion.prioridad)
        )
        .scalars()
        .all()
    )

    deducciones_disponibles = (
        db.session.execute(db.select(Deduccion).filter_by(activo=True).order_by(Deduccion.nombre)).scalars().all()
    )

    return render_template(
        "modules/planilla/config_deducciones.html",
        planilla=planilla,
        deducciones_asignadas=deducciones_asignadas,
        deducciones_disponibles=deducciones_disponibles,
    )


@planilla_bp.route("/<planilla_id>/config/prestaciones")
@require_write_access()
def config_prestaciones(planilla_id: str):
    """Manage benefits associated with a planilla."""
    planilla = db.get_or_404(Planilla, planilla_id)

    prestaciones_asignadas = (
        db.session.execute(db.select(PlanillaPrestacion).filter_by(planilla_id=planilla_id)).scalars().all()
    )

    prestaciones_disponibles = (
        db.session.execute(db.select(Prestacion).filter_by(activo=True).order_by(Prestacion.nombre)).scalars().all()
    )

    return render_template(
        "modules/planilla/config_prestaciones.html",
        planilla=planilla,
        prestaciones_asignadas=prestaciones_asignadas,
        prestaciones_disponibles=prestaciones_disponibles,
    )


@planilla_bp.route("/<planilla_id>/config/reglas")
@require_read_access()
def config_reglas(planilla_id: str):
    """View calculation rules associated with a planilla."""
    planilla = db.get_or_404(Planilla, planilla_id)

    reglas_asignadas = (
        db.session.execute(
            db.select(PlanillaReglaCalculo).filter_by(planilla_id=planilla_id).order_by(PlanillaReglaCalculo.orden)
        )
        .scalars()
        .all()
    )

    reglas_disponibles = (
        db.session.execute(db.select(ReglaCalculo).filter_by(activo=True).order_by(ReglaCalculo.nombre)).scalars().all()
    )

    return render_template(
        "modules/planilla/config_reglas.html",
        planilla=planilla,
        reglas_asignadas=reglas_asignadas,
        reglas_disponibles=reglas_disponibles,
    )


@planilla_bp.route("/<planilla_id>/delete", methods=["POST"])
@require_write_access()
def delete(planilla_id: str):
    """Delete a planilla."""
    planilla = db.get_or_404(Planilla, planilla_id)

    # Check if planilla has nominas (payroll runs)
    if planilla.nominas:
        flash(_("No se puede eliminar una planilla con nóminas generadas."), "error")
        return redirect(url_for("planilla.index"))

    db.session.delete(planilla)
    db.session.commit()
    flash(_("Planilla eliminada exitosamente."), "success")
    return redirect(url_for("planilla.index"))


# ============================================================================
# EMPLOYEE ASSOCIATIONS
# ============================================================================


@planilla_bp.route("/<planilla_id>/empleado/add", methods=["POST"])
@require_write_access()
def add_empleado(planilla_id: str):
    """Add an employee to the planilla."""
    # Verify planilla exists (raises 404 if not found)
    planilla = db.get_or_404(Planilla, planilla_id)
    empleado_id = request.form.get("empleado_id")

    if not empleado_id:
        flash(_("Debe seleccionar un empleado."), "error")
        return redirect(url_for("planilla.config_empleados", planilla_id=planilla_id))

    # Check if already exists
    existing = db.session.execute(
        db.select(PlanillaEmpleado).filter_by(planilla_id=planilla_id, empleado_id=empleado_id)
    ).scalar_one_or_none()

    if existing:
        flash(_("El empleado ya está asignado a esta planilla."), "warning")
        return redirect(url_for("planilla.config_empleados", planilla_id=planilla_id))

    # Validate that employee and planilla belong to the same company
    empleado = db.get_or_404(Empleado, empleado_id)
    if planilla.empresa_id and empleado.empresa_id and planilla.empresa_id != empleado.empresa_id:
        flash(_("El empleado y la planilla deben pertenecer a la misma empresa."), "error")
        return redirect(url_for("planilla.config_empleados", planilla_id=planilla_id))

    association = PlanillaEmpleado(
        planilla_id=planilla_id,
        empleado_id=empleado_id,
        fecha_inicio=date.today(),
        activo=True,
        creado_por=current_user.usuario,
    )
    db.session.add(association)
    db.session.commit()
    flash(_("Empleado agregado exitosamente."), "success")
    return redirect(url_for("planilla.config_empleados", planilla_id=planilla_id))


@planilla_bp.route("/<planilla_id>/empleado/<association_id>/remove", methods=["POST"])
@require_write_access()
def remove_empleado(planilla_id: str, association_id: str):
    """Remove an employee from the planilla."""
    association = db.get_or_404(PlanillaEmpleado, association_id)
    db.session.delete(association)
    db.session.commit()
    flash(_("Empleado removido exitosamente."), "success")
    return redirect(url_for("planilla.config_empleados", planilla_id=planilla_id))


# ============================================================================
# PERCEPTION ASSOCIATIONS
# ============================================================================


@planilla_bp.route("/<planilla_id>/percepcion/add", methods=["POST"])
@require_write_access()
def add_percepcion(planilla_id: str):
    """Add a perception to the planilla."""
    # Verify planilla exists (raises 404 if not found)
    db.get_or_404(Planilla, planilla_id)
    percepcion_id = request.form.get("percepcion_id")

    if not percepcion_id:
        flash(_("Debe seleccionar una percepción."), "error")
        return redirect(url_for("planilla.config_percepciones", planilla_id=planilla_id))

    existing = db.session.execute(
        db.select(PlanillaIngreso).filter_by(planilla_id=planilla_id, percepcion_id=percepcion_id)
    ).scalar_one_or_none()

    if existing:
        flash(_("La percepción ya está asignada a esta planilla."), "warning")
        return redirect(url_for("planilla.config_percepciones", planilla_id=planilla_id))

    orden = request.form.get("orden", 0, type=int)

    association = PlanillaIngreso(
        planilla_id=planilla_id,
        percepcion_id=percepcion_id,
        orden=orden,
        editable=True,
        activo=True,
        creado_por=current_user.usuario,
    )
    db.session.add(association)
    db.session.commit()
    flash(_("Percepción agregada exitosamente."), "success")
    return redirect(url_for("planilla.config_percepciones", planilla_id=planilla_id))


@planilla_bp.route("/<planilla_id>/percepcion/<association_id>/remove", methods=["POST"])
@require_write_access()
def remove_percepcion(planilla_id: str, association_id: str):
    """Remove a perception from the planilla."""
    association = db.get_or_404(PlanillaIngreso, association_id)
    db.session.delete(association)
    db.session.commit()
    flash(_("Percepción removida exitosamente."), "success")
    return redirect(url_for("planilla.config_percepciones", planilla_id=planilla_id))


# ============================================================================
# DEDUCTION ASSOCIATIONS (with priority)
# ============================================================================


@planilla_bp.route("/<planilla_id>/deduccion/add", methods=["POST"])
@require_write_access()
def add_deduccion(planilla_id: str):
    """Add a deduction to the planilla with priority."""
    # Verify planilla exists (raises 404 if not found)
    db.get_or_404(Planilla, planilla_id)
    deduccion_id = request.form.get("deduccion_id")

    if not deduccion_id:
        flash(_("Debe seleccionar una deducción."), "error")
        return redirect(url_for("planilla.config_deducciones", planilla_id=planilla_id))

    existing = db.session.execute(
        db.select(PlanillaDeduccion).filter_by(planilla_id=planilla_id, deduccion_id=deduccion_id)
    ).scalar_one_or_none()

    if existing:
        flash(_("La deducción ya está asignada a esta planilla."), "warning")
        return redirect(url_for("planilla.config_deducciones", planilla_id=planilla_id))

    prioridad = request.form.get("prioridad", 100, type=int)
    es_obligatoria = request.form.get("es_obligatoria") == "on"

    association = PlanillaDeduccion(
        planilla_id=planilla_id,
        deduccion_id=deduccion_id,
        prioridad=prioridad,
        es_obligatoria=es_obligatoria,
        editable=True,
        activo=True,
        creado_por=current_user.usuario,
    )
    db.session.add(association)
    db.session.commit()
    flash(_("Deducción agregada exitosamente."), "success")
    return redirect(url_for("planilla.config_deducciones", planilla_id=planilla_id))


@planilla_bp.route("/<planilla_id>/deduccion/<association_id>/remove", methods=["POST"])
@require_write_access()
def remove_deduccion(planilla_id: str, association_id: str):
    """Remove a deduction from the planilla."""
    association = db.get_or_404(PlanillaDeduccion, association_id)
    db.session.delete(association)
    db.session.commit()
    flash(_("Deducción removida exitosamente."), "success")
    return redirect(url_for("planilla.config_deducciones", planilla_id=planilla_id))


@planilla_bp.route("/<planilla_id>/deduccion/<association_id>/update-priority", methods=["POST"])
@require_write_access()
def update_deduccion_priority(planilla_id: str, association_id: str):
    """Update the priority of a deduction."""
    association = db.get_or_404(PlanillaDeduccion, association_id)

    prioridad = request.form.get("prioridad", type=int)
    if prioridad is not None:
        association.prioridad = prioridad
        association.modificado_por = current_user.usuario
        db.session.commit()
        flash(_("Prioridad actualizada."), "success")

    return redirect(url_for("planilla.config_deducciones", planilla_id=planilla_id))


# ============================================================================
# BENEFIT (PRESTACION) ASSOCIATIONS
# ============================================================================


@planilla_bp.route("/<planilla_id>/prestacion/add", methods=["POST"])
@require_write_access()
def add_prestacion(planilla_id: str):
    """Add a benefit (prestacion) to the planilla."""
    # Verify planilla exists (raises 404 if not found)
    db.get_or_404(Planilla, planilla_id)
    prestacion_id = request.form.get("prestacion_id")

    if not prestacion_id:
        flash(_("Debe seleccionar una prestación."), "error")
        return redirect(url_for("planilla.config_prestaciones", planilla_id=planilla_id))

    existing = db.session.execute(
        db.select(PlanillaPrestacion).filter_by(planilla_id=planilla_id, prestacion_id=prestacion_id)
    ).scalar_one_or_none()

    if existing:
        flash(_("La prestación ya está asignada a esta planilla."), "warning")
        return redirect(url_for("planilla.config_prestaciones", planilla_id=planilla_id))

    orden = request.form.get("orden", 0, type=int)

    association = PlanillaPrestacion(
        planilla_id=planilla_id,
        prestacion_id=prestacion_id,
        orden=orden,
        editable=True,
        activo=True,
        creado_por=current_user.usuario,
    )
    db.session.add(association)
    db.session.commit()
    flash(_("Prestación agregada exitosamente."), "success")
    return redirect(url_for("planilla.config_prestaciones", planilla_id=planilla_id))


@planilla_bp.route("/<planilla_id>/prestacion/<association_id>/remove", methods=["POST"])
@require_write_access()
def remove_prestacion(planilla_id: str, association_id: str):
    """Remove a benefit from the planilla."""
    association = db.get_or_404(PlanillaPrestacion, association_id)
    db.session.delete(association)
    db.session.commit()
    flash(_("Prestación removida exitosamente."), "success")
    return redirect(url_for("planilla.config_prestaciones", planilla_id=planilla_id))


# ============================================================================
# CALCULATION RULE ASSOCIATIONS
# ============================================================================


@planilla_bp.route("/<planilla_id>/regla/add", methods=["POST"])
@require_write_access()
def add_regla(planilla_id: str):
    """Add a calculation rule to the planilla."""
    # Verify planilla exists (raises 404 if not found)
    db.get_or_404(Planilla, planilla_id)
    regla_calculo_id = request.form.get("regla_calculo_id")

    if not regla_calculo_id:
        flash(_("Debe seleccionar una regla de cálculo."), "error")
        return redirect(url_for("planilla.config_reglas", planilla_id=planilla_id))

    existing = db.session.execute(
        db.select(PlanillaReglaCalculo).filter_by(planilla_id=planilla_id, regla_calculo_id=regla_calculo_id)
    ).scalar_one_or_none()

    if existing:
        flash(_("La regla ya está asignada a esta planilla."), "warning")
        return redirect(url_for("planilla.config_reglas", planilla_id=planilla_id))

    orden = request.form.get("orden", 0, type=int)

    association = PlanillaReglaCalculo(
        planilla_id=planilla_id,
        regla_calculo_id=regla_calculo_id,
        orden=orden,
        activo=True,
        creado_por=current_user.usuario,
    )
    db.session.add(association)
    db.session.commit()
    flash(_("Regla de cálculo agregada exitosamente."), "success")
    return redirect(url_for("planilla.config_reglas", planilla_id=planilla_id))


@planilla_bp.route("/<planilla_id>/regla/<association_id>/remove", methods=["POST"])
@require_write_access()
def remove_regla(planilla_id: str, association_id: str):
    """Remove a calculation rule from the planilla."""
    association = db.get_or_404(PlanillaReglaCalculo, association_id)
    db.session.delete(association)
    db.session.commit()
    flash(_("Regla de cálculo removida exitosamente."), "success")
    return redirect(url_for("planilla.config_reglas", planilla_id=planilla_id))


# ============================================================================
# NOMINA EXECUTION
# ============================================================================


@planilla_bp.route("/<planilla_id>/ejecutar", methods=["GET", "POST"])
@require_write_access()
def ejecutar_nomina(planilla_id: str):
    """Execute a payroll run for a planilla."""
    from datetime import date
    from flask import current_app
    from coati_payroll.nomina_engine import NominaEngine
    from coati_payroll.model import Nomina
    from coati_payroll.enums import NominaEstado
    from coati_payroll.queue import get_queue_driver

    planilla = db.get_or_404(Planilla, planilla_id)

    if request.method == "POST":
        periodo_inicio = request.form.get("periodo_inicio")
        periodo_fin = request.form.get("periodo_fin")
        fecha_calculo = request.form.get("fecha_calculo")

        if not periodo_inicio or not periodo_fin:
            flash(_("Debe especificar el período de la nómina."), "error")
            return redirect(url_for("planilla.ejecutar_nomina", planilla_id=planilla_id))

        # Parse dates
        try:
            periodo_inicio = date.fromisoformat(periodo_inicio)
            periodo_fin = date.fromisoformat(periodo_fin)
            fecha_calculo = date.fromisoformat(fecha_calculo) if fecha_calculo else date.today()
        except ValueError:
            flash(_("Formato de fecha inválido."), "error")
            return redirect(url_for("planilla.ejecutar_nomina", planilla_id=planilla_id))

        # Count active employees
        num_empleados = sum(1 for pe in planilla.planilla_empleados if pe.activo and pe.empleado.activo)

        # Get configurable threshold for background processing
        # Default is 100, but can be adjusted via BACKGROUND_PAYROLL_THRESHOLD env var
        # - Lower (25-50) for systems with complex formulas or slow performance
        # - Higher (200-500) for high-performance systems with simple formulas
        threshold = current_app.config.get("BACKGROUND_PAYROLL_THRESHOLD", 100)

        # Determine if we should process in background
        # For large payrolls (>threshold employees), use background processing
        if num_empleados > threshold:
            # Create nomina record with "calculando" status
            nomina = Nomina(
                planilla_id=planilla_id,
                periodo_inicio=periodo_inicio,
                periodo_fin=periodo_fin,
                generado_por=current_user.usuario,
                estado=NominaEstado.CALCULANDO,
                total_bruto=0,
                total_deducciones=0,
                total_neto=0,
                total_empleados=num_empleados,
                empleados_procesados=0,
                empleados_con_error=0,
                procesamiento_en_background=True,
            )
            db.session.add(nomina)
            db.session.commit()

            # Enqueue background task
            try:
                queue = get_queue_driver()
                queue.enqueue(
                    "process_large_payroll",
                    nomina_id=nomina.id,
                    planilla_id=planilla_id,
                    periodo_inicio=periodo_inicio.isoformat(),
                    periodo_fin=periodo_fin.isoformat(),
                    fecha_calculo=fecha_calculo.isoformat(),
                    usuario=current_user.usuario,
                )
                flash(
                    _(
                        "La nómina está siendo calculada en segundo plano. "
                        "Se procesarán %(num)d empleados. "
                        "Por favor, revise el progreso en unos momentos.",
                        num=num_empleados,
                    ),
                    "info",
                )
                return redirect(
                    url_for(
                        "planilla.ver_nomina",
                        planilla_id=planilla_id,
                        nomina_id=nomina.id,
                    )
                )
            except Exception as e:
                # If background processing fails, mark nomina as error
                nomina.estado = NominaEstado.ERROR
                nomina.errores_calculo = {"background_task_initialization_error": str(e)}
                db.session.commit()
                flash(
                    _(
                        "Error al iniciar el procesamiento en segundo plano: %(error)s",
                        error=str(e),
                    ),
                    "error",
                )
                return redirect(url_for("planilla.ejecutar_nomina", planilla_id=planilla_id))
        else:
            # For smaller payrolls, process synchronously as before
            # Execute the payroll
            engine = NominaEngine(
                planilla=planilla,
                periodo_inicio=periodo_inicio,
                periodo_fin=periodo_fin,
                fecha_calculo=fecha_calculo,
                usuario=current_user.usuario,
            )

            nomina = engine.ejecutar()

            if engine.errors:
                for error in engine.errors:
                    flash(error, "error")

            if engine.warnings:
                for warning in engine.warnings:
                    flash(warning, "warning")

            if nomina:
                flash(_("Nómina generada exitosamente."), "success")
                return redirect(
                    url_for(
                        "planilla.ver_nomina",
                        planilla_id=planilla_id,
                        nomina_id=nomina.id,
                    )
                )
            else:
                return redirect(url_for("planilla.ejecutar_nomina", planilla_id=planilla_id))

    # GET - show execution form
    # Get last nomina for default dates
    ultima_nomina = db.session.execute(
        db.select(Nomina).filter_by(planilla_id=planilla_id).order_by(Nomina.periodo_fin.desc())
    ).scalar()

    # Calculate suggested period
    from datetime import timedelta

    hoy = date.today()

    if ultima_nomina:
        # Start from the day after last period ended
        periodo_inicio_sugerido = ultima_nomina.periodo_fin + timedelta(days=1)
    else:
        # First day of current month
        periodo_inicio_sugerido = hoy.replace(day=1)

    # Calculate end of period based on tipo_planilla
    tipo = planilla.tipo_planilla
    match tipo.periodicidad if tipo else "mensual":
        case "semanal":
            periodo_fin_sugerido = periodo_inicio_sugerido + timedelta(days=6)
        case "quincenal":
            if periodo_inicio_sugerido.day <= 15:
                periodo_fin_sugerido = periodo_inicio_sugerido.replace(day=15)
            else:
                # End of month
                next_month = periodo_inicio_sugerido.replace(day=28) + timedelta(days=4)
                periodo_fin_sugerido = next_month - timedelta(days=next_month.day)
        case _:  # mensual or other
            # End of month
            next_month = periodo_inicio_sugerido.replace(day=28) + timedelta(days=4)
            periodo_fin_sugerido = next_month - timedelta(days=next_month.day)

    return render_template(
        "modules/planilla/ejecutar_nomina.html",
        planilla=planilla,
        periodo_inicio=periodo_inicio_sugerido,
        periodo_fin=periodo_fin_sugerido,
        fecha_calculo=hoy,
        ultima_nomina=ultima_nomina,
    )


@planilla_bp.route("/<planilla_id>/nominas")
@require_read_access()
def listar_nominas(planilla_id: str):
    """List all nominas for a planilla."""
    from coati_payroll.model import Nomina

    planilla = db.get_or_404(Planilla, planilla_id)
    nominas = (
        db.session.execute(db.select(Nomina).filter_by(planilla_id=planilla_id).order_by(Nomina.periodo_fin.desc()))
        .scalars()
        .all()
    )

    return render_template(
        "modules/planilla/listar_nominas.html",
        planilla=planilla,
        nominas=nominas,
    )


@planilla_bp.route("/<planilla_id>/nomina/<nomina_id>")
@require_read_access()
def ver_nomina(planilla_id: str, nomina_id: str):
    """View details of a specific nomina."""
    from coati_payroll.model import Nomina, NominaEmpleado

    planilla = db.get_or_404(Planilla, planilla_id)
    nomina = db.get_or_404(Nomina, nomina_id)

    if nomina.planilla_id != planilla_id:
        flash(_("La nómina no pertenece a esta planilla."), "error")
        return redirect(url_for("planilla.listar_nominas", planilla_id=planilla_id))

    nomina_empleados = db.session.execute(db.select(NominaEmpleado).filter_by(nomina_id=nomina_id)).scalars().all()

    return render_template(
        "modules/planilla/ver_nomina.html",
        planilla=planilla,
        nomina=nomina,
        nomina_empleados=nomina_empleados,
    )


@planilla_bp.route("/<planilla_id>/nomina/<nomina_id>/empleado/<nomina_empleado_id>")
@require_read_access()
def ver_nomina_empleado(planilla_id: str, nomina_id: str, nomina_empleado_id: str):
    """View details of an employee's payroll."""
    from coati_payroll.model import Nomina, NominaEmpleado, NominaDetalle

    planilla = db.get_or_404(Planilla, planilla_id)
    nomina = db.get_or_404(Nomina, nomina_id)
    nomina_empleado = db.get_or_404(NominaEmpleado, nomina_empleado_id)

    if nomina_empleado.nomina_id != nomina_id:
        flash(_("El detalle no pertenece a esta nómina."), "error")
        return redirect(url_for("planilla.ver_nomina", planilla_id=planilla_id, nomina_id=nomina_id))

    detalles = (
        db.session.execute(
            db.select(NominaDetalle).filter_by(nomina_empleado_id=nomina_empleado_id).order_by(NominaDetalle.orden)
        )
        .scalars()
        .all()
    )

    # Separate by type
    percepciones = [d for d in detalles if d.tipo == "ingreso"]
    deducciones = [d for d in detalles if d.tipo == "deduccion"]
    prestaciones = [d for d in detalles if d.tipo == "prestacion"]

    return render_template(
        "modules/planilla/ver_nomina_empleado.html",
        planilla=planilla,
        nomina=nomina,
        nomina_empleado=nomina_empleado,
        percepciones=percepciones,
        deducciones=deducciones,
        prestaciones=prestaciones,
    )


@planilla_bp.route("/<planilla_id>/nomina/<nomina_id>/progreso")
@require_read_access()
def progreso_nomina(planilla_id: str, nomina_id: str):
    """API endpoint to check calculation progress of a nomina."""
    from flask import jsonify
    from coati_payroll.model import Nomina

    nomina = db.get_or_404(Nomina, nomina_id)

    if nomina.planilla_id != planilla_id:
        return jsonify({"error": "Nomina does not belong to this planilla"}), 404

    return jsonify(
        {
            "estado": nomina.estado,
            "total_empleados": nomina.total_empleados or 0,
            "empleados_procesados": nomina.empleados_procesados or 0,
            "empleados_con_error": nomina.empleados_con_error or 0,
            "progreso_porcentaje": (
                int((nomina.empleados_procesados / nomina.total_empleados) * 100)
                if nomina.total_empleados and nomina.total_empleados > 0
                else 0
            ),
            "errores_calculo": nomina.errores_calculo or {},
            "procesamiento_en_background": nomina.procesamiento_en_background,
            "empleado_actual": nomina.empleado_actual or "",
            "log_procesamiento": nomina.log_procesamiento or [],
        }
    )


@planilla_bp.route("/<planilla_id>/nomina/<nomina_id>/aprobar", methods=["POST"])
@require_write_access()
def aprobar_nomina(planilla_id: str, nomina_id: str):
    """Approve a nomina for payment."""
    from coati_payroll.model import Nomina

    nomina = db.get_or_404(Nomina, nomina_id)

    if nomina.planilla_id != planilla_id:
        flash(_("La nómina no pertenece a esta planilla."), "error")
        return redirect(url_for("planilla.listar_nominas", planilla_id=planilla_id))

    if nomina.estado != "generado":
        flash(_("Solo se pueden aprobar nóminas en estado 'generado'."), "error")
        return redirect(url_for("planilla.ver_nomina", planilla_id=planilla_id, nomina_id=nomina_id))

    nomina.estado = "aprobado"
    nomina.modificado_por = current_user.usuario
    db.session.commit()

    flash(_("Nómina aprobada exitosamente."), "success")
    return redirect(url_for("planilla.ver_nomina", planilla_id=planilla_id, nomina_id=nomina_id))


@planilla_bp.route("/<planilla_id>/nomina/<nomina_id>/aplicar", methods=["POST"])
@require_write_access()
def aplicar_nomina(planilla_id: str, nomina_id: str):
    """Mark a nomina as applied (paid)."""
    from coati_payroll.model import Nomina

    nomina = db.get_or_404(Nomina, nomina_id)

    if nomina.planilla_id != planilla_id:
        flash(_("La nómina no pertenece a esta planilla."), "error")
        return redirect(url_for("planilla.listar_nominas", planilla_id=planilla_id))

    if nomina.estado != "aprobado":
        flash(_("Solo se pueden aplicar nóminas en estado 'aprobado'."), "error")
        return redirect(url_for("planilla.ver_nomina", planilla_id=planilla_id, nomina_id=nomina_id))

    nomina.estado = "aplicado"
    nomina.modificado_por = current_user.usuario

    # Actualizar estado de todas las novedades asociadas a "ejecutada"
    from coati_payroll.model import NominaNovedad
    from coati_payroll.enums import NovedadEstado

    # Obtener empleados de la planilla para filtrar novedades
    planilla = db.get_or_404(Planilla, planilla_id)
    empleado_ids = [pe.empleado_id for pe in planilla.planilla_empleados if pe.activo]

    # Actualizar novedades que corresponden a este período
    novedades = (
        db.session.execute(
            db.select(NominaNovedad).filter(
                NominaNovedad.empleado_id.in_(empleado_ids),
                NominaNovedad.fecha_novedad >= nomina.periodo_inicio,
                NominaNovedad.fecha_novedad <= nomina.periodo_fin,
                NominaNovedad.estado == NovedadEstado.PENDIENTE,
            )
        )
        .scalars()
        .all()
    )

    for novedad in novedades:
        novedad.estado = NovedadEstado.EJECUTADA
        novedad.modificado_por = current_user.usuario

    db.session.commit()

    flash(
        _("Nómina aplicada exitosamente. {} novedad(es) marcadas como ejecutadas.").format(len(novedades)),
        "success",
    )
    return redirect(url_for("planilla.ver_nomina", planilla_id=planilla_id, nomina_id=nomina_id))


@planilla_bp.route("/<planilla_id>/nomina/<nomina_id>/recalcular", methods=["POST"])
@require_write_access()
def recalcular_nomina(planilla_id: str, nomina_id: str):
    """Recalculate an existing nomina.

    Allows recalculation of a payroll unless it is in 'aplicado' (paid) status.
    This is useful when errors are found during review and the payroll needs
    to be completely recalculated.

    The existing NominaEmpleado, NominaDetalle, and related AdelantoAbono records
    are deleted and recreated during recalculation.
    """
    from datetime import date as date_type
    from coati_payroll.model import (
        Nomina,
        NominaEmpleado,
        NominaDetalle,
        NominaNovedad,
        AdelantoAbono,
    )
    from coati_payroll.nomina_engine import NominaEngine

    nomina = db.get_or_404(Nomina, nomina_id)
    planilla = db.get_or_404(Planilla, planilla_id)

    if nomina.planilla_id != planilla_id:
        flash(_("La nómina no pertenece a esta planilla."), "error")
        return redirect(url_for("planilla.listar_nominas", planilla_id=planilla_id))

    if nomina.estado == "aplicado":
        flash(
            _("No se puede recalcular una nómina en estado 'aplicado' (pagada)."),
            "error",
        )
        return redirect(url_for("planilla.ver_nomina", planilla_id=planilla_id, nomina_id=nomina_id))

    # Store the original period information
    periodo_inicio = nomina.periodo_inicio
    periodo_fin = nomina.periodo_fin

    # Delete related AdelantoAbono records (payments towards loans/advances)
    # that were created for this nomina
    db.session.execute(db.delete(AdelantoAbono).where(AdelantoAbono.nomina_id == nomina_id))

    # Delete NominaNovedad records (novelties) associated with this nomina
    db.session.execute(db.delete(NominaNovedad).where(NominaNovedad.nomina_id == nomina_id))

    # Delete NominaDetalle records for all NominaEmpleado of this nomina
    # using a subquery for efficiency
    db.session.execute(
        db.delete(NominaDetalle).where(
            NominaDetalle.nomina_empleado_id.in_(
                db.select(NominaEmpleado.id).where(NominaEmpleado.nomina_id == nomina_id)
            )
        )
    )

    # Delete all NominaEmpleado records
    db.session.execute(db.delete(NominaEmpleado).where(NominaEmpleado.nomina_id == nomina_id))

    # Delete the nomina record itself
    db.session.delete(nomina)
    db.session.commit()

    # Re-execute the payroll with the same period
    engine = NominaEngine(
        planilla=planilla,
        periodo_inicio=periodo_inicio,
        periodo_fin=periodo_fin,
        fecha_calculo=date_type.today(),
        usuario=current_user.usuario,
    )

    new_nomina = engine.ejecutar()

    if engine.errors:
        for error in engine.errors:
            flash(error, "error")

    if engine.warnings:
        for warning in engine.warnings:
            flash(warning, "warning")

    if new_nomina:
        flash(_("Nómina recalculada exitosamente."), "success")
        return redirect(url_for("planilla.ver_nomina", planilla_id=planilla_id, nomina_id=new_nomina.id))
    else:
        flash(_("Error al recalcular la nómina."), "error")
        return redirect(url_for("planilla.listar_nominas", planilla_id=planilla_id))


# ============================================================================
# HELPER FUNCTIONS
# ============================================================================


def _get_planilla_component_counts(planilla_id: str) -> dict:
    """Get counts of all components associated with a planilla.

    Returns a dictionary with counts for empleados, percepciones, deducciones,
    prestaciones, and reglas.
    """
    from sqlalchemy import func, select

    return {
        "empleados_count": db.session.execute(
            select(func.count()).select_from(PlanillaEmpleado).filter_by(planilla_id=planilla_id)
        ).scalar(),
        "percepciones_count": db.session.execute(
            select(func.count()).select_from(PlanillaIngreso).filter_by(planilla_id=planilla_id)
        ).scalar(),
        "deducciones_count": db.session.execute(
            select(func.count()).select_from(PlanillaDeduccion).filter_by(planilla_id=planilla_id)
        ).scalar(),
        "prestaciones_count": db.session.execute(
            select(func.count()).select_from(PlanillaPrestacion).filter_by(planilla_id=planilla_id)
        ).scalar(),
        "reglas_count": db.session.execute(
            select(func.count()).select_from(PlanillaReglaCalculo).filter_by(planilla_id=planilla_id)
        ).scalar(),
    }


def _populate_form_choices(form: PlanillaForm):
    """Populate form select choices from database."""
    from coati_payroll.model import Empresa

    tipos = (
        db.session.execute(db.select(TipoPlanilla).filter_by(activo=True).order_by(TipoPlanilla.codigo)).scalars().all()
    )
    form.tipo_planilla_id.choices = [("", _("-- Seleccionar --"))] + [
        (t.id, f"{t.codigo} - {t.descripcion or t.codigo}") for t in tipos
    ]

    monedas = db.session.execute(db.select(Moneda).filter_by(activo=True).order_by(Moneda.codigo)).scalars().all()
    form.moneda_id.choices = [("", _("-- Seleccionar --"))] + [(m.id, f"{m.codigo} - {m.nombre}") for m in monedas]

    empresas = (
        db.session.execute(db.select(Empresa).filter_by(activo=True).order_by(Empresa.razon_social)).scalars().all()
    )
    form.empresa_id.choices = [("", _("-- Seleccionar --"))] + [
        (e.id, f"{e.codigo} - {e.razon_social}") for e in empresas
    ]


# ============================================================================
# NOMINA NOVEDADES (NOVELTIES)
# ============================================================================


@planilla_bp.route("/<planilla_id>/nomina/<nomina_id>/novedades")
@require_read_access()
def listar_novedades(planilla_id: str, nomina_id: str):
    """List all novedades (novelties) for a specific nomina.

    Shows novedades that fall within the nomina's period dates,
    regardless of which nomina_id they were originally associated with.
    This ensures that recalculations show the correct novedades.
    """
    from coati_payroll.model import Nomina, NominaNovedad

    planilla = db.get_or_404(Planilla, planilla_id)
    nomina = db.get_or_404(Nomina, nomina_id)

    if nomina.planilla_id != planilla_id:
        flash(_("La nómina no pertenece a esta planilla."), "error")
        return redirect(url_for("planilla.listar_nominas", planilla_id=planilla_id))

    # Get all employees in this planilla
    empleado_ids = [pe.empleado_id for pe in planilla.planilla_empleados if pe.activo]

    # Query novedades that fall within the nomina period and are for employees in this planilla
    novedades = (
        db.session.execute(
            db.select(NominaNovedad)
            .filter(
                NominaNovedad.empleado_id.in_(empleado_ids),
                NominaNovedad.fecha_novedad >= nomina.periodo_inicio,
                NominaNovedad.fecha_novedad <= nomina.periodo_fin,
            )
            .order_by(NominaNovedad.fecha_novedad.desc(), NominaNovedad.timestamp.desc())
        )
        .scalars()
        .all()
    )

    return render_template(
        "modules/planilla/novedades/index.html",
        planilla=planilla,
        nomina=nomina,
        novedades=novedades,
    )


@planilla_bp.route("/<planilla_id>/nomina/<nomina_id>/novedades/new", methods=["GET", "POST"])
@require_write_access()
def nueva_novedad(planilla_id: str, nomina_id: str):
    """Add a new novedad to a nomina."""
    from decimal import Decimal
    from coati_payroll.model import Nomina, NominaNovedad
    from coati_payroll.forms import NominaNovedadForm

    planilla = db.get_or_404(Planilla, planilla_id)
    nomina = db.get_or_404(Nomina, nomina_id)

    if nomina.planilla_id != planilla_id:
        flash(_("La nómina no pertenece a esta planilla."), "error")
        return redirect(url_for("planilla.listar_nominas", planilla_id=planilla_id))

    if nomina.estado == "aplicado":
        flash(_("No se pueden agregar novedades a una nómina aplicada."), "error")
        return redirect(
            url_for(
                "planilla.listar_novedades",
                planilla_id=planilla_id,
                nomina_id=nomina_id,
            )
        )

    form = NominaNovedadForm()
    _populate_novedad_form_choices(form, nomina_id)

    if form.validate_on_submit():
        # Validate that fecha_novedad falls within the nomina period
        if form.fecha_novedad.data:
            if form.fecha_novedad.data < nomina.periodo_inicio or form.fecha_novedad.data > nomina.periodo_fin:
                flash(
                    _("La fecha de la novedad debe estar dentro del período de la nómina " "({} a {}).").format(
                        nomina.periodo_inicio.strftime("%d/%m/%Y"),
                        nomina.periodo_fin.strftime("%d/%m/%Y"),
                    ),
                    "error",
                )
                return render_template(
                    "modules/planilla/novedades/form.html",
                    form=form,
                    planilla=planilla,
                    nomina=nomina,
                )

        # Determine percepcion_id or deduccion_id based on tipo_concepto
        percepcion_id, deduccion_id = _get_concepto_ids_from_form(form)

        novedad = NominaNovedad(
            nomina_id=nomina_id,
            empleado_id=form.empleado_id.data,
            codigo_concepto=form.codigo_concepto.data,
            tipo_valor=form.tipo_valor.data,
            valor_cantidad=Decimal(str(form.valor_cantidad.data)),
            fecha_novedad=form.fecha_novedad.data,
            percepcion_id=percepcion_id,
            deduccion_id=deduccion_id,
            creado_por=current_user.usuario,
        )
        db.session.add(novedad)
        db.session.commit()
        flash(_("Novedad agregada exitosamente."), "success")
        return redirect(
            url_for(
                "planilla.listar_novedades",
                planilla_id=planilla_id,
                nomina_id=nomina_id,
            )
        )

    return render_template(
        "modules/planilla/novedades/form.html",
        planilla=planilla,
        nomina=nomina,
        form=form,
        is_edit=False,
    )


@planilla_bp.route(
    "/<planilla_id>/nomina/<nomina_id>/novedades/<novedad_id>/edit",
    methods=["GET", "POST"],
)
@require_write_access()
def editar_novedad(planilla_id: str, nomina_id: str, novedad_id: str):
    """Edit an existing novedad."""
    from decimal import Decimal
    from coati_payroll.model import Nomina, NominaNovedad
    from coati_payroll.forms import NominaNovedadForm

    planilla = db.get_or_404(Planilla, planilla_id)
    nomina = db.get_or_404(Nomina, nomina_id)
    novedad = db.get_or_404(NominaNovedad, novedad_id)

    if nomina.planilla_id != planilla_id:
        flash(_("La nómina no pertenece a esta planilla."), "error")
        return redirect(url_for("planilla.listar_nominas", planilla_id=planilla_id))

    if novedad.nomina_id != nomina_id:
        flash(_("La novedad no pertenece a esta nómina."), "error")
        return redirect(
            url_for(
                "planilla.listar_novedades",
                planilla_id=planilla_id,
                nomina_id=nomina_id,
            )
        )

    if nomina.estado == "aplicado":
        flash(_("No se pueden editar novedades de una nómina aplicada."), "error")
        return redirect(
            url_for(
                "planilla.listar_novedades",
                planilla_id=planilla_id,
                nomina_id=nomina_id,
            )
        )

    form = NominaNovedadForm(obj=novedad)
    _populate_novedad_form_choices(form, nomina_id)

    # Set tipo_concepto based on existing data
    if request.method == "GET":
        if novedad.percepcion_id:
            form.tipo_concepto.data = "percepcion"
            form.percepcion_id.data = novedad.percepcion_id
        elif novedad.deduccion_id:
            form.tipo_concepto.data = "deduccion"
            form.deduccion_id.data = novedad.deduccion_id

    if form.validate_on_submit():
        # Validate that fecha_novedad falls within the nomina period
        if form.fecha_novedad.data:
            if form.fecha_novedad.data < nomina.periodo_inicio or form.fecha_novedad.data > nomina.periodo_fin:
                flash(
                    _("La fecha de la novedad debe estar dentro del período de la nómina " "({} a {}).").format(
                        nomina.periodo_inicio.strftime("%d/%m/%Y"),
                        nomina.periodo_fin.strftime("%d/%m/%Y"),
                    ),
                    "error",
                )
                return render_template(
                    "modules/planilla/novedades/form.html",
                    form=form,
                    planilla=planilla,
                    nomina=nomina,
                    novedad=novedad,
                )

        novedad.empleado_id = form.empleado_id.data
        novedad.codigo_concepto = form.codigo_concepto.data
        novedad.tipo_valor = form.tipo_valor.data
        novedad.valor_cantidad = Decimal(str(form.valor_cantidad.data))
        novedad.fecha_novedad = form.fecha_novedad.data

        # Update percepcion_id or deduccion_id based on tipo_concepto
        percepcion_id, deduccion_id = _get_concepto_ids_from_form(form)
        novedad.percepcion_id = percepcion_id
        novedad.deduccion_id = deduccion_id

        novedad.modificado_por = current_user.usuario
        db.session.commit()
        flash(_("Novedad actualizada exitosamente."), "success")
        return redirect(
            url_for(
                "planilla.listar_novedades",
                planilla_id=planilla_id,
                nomina_id=nomina_id,
            )
        )

    return render_template(
        "modules/planilla/novedades/form.html",
        planilla=planilla,
        nomina=nomina,
        form=form,
        novedad=novedad,
        is_edit=True,
    )


@planilla_bp.route(
    "/<planilla_id>/nomina/<nomina_id>/novedades/<novedad_id>/delete",
    methods=["POST"],
)
@require_write_access()
def eliminar_novedad(planilla_id: str, nomina_id: str, novedad_id: str):
    """Delete a novedad from a nomina."""
    from coati_payroll.model import Nomina, NominaNovedad

    nomina = db.get_or_404(Nomina, nomina_id)
    novedad = db.get_or_404(NominaNovedad, novedad_id)

    if nomina.planilla_id != planilla_id:
        flash(_("La nómina no pertenece a esta planilla."), "error")
        return redirect(url_for("planilla.listar_nominas", planilla_id=planilla_id))

    if novedad.nomina_id != nomina_id:
        flash(_("La novedad no pertenece a esta nómina."), "error")
        return redirect(
            url_for(
                "planilla.listar_novedades",
                planilla_id=planilla_id,
                nomina_id=nomina_id,
            )
        )

    if nomina.estado == "aplicado":
        flash(_("No se pueden eliminar novedades de una nómina aplicada."), "error")
        return redirect(
            url_for(
                "planilla.listar_novedades",
                planilla_id=planilla_id,
                nomina_id=nomina_id,
            )
        )

    db.session.delete(novedad)
    db.session.commit()
    flash(_("Novedad eliminada exitosamente."), "success")
    return redirect(url_for("planilla.listar_novedades", planilla_id=planilla_id, nomina_id=nomina_id))


def _populate_novedad_form_choices(form, nomina_id: str):
    """Populate form select choices for novedad form."""
    from coati_payroll.model import NominaEmpleado, Percepcion, Deduccion

    # Get employees associated with this nomina
    nomina_empleados = db.session.execute(db.select(NominaEmpleado).filter_by(nomina_id=nomina_id)).scalars().all()
    form.empleado_id.choices = [("", _("-- Seleccionar Empleado --"))] + [
        (
            ne.empleado.id,
            f"{ne.empleado.primer_nombre} {ne.empleado.primer_apellido} ({ne.empleado.codigo_empleado})",
        )
        for ne in nomina_empleados
    ]

    # Get active percepciones
    percepciones = (
        db.session.execute(db.select(Percepcion).filter_by(activo=True).order_by(Percepcion.nombre)).scalars().all()
    )
    form.percepcion_id.choices = [("", _("-- Seleccionar Percepción --"))] + [
        (p.id, f"{p.codigo} - {p.nombre}") for p in percepciones
    ]

    # Get active deducciones
    deducciones = (
        db.session.execute(db.select(Deduccion).filter_by(activo=True).order_by(Deduccion.nombre)).scalars().all()
    )
    form.deduccion_id.choices = [("", _("-- Seleccionar Deducción --"))] + [
        (d.id, f"{d.codigo} - {d.nombre}") for d in deducciones
    ]


def _get_concepto_ids_from_form(form) -> tuple[str | None, str | None]:
    """Extract percepcion_id and deduccion_id from form based on tipo_concepto.

    Args:
        form: The NominaNovedadForm with submitted data

    Returns:
        Tuple of (percepcion_id, deduccion_id) - one will be set, other will be None
    """
    percepcion_id = None
    deduccion_id = None

    if form.tipo_concepto.data == "percepcion":
        percepcion_id = form.percepcion_id.data if form.percepcion_id.data else None
    else:
        deduccion_id = form.deduccion_id.data if form.deduccion_id.data else None

    return percepcion_id, deduccion_id


def _check_openpyxl_available():
    """Check if openpyxl is available and return necessary classes.

    Returns:
        tuple: (Workbook, Font, Alignment, PatternFill, Border, Side) or None if not available
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        return Workbook, Font, Alignment, PatternFill, Border, Side
    except ImportError:
        return None


@planilla_bp.route("/<planilla_id>/nomina/<nomina_id>/exportar-excel")
@login_required
def exportar_nomina_excel(planilla_id: str, nomina_id: str):
    """Export nomina to Excel with employee details and calculations."""
    from io import BytesIO
    from flask import send_file
    from coati_payroll.model import Nomina, NominaEmpleado

    openpyxl_classes = _check_openpyxl_available()
    if not openpyxl_classes:
        flash(_("Excel export no disponible. Instale openpyxl."), "warning")
        return redirect(url_for("planilla.ver_nomina", planilla_id=planilla_id, nomina_id=nomina_id))

    Workbook, Font, Alignment, PatternFill, Border, Side = openpyxl_classes

    planilla = db.get_or_404(Planilla, planilla_id)
    nomina = db.get_or_404(Nomina, nomina_id)

    if nomina.planilla_id != planilla_id:
        flash(_("La nómina no pertenece a esta planilla."), "error")
        return redirect(url_for("planilla.listar_nominas", planilla_id=planilla_id))

    # Get all nomina employees
    nomina_empleados = db.session.execute(db.select(NominaEmpleado).filter_by(nomina_id=nomina_id)).scalars().all()

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Nómina"

    # Define styles
    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    subheader_font = Font(bold=True, size=11)
    subheader_fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Title
    ws.merge_cells("A1:P1")
    title_cell = ws["A1"]
    title_cell.value = f"NÓMINA - {planilla.nombre}"
    title_cell.font = header_font
    title_cell.fill = header_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Nomina info
    row = 3
    # Company information if linked
    if planilla.empresa_id and planilla.empresa:
        ws[f"A{row}"] = "Empresa:"
        ws[f"B{row}"] = planilla.empresa.razon_social
        row += 1
        if planilla.empresa.ruc:
            ws[f"A{row}"] = "RUC:"
            ws[f"B{row}"] = planilla.empresa.ruc
            row += 1
    ws[f"A{row}"] = "Período:"
    ws[f"B{row}"] = f"{nomina.periodo_inicio.strftime('%d/%m/%Y')} - {nomina.periodo_fin.strftime('%d/%m/%Y')}"
    row += 1
    ws[f"A{row}"] = "Estado:"
    ws[f"B{row}"] = nomina.estado
    row += 1
    ws[f"A{row}"] = "Generado por:"
    ws[f"B{row}"] = nomina.generado_por or ""
    row += 2

    # Table headers
    headers = [
        "Cód. Empleado",
        "Identificación",
        "No. Seg. Social",
        "ID Fiscal",
        "Nombres",
        "Apellidos",
        "Cargo",
        "Salario Base",
        "Total Percepciones",
        "Total Deducciones",
        "Salario Neto",
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # Data rows
    for ne in nomina_empleados:
        row += 1
        emp = ne.empleado

        ws.cell(row=row, column=1, value=emp.codigo_empleado).border = border
        ws.cell(row=row, column=2, value=emp.identificacion_personal).border = border
        ws.cell(row=row, column=3, value=emp.id_seguridad_social or "").border = border
        ws.cell(row=row, column=4, value=emp.id_fiscal or "").border = border
        ws.cell(row=row, column=5, value=f"{emp.primer_nombre} {emp.segundo_nombre or ''}".strip()).border = border
        ws.cell(row=row, column=6, value=f"{emp.primer_apellido} {emp.segundo_apellido or ''}".strip()).border = border
        ws.cell(row=row, column=7, value=ne.cargo_snapshot or emp.cargo or "").border = border
        ws.cell(row=row, column=8, value=float(ne.sueldo_base_historico)).border = border
        ws.cell(row=row, column=9, value=float(ne.total_ingresos)).border = border
        ws.cell(row=row, column=10, value=float(ne.total_deducciones)).border = border
        ws.cell(row=row, column=11, value=float(ne.salario_neto)).border = border

    # Auto-adjust column widths
    for col in range(1, 12):
        ws.column_dimensions[chr(64 + col)].width = 15

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"nomina_{planilla.nombre}_{nomina.periodo_inicio.strftime('%Y%m%d')}_{nomina.id[:8]}.xlsx"

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@planilla_bp.route("/<planilla_id>/nomina/<nomina_id>/exportar-prestaciones-excel")
@login_required
def exportar_prestaciones_excel(planilla_id: str, nomina_id: str):
    """Export benefits (prestaciones) to Excel separately."""
    from io import BytesIO
    from flask import send_file
    from coati_payroll.model import Nomina, NominaEmpleado, NominaDetalle

    openpyxl_classes = _check_openpyxl_available()
    if not openpyxl_classes:
        flash(_("Excel export no disponible. Instale openpyxl."), "warning")
        return redirect(url_for("planilla.ver_nomina", planilla_id=planilla_id, nomina_id=nomina_id))

    Workbook, Font, Alignment, PatternFill, Border, Side = openpyxl_classes

    planilla = db.get_or_404(Planilla, planilla_id)
    nomina = db.get_or_404(Nomina, nomina_id)

    if nomina.planilla_id != planilla_id:
        flash(_("La nómina no pertenece a esta planilla."), "error")
        return redirect(url_for("planilla.listar_nominas", planilla_id=planilla_id))

    # Get all nomina employees
    nomina_empleados = db.session.execute(db.select(NominaEmpleado).filter_by(nomina_id=nomina_id)).scalars().all()

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Prestaciones"

    # Define styles
    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    subheader_font = Font(bold=True, size=11)
    subheader_fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Title
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = f"PRESTACIONES LABORALES - {planilla.nombre}"
    title_cell.font = header_font
    title_cell.fill = header_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Nomina info
    row = 3
    # Company information if linked
    if planilla.empresa_id and planilla.empresa:
        ws[f"A{row}"] = "Empresa:"
        ws[f"B{row}"] = planilla.empresa.razon_social
        row += 1
        if planilla.empresa.ruc:
            ws[f"A{row}"] = "RUC:"
            ws[f"B{row}"] = planilla.empresa.ruc
            row += 1
    ws[f"A{row}"] = "Período:"
    ws[f"B{row}"] = f"{nomina.periodo_inicio.strftime('%d/%m/%Y')} - {nomina.periodo_fin.strftime('%d/%m/%Y')}"
    row += 2

    # Table headers
    headers = ["Cód. Empleado", "Nombres", "Apellidos"]

    # Get all unique prestaciones
    prestaciones_set = set()
    for ne in nomina_empleados:
        detalles = (
            db.session.execute(db.select(NominaDetalle).filter_by(nomina_empleado_id=ne.id, tipo="prestacion"))
            .scalars()
            .all()
        )
        for d in detalles:
            prestaciones_set.add((d.codigo, d.descripcion))

    prestaciones_list = sorted(prestaciones_set, key=lambda x: x[0])
    headers.extend([p[1] or p[0] for p in prestaciones_list])

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # Data rows
    for ne in nomina_empleados:
        row += 1
        emp = ne.empleado

        ws.cell(row=row, column=1, value=emp.codigo_empleado).border = border
        ws.cell(row=row, column=2, value=f"{emp.primer_nombre} {emp.segundo_nombre or ''}".strip()).border = border
        ws.cell(row=row, column=3, value=f"{emp.primer_apellido} {emp.segundo_apellido or ''}".strip()).border = border

        # Get prestaciones for this employee
        detalles = (
            db.session.execute(
                db.select(NominaDetalle)
                .filter_by(nomina_empleado_id=ne.id, tipo="prestacion")
                .order_by(NominaDetalle.orden)
            )
            .scalars()
            .all()
        )

        prestaciones_dict = {d.codigo: float(d.monto) for d in detalles}

        # Fill prestacion amounts
        for col_idx, (codigo, _nombre) in enumerate(prestaciones_list, start=4):
            cell = ws.cell(row=row, column=col_idx, value=prestaciones_dict.get(codigo, 0.0))
            cell.border = border

    # Auto-adjust column widths (limit to 26 columns for simplicity)
    for col in range(1, min(len(headers) + 1, 27)):
        ws.column_dimensions[chr(64 + col)].width = 15

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"prestaciones_{planilla.nombre}_{nomina.periodo_inicio.strftime('%Y%m%d')}_{nomina.id[:8]}.xlsx"

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@planilla_bp.route("/<planilla_id>/nomina/<nomina_id>/exportar-comprobante-excel")
@login_required
def exportar_comprobante_excel(planilla_id: str, nomina_id: str):
    """Export accounting voucher (comprobante contable) to Excel."""
    from io import BytesIO
    from flask import send_file
    from coati_payroll.model import Nomina, NominaEmpleado, NominaDetalle, ComprobanteContable
    from decimal import Decimal

    openpyxl_classes = _check_openpyxl_available()
    if not openpyxl_classes:
        flash(_("Excel export no disponible. Instale openpyxl."), "warning")
        return redirect(url_for("planilla.ver_nomina", planilla_id=planilla_id, nomina_id=nomina_id))

    Workbook, Font, Alignment, PatternFill, Border, Side = openpyxl_classes

    planilla = db.get_or_404(Planilla, planilla_id)
    nomina = db.get_or_404(Nomina, nomina_id)

    if nomina.planilla_id != planilla_id:
        flash(_("La nómina no pertenece a esta planilla."), "error")
        return redirect(url_for("planilla.listar_nominas", planilla_id=planilla_id))

    # Get all nomina employees
    nomina_empleados = db.session.execute(db.select(NominaEmpleado).filter_by(nomina_id=nomina_id)).scalars().all()

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Comprobante Contable"

    # Define styles
    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    subheader_font = Font(bold=True, size=11)
    subheader_fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
    total_font = Font(bold=True, size=11)
    total_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Title
    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    title_cell.value = "COMPROBANTE CONTABLE - NÓMINA"
    title_cell.font = header_font
    title_cell.fill = header_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Header info
    row = 3
    # Company information if linked
    if planilla.empresa_id and planilla.empresa:
        ws[f"A{row}"] = "Empresa:"
        ws[f"B{row}"] = planilla.empresa.razon_social
        row += 1
        if planilla.empresa.ruc:
            ws[f"A{row}"] = "RUC:"
            ws[f"B{row}"] = planilla.empresa.ruc
            row += 1
    ws[f"A{row}"] = "Planilla:"
    ws[f"B{row}"] = planilla.nombre
    row += 1
    ws[f"A{row}"] = "Período:"
    ws[f"B{row}"] = f"{nomina.periodo_inicio.strftime('%d/%m/%Y')} - {nomina.periodo_fin.strftime('%d/%m/%Y')}"
    row += 1
    ws[f"A{row}"] = "Estado:"
    ws[f"B{row}"] = nomina.estado
    row += 1
    ws[f"A{row}"] = "Generado por:"
    ws[f"B{row}"] = nomina.generado_por or ""
    row += 2

    # Table headers
    headers = ["Código Cuenta", "Descripción", "Centro de Costos", "Débitos", "Créditos"]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # Helper function to add or update accounting entry
    def add_accounting_entry(entries_dict, codigo_cuenta, centro_costos, descripcion, debito=None, credito=None):
        """Add or update an accounting entry in the entries dictionary."""
        key = (codigo_cuenta, centro_costos)
        if key not in entries_dict:
            entries_dict[key] = {
                "descripcion": descripcion,
                "debito": Decimal("0"),
                "credito": Decimal("0"),
            }
        if debito is not None:
            entries_dict[key]["debito"] += debito
        if credito is not None:
            entries_dict[key]["credito"] += credito

    # Helper function to add unique warnings
    def add_unique_warning(warnings_list, warning_message):
        """Add a warning message to the list if it's not already present."""
        if warning_message not in warnings_list:
            warnings_list.append(warning_message)

    # Collect accounting entries - summarized by (account code, cost center) combination
    # Key is (codigo_cuenta, centro_costos), value includes description and amounts
    # Structure: {(codigo_cuenta, centro_costos): {"descripcion": str, "debito": amount, "credito": amount}}
    accounting_entries = {}
    advertencias = []  # List of configuration warnings

    # 1. Salario Base (from planilla configuration)
    # Group by cost center
    salario_por_centro_costos = {}  # {centro_costos: amount}
    for ne in nomina_empleados:
        centro = ne.centro_costos_snapshot or ""
        if centro not in salario_por_centro_costos:
            salario_por_centro_costos[centro] = Decimal("0")
        salario_por_centro_costos[centro] += ne.sueldo_base_historico

    for centro_costos, total_salario in salario_por_centro_costos.items():
        if total_salario > 0:
            if planilla.codigo_cuenta_debe_salario:
                add_accounting_entry(
                    accounting_entries,
                    planilla.codigo_cuenta_debe_salario,
                    centro_costos,
                    planilla.descripcion_cuenta_debe_salario or "Salario Base",
                    debito=total_salario
                )
            else:
                add_unique_warning(advertencias, "Planilla: Falta configurar cuenta débito para salario base")

            if planilla.codigo_cuenta_haber_salario:
                add_accounting_entry(
                    accounting_entries,
                    planilla.codigo_cuenta_haber_salario,
                    centro_costos,
                    planilla.descripcion_cuenta_haber_salario or "Salario por Pagar",
                    credito=total_salario
                )
            else:
                add_unique_warning(advertencias, "Planilla: Falta configurar cuenta crédito para salario base")

    # 2. Process all detalles (percepciones, deducciones, prestaciones)
    for ne in nomina_empleados:
        centro_costos = ne.centro_costos_snapshot or ""

        detalles = (
            db.session.execute(
                db.select(NominaDetalle).filter_by(nomina_empleado_id=ne.id).order_by(NominaDetalle.orden)
            )
            .scalars()
            .all()
        )

        for detalle in detalles:
            # Get the corresponding concept to get account codes
            concepto = None
            concepto_tipo = ""
            if detalle.tipo == "ingreso" and detalle.percepcion_id:
                concepto = db.session.get(Percepcion, detalle.percepcion_id)
                concepto_tipo = "Percepción"
            elif detalle.tipo == "deduccion" and detalle.deduccion_id:
                concepto = db.session.get(Deduccion, detalle.deduccion_id)
                concepto_tipo = "Deducción"
            elif detalle.tipo == "prestacion" and detalle.prestacion_id:
                concepto = db.session.get(Prestacion, detalle.prestacion_id)
                concepto_tipo = "Prestación"
            else:
                continue

            if not concepto or not concepto.contabilizable:
                continue

            # Add debit entry with cost center
            if concepto.codigo_cuenta_debe:
                add_accounting_entry(
                    accounting_entries,
                    concepto.codigo_cuenta_debe,
                    centro_costos,
                    concepto.descripcion_cuenta_debe or detalle.descripcion or concepto.nombre,
                    debito=detalle.monto
                )
            else:
                warning = f"{concepto_tipo} '{concepto.codigo}': Falta configurar cuenta débito"
                add_unique_warning(advertencias, warning)

            # Add credit entry with cost center
            if concepto.codigo_cuenta_haber:
                add_accounting_entry(
                    accounting_entries,
                    concepto.codigo_cuenta_haber,
                    centro_costos,
                    concepto.descripcion_cuenta_haber or detalle.descripcion or concepto.nombre,
                    credito=detalle.monto
                )
            else:
                warning = f"{concepto_tipo} '{concepto.codigo}': Falta configurar cuenta crédito"
                add_unique_warning(advertencias, warning)

    # Remove duplicate warnings (already done inline, but keep for safety)
    advertencias = list(set(advertencias))

    # Write accounting entries (summarized by account code + cost center)
    total_debitos = Decimal("0")
    total_creditos = Decimal("0")
    asientos_json = []  # For database storage

    # Sort entries by account code, then by cost center
    for key in sorted(accounting_entries.keys(), key=lambda x: (x[0], x[1])):
        codigo_cuenta, centro_costos = key
        entry = accounting_entries[key]
        row += 1
        ws.cell(row=row, column=1, value=codigo_cuenta).border = border
        ws.cell(row=row, column=2, value=entry["descripcion"]).border = border
        ws.cell(row=row, column=3, value=centro_costos).border = border
        ws.cell(row=row, column=4, value=float(entry["debito"]) if entry["debito"] else "").border = border
        ws.cell(row=row, column=5, value=float(entry["credito"]) if entry["credito"] else "").border = border

        total_debitos += entry["debito"]
        total_creditos += entry["credito"]

        # Store for database
        asientos_json.append(
            {
                "codigo_cuenta": codigo_cuenta,
                "descripcion": entry["descripcion"],
                "centro_costos": centro_costos,
                "debito": float(entry["debito"]),
                "credito": float(entry["credito"]),
            }
        )

    # Totals row
    row += 1
    ws.cell(row=row, column=1, value="").border = border
    ws.cell(row=row, column=2, value="").border = border
    total_cell = ws.cell(row=row, column=3, value="TOTALES")
    total_cell.font = total_font
    total_cell.fill = total_fill
    total_cell.border = border
    debito_total = ws.cell(row=row, column=4, value=float(total_debitos))
    debito_total.font = total_font
    debito_total.fill = total_fill
    debito_total.border = border
    credito_total = ws.cell(row=row, column=5, value=float(total_creditos))
    credito_total.font = total_font
    credito_total.fill = total_fill
    credito_total.border = border

    # Balance check
    row += 2
    ws[f"A{row}"] = "Balance:"
    balance = total_debitos - total_creditos
    balance_cell = ws[f"B{row}"]
    balance_cell.value = float(balance)
    if abs(balance) < 0.01:  # Close to zero
        balance_cell.font = Font(bold=True, color="008000")  # Green
        ws[f"C{row}"] = "✓ Balanceado"
    else:
        balance_cell.font = Font(bold=True, color="FF0000")  # Red
        ws[f"C{row}"] = "⚠ Desbalanceado"

    # Save comprobante to database for audit trail
    comprobante = db.session.execute(db.select(ComprobanteContable).filter_by(nomina_id=nomina_id)).scalar_one_or_none()

    if comprobante:
        # Update existing
        comprobante.asientos_contables = asientos_json
        comprobante.total_debitos = total_debitos
        comprobante.total_creditos = total_creditos
        comprobante.balance = balance
        comprobante.advertencias = advertencias
        comprobante.modificado_por = current_user.usuario
    else:
        # Create new
        comprobante = ComprobanteContable(
            nomina_id=nomina_id,
            asientos_contables=asientos_json,
            total_debitos=total_debitos,
            total_creditos=total_creditos,
            balance=balance,
            advertencias=advertencias,
            creado_por=current_user.usuario,
        )
        db.session.add(comprobante)

    db.session.commit()

    # Auto-adjust column widths
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Sanitize planilla name for filename
    safe_planilla_name = re.sub(r'[^\w\s-]', '', planilla.nombre).strip().replace(' ', '_')
    filename = f"comprobante_{safe_planilla_name}_{nomina.periodo_inicio.strftime('%Y%m%d')}_{nomina.id[:8]}.xlsx"

    # Save warnings to nomina log
    if advertencias:
        # Initialize log if needed
        if not nomina.log_procesamiento:
            nomina.log_procesamiento = []

        # Add warnings to log
        for adv in advertencias:
            log_entry = {
                "timestamp": datetime.now(timezone.utc).isoformat(),
                "tipo": "advertencia_contabilidad",
                "mensaje": adv,
            }
            nomina.log_procesamiento.append(log_entry)

        db.session.commit()

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@planilla_bp.route("/<planilla_id>/nomina/<nomina_id>/exportar-comprobante-detallado-excel")
@login_required
def exportar_comprobante_detallado_excel(planilla_id: str, nomina_id: str):
    """Export detailed accounting voucher per employee to Excel."""
    from io import BytesIO
    from flask import send_file
    from coati_payroll.model import Nomina, NominaEmpleado, NominaDetalle
    from decimal import Decimal

    openpyxl_classes = _check_openpyxl_available()
    if not openpyxl_classes:
        flash(_("Excel export no disponible. Instale openpyxl."), "warning")
        return redirect(url_for("planilla.ver_nomina", planilla_id=planilla_id, nomina_id=nomina_id))

    Workbook, Font, Alignment, PatternFill, Border, Side = openpyxl_classes

    planilla = db.get_or_404(Planilla, planilla_id)
    nomina = db.get_or_404(Nomina, nomina_id)

    if nomina.planilla_id != planilla_id:
        flash(_("La nómina no pertenece a esta planilla."), "error")
        return redirect(url_for("planilla.listar_nominas", planilla_id=planilla_id))

    # Get all nomina employees
    nomina_empleados = db.session.execute(
        db.select(NominaEmpleado)
        .filter_by(nomina_id=nomina_id)
        .join(Empleado)
        .order_by(Empleado.codigo_empleado)
    ).scalars().all()

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Detalle por Empleado"

    # Define styles
    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    subheader_font = Font(bold=True, size=11)
    subheader_fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
    total_font = Font(bold=True, size=11)
    total_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Title
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = "COMPROBANTE CONTABLE DETALLADO - POR EMPLEADO"
    title_cell.font = header_font
    title_cell.fill = header_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Header info
    row = 3
    # Company information if linked
    if planilla.empresa_id and planilla.empresa:
        ws[f"A{row}"] = "Empresa:"
        ws[f"B{row}"] = planilla.empresa.razon_social
        row += 1
        if planilla.empresa.ruc:
            ws[f"A{row}"] = "RUC:"
            ws[f"B{row}"] = planilla.empresa.ruc
            row += 1
    ws[f"A{row}"] = "Planilla:"
    ws[f"B{row}"] = planilla.nombre
    row += 1
    ws[f"A{row}"] = "Período:"
    ws[f"B{row}"] = f"{nomina.periodo_inicio.strftime('%d/%m/%Y')} - {nomina.periodo_fin.strftime('%d/%m/%Y')}"
    row += 1
    ws[f"A{row}"] = "Estado:"
    ws[f"B{row}"] = nomina.estado
    row += 2

    # Table headers
    headers = ["Código Empleado", "Código Cuenta", "Descripción", "Centro de Costos", "Débitos", "Créditos"]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # Collect detailed entries per employee
    total_debitos = Decimal("0")
    total_creditos = Decimal("0")

    for ne in nomina_empleados:
        empleado = ne.empleado
        centro_costos = ne.centro_costos_snapshot or ""

        # 1. Salario Base
        if ne.sueldo_base_historico > 0:
            # Debit entry
            if planilla.codigo_cuenta_debe_salario:
                row += 1
                ws.cell(row=row, column=1, value=empleado.codigo_empleado).border = border
                ws.cell(row=row, column=2, value=planilla.codigo_cuenta_debe_salario).border = border
                desc_debe = planilla.descripcion_cuenta_debe_salario or "Salario Base"
                ws.cell(row=row, column=3, value=desc_debe).border = border
                ws.cell(row=row, column=4, value=centro_costos).border = border
                ws.cell(row=row, column=5, value=float(ne.sueldo_base_historico)).border = border
                ws.cell(row=row, column=6, value="").border = border
                total_debitos += ne.sueldo_base_historico

            # Credit entry
            if planilla.codigo_cuenta_haber_salario:
                row += 1
                ws.cell(row=row, column=1, value=empleado.codigo_empleado).border = border
                ws.cell(row=row, column=2, value=planilla.codigo_cuenta_haber_salario).border = border
                desc_haber = planilla.descripcion_cuenta_haber_salario or "Salario por Pagar"
                ws.cell(row=row, column=3, value=desc_haber).border = border
                ws.cell(row=row, column=4, value=centro_costos).border = border
                ws.cell(row=row, column=5, value="").border = border
                ws.cell(row=row, column=6, value=float(ne.sueldo_base_historico)).border = border
                total_creditos += ne.sueldo_base_historico

        # 2. Process all detalles
        detalles = (
            db.session.execute(
                db.select(NominaDetalle).filter_by(nomina_empleado_id=ne.id).order_by(NominaDetalle.orden)
            )
            .scalars()
            .all()
        )

        for detalle in detalles:
            # Get the corresponding concept to get account codes
            concepto = None
            if detalle.tipo == "ingreso" and detalle.percepcion_id:
                concepto = db.session.get(Percepcion, detalle.percepcion_id)
            elif detalle.tipo == "deduccion" and detalle.deduccion_id:
                concepto = db.session.get(Deduccion, detalle.deduccion_id)
            elif detalle.tipo == "prestacion" and detalle.prestacion_id:
                concepto = db.session.get(Prestacion, detalle.prestacion_id)
            else:
                continue

            if not concepto or not concepto.contabilizable:
                continue

            # Add debit entry
            if concepto.codigo_cuenta_debe:
                row += 1
                ws.cell(row=row, column=1, value=empleado.codigo_empleado).border = border
                ws.cell(row=row, column=2, value=concepto.codigo_cuenta_debe).border = border
                desc_debe_det = concepto.descripcion_cuenta_debe or detalle.descripcion or concepto.nombre
                ws.cell(row=row, column=3, value=desc_debe_det).border = border
                ws.cell(row=row, column=4, value=centro_costos).border = border
                ws.cell(row=row, column=5, value=float(detalle.monto)).border = border
                ws.cell(row=row, column=6, value="").border = border
                total_debitos += detalle.monto

            # Add credit entry
            if concepto.codigo_cuenta_haber:
                row += 1
                ws.cell(row=row, column=1, value=empleado.codigo_empleado).border = border
                ws.cell(row=row, column=2, value=concepto.codigo_cuenta_haber).border = border
                desc_haber_det = concepto.descripcion_cuenta_haber or detalle.descripcion or concepto.nombre
                ws.cell(row=row, column=3, value=desc_haber_det).border = border
                ws.cell(row=row, column=4, value=centro_costos).border = border
                ws.cell(row=row, column=5, value="").border = border
                ws.cell(row=row, column=6, value=float(detalle.monto)).border = border
                total_creditos += detalle.monto

    # Totals row
    row += 1
    ws.cell(row=row, column=1, value="").border = border
    ws.cell(row=row, column=2, value="").border = border
    ws.cell(row=row, column=3, value="").border = border
    total_cell = ws.cell(row=row, column=4, value="TOTALES")
    total_cell.font = total_font
    total_cell.fill = total_fill
    total_cell.border = border
    debito_total = ws.cell(row=row, column=5, value=float(total_debitos))
    debito_total.font = total_font
    debito_total.fill = total_fill
    debito_total.border = border
    credito_total = ws.cell(row=row, column=6, value=float(total_creditos))
    credito_total.font = total_font
    credito_total.fill = total_fill
    credito_total.border = border

    # Balance check
    row += 2
    ws[f"A{row}"] = "Balance:"
    balance = total_debitos - total_creditos
    balance_cell = ws[f"B{row}"]
    balance_cell.value = float(balance)
    if abs(balance) < 0.01:  # Close to zero
        balance_cell.font = Font(bold=True, color="008000")  # Green
        ws[f"C{row}"] = "✓ Balanceado"
    else:
        balance_cell.font = Font(bold=True, color="FF0000")  # Red
        ws[f"C{row}"] = "⚠ Desbalanceado"

    # Auto-adjust column widths
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Sanitize planilla name for filename
    safe_planilla_name = re.sub(r'[^\w\s-]', '', planilla.nombre).strip().replace(' ', '_')
    date_str = nomina.periodo_inicio.strftime('%Y%m%d')
    filename = f"comprobante_detallado_{safe_planilla_name}_{date_str}_{nomina.id[:8]}.xlsx"

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@planilla_bp.route("/<planilla_id>/nomina/<nomina_id>/log")
@login_required
def ver_log_nomina(planilla_id: str, nomina_id: str):
    """View execution log for a nomina including warnings and errors."""
    from coati_payroll.model import Nomina

    planilla = db.get_or_404(Planilla, planilla_id)
    nomina = db.get_or_404(Nomina, nomina_id)

    if nomina.planilla_id != planilla_id:
        flash(_("La nómina no pertenece a esta planilla."), "error")
        return redirect(url_for("planilla.listar_nominas", planilla_id=planilla_id))

    # Get log entries
    log_entries = nomina.log_procesamiento or []

    # Get comprobante warnings if exists
    from coati_payroll.model import ComprobanteContable

    comprobante = db.session.execute(db.select(ComprobanteContable).filter_by(nomina_id=nomina_id)).scalar_one_or_none()

    comprobante_warnings = comprobante.advertencias if comprobante else []

    return render_template(
        "modules/planilla/log_nomina.html",
        planilla=planilla,
        nomina=nomina,
        log_entries=log_entries,
        comprobante_warnings=comprobante_warnings,
    )
