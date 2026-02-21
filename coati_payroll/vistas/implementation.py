# SPDX-License-Identifier: Apache-2.0
# SPDX-FileCopyrightText: 2025 - 2026 BMO Soluciones, S.A.
"""Implementation helpers for initial system setup."""

from __future__ import annotations

from flask import Blueprint, flash, redirect, render_template, request, url_for

from coati_payroll.enums import TipoUsuario
from coati_payroll.i18n import _
from coati_payroll.rbac import require_role
from coati_payroll.vistas.planilla.helpers import check_openpyxl_available
from coati_payroll.vistas.implementation_helpers import import_accounting_configuration_rows

implementation_bp = Blueprint("implementation", __name__, url_prefix="/settings/helpers")


@implementation_bp.route("/")
@require_role(TipoUsuario.ADMIN)
def index():
    """Display the initial implementation helpers menu."""
    return render_template("modules/settings/implementation/index.html")


@implementation_bp.route("/accounting-config/import", methods=["POST"])
@require_role(TipoUsuario.ADMIN)
def import_accounting_configuration():
    """Import accounting account configuration from an Excel template."""
    if not check_openpyxl_available():
        flash(_("Carga no disponible. Instale openpyxl."), "warning")
        return redirect(url_for("implementation.index"))

    uploaded_file = request.files.get("accounting_file")
    if uploaded_file is None or not uploaded_file.filename:
        flash(_("Debe seleccionar un archivo Excel para importar."), "error")
        return redirect(url_for("implementation.index"))

    try:
        from openpyxl import load_workbook

        workbook = load_workbook(uploaded_file, data_only=True)
        worksheet = workbook.active
        rows = [list(row) for row in worksheet.iter_rows(values_only=True)]

        result = import_accounting_configuration_rows(rows)

        flash(
            _("Importación completada: {} filas actualizadas, {} filas omitidas.").format(
                result.updated_rows, result.skipped_rows
            ),
            "success",
        )
    except ValueError as exc:
        flash(str(exc), "error")
    except Exception as exc:  # pragma: no cover - defensive fallback
        flash(_("No se pudo procesar el archivo: {}").format(str(exc)), "error")

    return redirect(url_for("implementation.index"))
