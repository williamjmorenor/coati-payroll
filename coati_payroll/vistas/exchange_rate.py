# SPDX-License-Identifier: Apache-2.0
# SPDX-FileCopyrightText: 2025 - 2026 BMO Soluciones, S.A.
# Copyright 2025 - 2026 BMO Soluciones, S.A.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#    http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
"""Exchange Rate CRUD routes."""

from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal

from flask import Blueprint, flash, redirect, render_template, request, url_for
from flask_login import current_user
from openpyxl import load_workbook

from coati_payroll.forms import ExchangeRateForm
from coati_payroll.i18n import _
from coati_payroll.rbac import require_read_access, require_write_access
from coati_payroll.model import Moneda, TipoCambio, db
from coati_payroll.vistas.constants import PER_PAGE

exchange_rate_bp = Blueprint("exchange_rate", __name__, url_prefix="/exchange_rate")


def get_currency_choices():
    """Get list of currencies for select fields."""
    currencies = db.session.execute(db.select(Moneda).filter_by(activo=True).order_by(Moneda.codigo)).scalars().all()
    return [(c.id, f"{c.codigo} - {c.nombre}") for c in currencies]


@exchange_rate_bp.route("/")
@require_read_access()
def index():
    """List all exchange rates with pagination and filters."""
    page = request.args.get("page", 1, type=int)

    # Get filter parameters
    fecha_desde = request.args.get("fecha_desde", type=str)
    fecha_hasta = request.args.get("fecha_hasta", type=str)
    moneda_origen_id = request.args.get("moneda_origen_id", type=str) if request.args.get("moneda_origen_id") else None
    moneda_destino_id = (
        request.args.get("moneda_destino_id", type=str) if request.args.get("moneda_destino_id") else None
    )

    # Build query with filters
    query = db.select(TipoCambio)

    if fecha_desde:
        query = query.filter(TipoCambio.fecha >= fecha_desde)
    if fecha_hasta:
        query = query.filter(TipoCambio.fecha <= fecha_hasta)
    if moneda_origen_id:
        query = query.filter(TipoCambio.moneda_origen_id == moneda_origen_id)
    if moneda_destino_id:
        query = query.filter(TipoCambio.moneda_destino_id == moneda_destino_id)

    query = query.order_by(TipoCambio.fecha.desc())

    pagination = db.paginate(
        query,
        page=page,
        per_page=PER_PAGE,
        error_out=False,
    )

    # Get currencies for filter dropdowns
    currencies = get_currency_choices()

    return render_template(
        "modules/exchange_rate/index.html",
        exchange_rates=pagination.items,
        pagination=pagination,
        currencies=currencies,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
        moneda_origen_id=moneda_origen_id,
        moneda_destino_id=moneda_destino_id,
    )


@exchange_rate_bp.route("/new", methods=["GET", "POST"])
@require_write_access()
def new():
    """Create a new exchange rate."""
    form = ExchangeRateForm()
    form.moneda_origen_id.choices = [("", _("Seleccionar..."))] + get_currency_choices()
    form.moneda_destino_id.choices = [("", _("Seleccionar..."))] + get_currency_choices()

    if form.validate_on_submit():
        exchange_rate = TipoCambio()
        exchange_rate.fecha = form.fecha.data
        exchange_rate.moneda_origen_id = form.moneda_origen_id.data
        exchange_rate.moneda_destino_id = form.moneda_destino_id.data
        exchange_rate.tasa = form.tasa.data
        exchange_rate.creado_por = current_user.usuario

        db.session.add(exchange_rate)
        db.session.commit()
        flash(_("Tipo de cambio creado exitosamente."), "success")
        return redirect(url_for("exchange_rate.index"))

    # Default date to today
    if not form.fecha.data:
        form.fecha.data = date.today()

    return render_template("modules/exchange_rate/form.html", form=form, title=_("Nuevo Tipo de Cambio"))


@exchange_rate_bp.route("/edit/<string:id>", methods=["GET", "POST"])
@require_write_access()
def edit(id: str):
    """Edit an existing exchange rate."""
    exchange_rate = db.session.get(TipoCambio, id)
    if not exchange_rate:
        flash(_("Tipo de cambio no encontrado."), "error")
        return redirect(url_for("exchange_rate.index"))

    form = ExchangeRateForm(obj=exchange_rate)
    form.moneda_origen_id.choices = [("", _("Seleccionar..."))] + get_currency_choices()
    form.moneda_destino_id.choices = [("", _("Seleccionar..."))] + get_currency_choices()

    if form.validate_on_submit():
        exchange_rate.fecha = form.fecha.data
        exchange_rate.moneda_origen_id = form.moneda_origen_id.data
        exchange_rate.moneda_destino_id = form.moneda_destino_id.data
        exchange_rate.tasa = form.tasa.data
        exchange_rate.modificado_por = current_user.usuario

        db.session.commit()
        flash(_("Tipo de cambio actualizado exitosamente."), "success")
        return redirect(url_for("exchange_rate.index"))

    return render_template(
        "modules/exchange_rate/form.html",
        form=form,
        title=_("Editar Tipo de Cambio"),
        exchange_rate=exchange_rate,
    )


@exchange_rate_bp.route("/delete/<string:id>", methods=["POST"])
@require_write_access()
def delete(id: str):
    """Delete an exchange rate."""
    exchange_rate = db.session.get(TipoCambio, id)
    if not exchange_rate:
        flash(_("Tipo de cambio no encontrado."), "error")
        return redirect(url_for("exchange_rate.index"))

    db.session.delete(exchange_rate)
    db.session.commit()
    flash(_("Tipo de cambio eliminado exitosamente."), "success")
    return redirect(url_for("exchange_rate.index"))


# Constants for Excel import
EXPECTED_COLUMNS = 4
MAX_ERRORS_DISPLAYED = 10


@exchange_rate_bp.route("/import", methods=["GET", "POST"])
@require_write_access()
def import_excel():
    """Import exchange rates from Excel file."""
    if request.method == "GET":
        return render_template("modules/exchange_rate/import.html")

    # Check if file is in request
    if "file" not in request.files:
        flash(_("No se seleccionó ningún archivo."), "error")
        return redirect(url_for("exchange_rate.import_excel"))

    file = request.files["file"]

    # Check if file has a name
    if file.filename == "":
        flash(_("No se seleccionó ningún archivo."), "error")
        return redirect(url_for("exchange_rate.import_excel"))

    # Check if file is Excel
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        flash(_("El archivo debe ser un archivo Excel (.xlsx o .xls)."), "error")
        return redirect(url_for("exchange_rate.import_excel"))

    try:
        # Load the workbook
        workbook = load_workbook(file, data_only=True)
        sheet = workbook.active

        # Track statistics
        imported_count = 0
        updated_count = 0
        error_count = 0
        errors = []

        # Get all active currencies for lookup
        currencies = db.session.execute(db.select(Moneda).filter_by(activo=True)).scalars().all()
        currency_map = {c.codigo.upper(): c for c in currencies}

        # Process rows (skip header)
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not any(row):  # Skip empty rows
                continue

            try:
                # Expected columns: Fecha | Moneda Base | Moneda Destino | Tipo de Cambio
                if len(row) < EXPECTED_COLUMNS:
                    errors.append(
                        _("Fila {}: formato incorrecto, se esperan {} columnas.").format(row_idx, EXPECTED_COLUMNS)
                    )
                    error_count += 1
                    continue

                fecha_val, moneda_origen_codigo, moneda_destino_codigo, tasa_val = row[0], row[1], row[2], row[3]

                # Validate fecha
                if isinstance(fecha_val, datetime):
                    fecha = fecha_val.date()
                elif isinstance(fecha_val, date):
                    fecha = fecha_val
                elif isinstance(fecha_val, str):
                    try:
                        fecha = datetime.strptime(fecha_val, "%Y-%m-%d").date()
                    except ValueError:
                        try:
                            fecha = datetime.strptime(fecha_val, "%d/%m/%Y").date()
                        except ValueError:
                            errors.append(_("Fila {}: fecha inválida '{}'.").format(row_idx, fecha_val))
                            error_count += 1
                            continue
                else:
                    errors.append(_("Fila {}: fecha inválida.").format(row_idx))
                    error_count += 1
                    continue

                # Validate moneda_origen
                if not moneda_origen_codigo:
                    errors.append(_("Fila {}: moneda origen vacía.").format(row_idx))
                    error_count += 1
                    continue

                moneda_origen_key = str(moneda_origen_codigo).strip().upper()
                moneda_origen = currency_map.get(moneda_origen_key)
                if not moneda_origen:
                    errors.append(_("Fila {}: moneda origen '{}' no encontrada.").format(row_idx, moneda_origen_codigo))
                    error_count += 1
                    continue

                # Validate moneda_destino
                if not moneda_destino_codigo:
                    errors.append(_("Fila {}: moneda destino vacía.").format(row_idx))
                    error_count += 1
                    continue

                moneda_destino_key = str(moneda_destino_codigo).strip().upper()
                moneda_destino = currency_map.get(moneda_destino_key)
                if not moneda_destino:
                    errors.append(
                        _("Fila {}: moneda destino '{}' no encontrada.").format(row_idx, moneda_destino_codigo)
                    )
                    error_count += 1
                    continue

                # Validate tasa
                try:
                    if isinstance(tasa_val, (int, float)):
                        tasa = Decimal(str(tasa_val))
                    elif isinstance(tasa_val, str):
                        tasa = Decimal(tasa_val.strip())
                    else:
                        tasa = Decimal(str(tasa_val))

                    if tasa <= 0:
                        errors.append(_("Fila {}: tasa debe ser mayor que cero.").format(row_idx))
                        error_count += 1
                        continue
                except (ValueError, TypeError):
                    errors.append(_("Fila {}: tasa inválida '{}'.").format(row_idx, tasa_val))
                    error_count += 1
                    continue

                # Check if exchange rate already exists
                existing = db.session.execute(
                    db.select(TipoCambio).filter_by(
                        fecha=fecha, moneda_origen_id=moneda_origen.id, moneda_destino_id=moneda_destino.id
                    )
                ).scalar_one_or_none()

                if existing:
                    # Update existing
                    existing.tasa = tasa
                    existing.modificado_por = current_user.usuario
                    updated_count += 1
                else:
                    # Create new
                    exchange_rate = TipoCambio()
                    exchange_rate.fecha = fecha
                    exchange_rate.moneda_origen_id = moneda_origen.id
                    exchange_rate.moneda_destino_id = moneda_destino.id
                    exchange_rate.tasa = tasa
                    exchange_rate.creado_por = current_user.usuario
                    db.session.add(exchange_rate)
                    imported_count += 1

            except Exception as e:
                errors.append(_("Fila {}: error inesperado - {}.").format(row_idx, str(e)))
                error_count += 1
                continue

        # Commit all changes
        db.session.commit()

        # Show results
        if imported_count > 0 or updated_count > 0:
            flash(
                _("Importación completada: {} creados, {} actualizados.").format(imported_count, updated_count),
                "success",
            )

        if error_count > 0:
            flash(_("{} errores encontrados durante la importación.").format(error_count), "warning")
            for error in errors[:MAX_ERRORS_DISPLAYED]:
                flash(error, "error")
            if len(errors) > MAX_ERRORS_DISPLAYED:
                flash(_("... y {} errores más.").format(len(errors) - MAX_ERRORS_DISPLAYED), "error")

        return redirect(url_for("exchange_rate.index"))

    except Exception as e:
        db.session.rollback()
        flash(_("Error al procesar el archivo: {}.").format(str(e)), "error")
        return redirect(url_for("exchange_rate.import_excel"))
