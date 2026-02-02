# SPDX-License-Identifier: Apache-2.0
# SPDX-FileCopyrightText: 2025 - 2026 BMO Soluciones, S.A.
"""Excel helper functions for planilla views."""


def check_openpyxl_available():
    """Check if openpyxl is available and return necessary classes.

    Returns:
        tuple: (Workbook, Font, Alignment, PatternFill, Border, Side) or None if not available
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        return Workbook, Font, Alignment, PatternFill, Border, Side
    except ImportError:
        return None
