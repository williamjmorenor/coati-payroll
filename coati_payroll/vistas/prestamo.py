# SPDX-License-Identifier: Apache-2.0
# SPDX-FileCopyrightText: 2025 - 2026 BMO Soluciones, S.A.
"""Views for managing loans and salary advances (Préstamos y Adelantos).

This module handles:
- Creating loan/advance requests
- Approving/rejecting loans
- Viewing payment schedules
- Exporting payment tables to Excel/PDF
- Tracking loan balances and payments
"""

from __future__ import annotations

from datetime import date
from decimal import Decimal
from io import BytesIO
from typing import Any, cast

from flask import (
    Blueprint,
    flash,
    redirect,
    render_template,
    request,
    url_for,
    send_file,
    Response,
)
from flask_login import current_user

from coati_payroll.model import (
    db,
    Adelanto,
    Empleado,
    Moneda,
    Deduccion,
    AdelantoAbono,
)
from coati_payroll.forms import (
    PrestamoForm,
    PrestamoApprovalForm,
    PagoExtraordinarioForm,
    CondonacionForm,
)
from coati_payroll.i18n import _
from coati_payroll.enums import AdelantoEstado, AdelantoTipo
from coati_payroll.rbac import require_read_access, require_write_access

prestamo_bp = Blueprint("prestamo", __name__, url_prefix="/prestamo")


@prestamo_bp.route("/")
@require_read_access()
def index():
    """List all loans and advances with filtering options."""
    # Get filter parameters
    empleado_id = request.args.get("empleado_id", "")
    estado = request.args.get("estado", "")
    tipo = request.args.get("tipo", "")

    # Build query
    query = db.select(Adelanto).join(Empleado)

    if empleado_id:
        query = query.filter(Adelanto.empleado_id == empleado_id)
    if estado:
        query = query.filter(Adelanto.estado == estado)
    if tipo:
        query = query.filter(Adelanto.tipo == tipo)

    # Order by most recent first
    query = query.order_by(Adelanto.fecha_solicitud.desc())

    prestamos = db.session.execute(query).scalars().all()

    # Get all employees for filter dropdown
    empleados = (
        db.session.execute(db.select(Empleado).filter_by(activo=True).order_by(Empleado.primer_apellido))
        .scalars()
        .all()
    )

    return render_template(
        "modules/prestamo/index.html",
        prestamos=prestamos,
        empleados=empleados,
        filtro_empleado=empleado_id,
        filtro_estado=estado,
        filtro_tipo=tipo,
        estados=AdelantoEstado,
        tipos=AdelantoTipo,
    )


@prestamo_bp.route("/new", methods=["GET", "POST"])
@require_write_access()
def new():
    """Create a new loan or salary advance."""
    form = PrestamoForm()

    # Populate select fields
    empleados = (
        db.session.execute(db.select(Empleado).filter_by(activo=True).order_by(Empleado.primer_apellido))
        .scalars()
        .all()
    )
    form.empleado_id.choices = [
        (emp.id, f"{emp.primer_nombre} {emp.primer_apellido} - {emp.codigo_empleado}") for emp in empleados
    ]

    monedas = db.session.execute(db.select(Moneda).filter_by(activo=True).order_by(Moneda.nombre)).scalars().all()
    form.moneda_id.choices = [(m.id, f"{m.nombre} ({m.codigo})") for m in monedas]

    deducciones = (
        db.session.execute(db.select(Deduccion).filter_by(activo=True).order_by(Deduccion.nombre)).scalars().all()
    )
    form.deduccion_id.choices = [("", _("-- Sin deducción asociada --"))] + [(d.id, d.nombre) for d in deducciones]

    if form.validate_on_submit():
        prestamo = Adelanto()
        prestamo.empleado_id = form.empleado_id.data
        prestamo.tipo = form.tipo.data
        prestamo.fecha_solicitud = form.fecha_solicitud.data
        prestamo.monto_solicitado = form.monto_solicitado.data
        prestamo.moneda_id = form.moneda_id.data
        prestamo.cuotas_pactadas = form.cuotas_pactadas.data
        prestamo.tasa_interes = form.tasa_interes.data or Decimal("0.0000")
        prestamo.tipo_interes = form.tipo_interes.data
        prestamo.metodo_amortizacion = form.metodo_amortizacion.data
        prestamo.cuenta_debe = form.cuenta_debe.data
        prestamo.cuenta_haber = form.cuenta_haber.data
        prestamo.motivo = form.motivo.data
        prestamo.estado = AdelantoEstado.BORRADOR
        prestamo.saldo_pendiente = Decimal("0.00")  # Will be set upon approval
        prestamo.creado_por = current_user.usuario

        # Set deduccion_id only if a valid value was selected
        if form.deduccion_id.data:
            prestamo.deduccion_id = form.deduccion_id.data

        db.session.add(prestamo)
        db.session.commit()

        flash(
            _("Préstamo/Adelanto creado exitosamente. Estado: Borrador."),
            "success",
        )
        return redirect(url_for("prestamo.detail", prestamo_id=prestamo.id))

    # Set default values
    if request.method == "GET":
        form.fecha_solicitud.data = date.today()
        form.tipo_interes.data = "ninguno"
        form.tasa_interes.data = Decimal("0.0000")
        form.metodo_amortizacion.data = "frances"

    return render_template("modules/prestamo/form.html", form=form, prestamo=None)


@prestamo_bp.route("/<prestamo_id>")
@require_read_access()
def detail(prestamo_id):
    """View loan details including payment schedule."""
    prestamo = db.session.get(Adelanto, prestamo_id)
    if not prestamo:
        flash("Préstamo no encontrado.", "danger")
        return redirect(url_for("prestamo.index"))

    # Touch relationship to ensure it is loaded before rendering
    prestamo.empleado

    # Generate payment schedule
    tabla_pago = generar_tabla_pago(prestamo)

    # Get payment history
    abonos = (
        db.session.execute(
            db.select(AdelantoAbono).filter_by(adelanto_id=prestamo_id).order_by(AdelantoAbono.fecha_abono.desc())
        )
        .scalars()
        .all()
    )

    # Get interest journal if loan has interest
    from coati_payroll.model import InteresAdelanto

    intereses = []
    if prestamo.tasa_interes and prestamo.tasa_interes > 0:
        intereses = (
            db.session.execute(
                db.select(InteresAdelanto)
                .filter_by(adelanto_id=prestamo_id)
                .order_by(InteresAdelanto.fecha_hasta.desc())
            )
            .scalars()
            .all()
        )

    return render_template(
        "modules/prestamo/detail.html",
        prestamo=prestamo,
        tabla_pago=tabla_pago,
        abonos=abonos,
        intereses=intereses,
    )


@prestamo_bp.route("/<prestamo_id>/edit", methods=["GET", "POST"])
@require_write_access()
def edit(prestamo_id):
    """Edit a loan (only allowed in draft or pending state)."""
    prestamo = db.session.get(Adelanto, prestamo_id)
    if not prestamo:
        flash(_("Préstamo no encontrado."), "danger")
        return redirect(url_for("prestamo.index"))

    # Only allow editing in draft or pending state
    if prestamo.estado not in [AdelantoEstado.BORRADOR, AdelantoEstado.PENDIENTE]:
        flash(
            _("No se puede editar un préstamo en estado '{estado}'.".format(estado=prestamo.estado)),
            "warning",
        )
        return redirect(url_for("prestamo.detail", prestamo_id=prestamo_id))

    form = PrestamoForm(obj=prestamo)

    # Populate select fields
    empleados = (
        db.session.execute(db.select(Empleado).filter_by(activo=True).order_by(Empleado.primer_apellido))
        .scalars()
        .all()
    )
    form.empleado_id.choices = [
        (emp.id, f"{emp.primer_nombre} {emp.primer_apellido} - {emp.codigo_empleado}") for emp in empleados
    ]

    monedas = db.session.execute(db.select(Moneda).filter_by(activo=True).order_by(Moneda.nombre)).scalars().all()
    form.moneda_id.choices = [(m.id, f"{m.nombre} ({m.codigo})") for m in monedas]

    deducciones = (
        db.session.execute(db.select(Deduccion).filter_by(activo=True).order_by(Deduccion.nombre)).scalars().all()
    )
    form.deduccion_id.choices = [("", _("-- Sin deducción asociada --"))] + [(d.id, d.nombre) for d in deducciones]

    if form.validate_on_submit():
        prestamo.empleado_id = form.empleado_id.data
        prestamo.tipo = form.tipo.data
        prestamo.fecha_solicitud = form.fecha_solicitud.data
        prestamo.monto_solicitado = form.monto_solicitado.data
        prestamo.moneda_id = form.moneda_id.data
        prestamo.cuotas_pactadas = form.cuotas_pactadas.data
        prestamo.tasa_interes = form.tasa_interes.data or Decimal("0.0000")
        prestamo.tipo_interes = form.tipo_interes.data
        prestamo.metodo_amortizacion = form.metodo_amortizacion.data
        prestamo.cuenta_debe = form.cuenta_debe.data
        prestamo.cuenta_haber = form.cuenta_haber.data
        prestamo.motivo = form.motivo.data
        prestamo.modificado_por = current_user.usuario

        # Set deduccion_id
        if form.deduccion_id.data:
            prestamo.deduccion_id = form.deduccion_id.data
        else:
            prestamo.deduccion_id = None

        db.session.commit()

        flash(_("Préstamo/Adelanto actualizado exitosamente."), "success")
        return redirect(url_for("prestamo.detail", prestamo_id=prestamo_id))

    return render_template("modules/prestamo/form.html", form=form, prestamo=prestamo)


@prestamo_bp.route("/<prestamo_id>/submit", methods=["POST"])
@require_write_access()
def submit(prestamo_id):
    """Submit a loan for approval (change from draft to pending)."""
    prestamo = db.session.get(Adelanto, prestamo_id)
    if not prestamo:
        flash(_("Préstamo no encontrado."), "danger")
        return redirect(url_for("prestamo.index"))

    if prestamo.estado != AdelantoEstado.BORRADOR:
        flash(_("Solo los préstamos en borrador pueden ser enviados."), "warning")
        return redirect(url_for("prestamo.detail", prestamo_id=prestamo_id))

    prestamo.estado = AdelantoEstado.PENDIENTE
    prestamo.modificado_por = current_user.usuario
    db.session.commit()

    flash(_("Préstamo enviado para aprobación."), "success")
    return redirect(url_for("prestamo.detail", prestamo_id=prestamo_id))


@prestamo_bp.route("/<prestamo_id>/approve", methods=["GET", "POST"])
@require_write_access()
def approve(prestamo_id):
    """Approve a loan and set it as active."""
    prestamo = db.session.get(Adelanto, prestamo_id)
    if not prestamo:
        flash(_("Préstamo no encontrado."), "danger")
        return redirect(url_for("prestamo.index"))

    if prestamo.estado not in [AdelantoEstado.PENDIENTE, AdelantoEstado.BORRADOR]:
        flash(_("Este préstamo no puede ser aprobado en su estado actual."), "warning")
        return redirect(url_for("prestamo.detail", prestamo_id=prestamo_id))

    form = PrestamoApprovalForm()

    if form.validate_on_submit():
        if form.aprobar.data:
            # Approve the loan
            prestamo.monto_aprobado = form.monto_aprobado.data
            prestamo.fecha_aprobacion = form.fecha_aprobacion.data
            prestamo.fecha_desembolso = form.fecha_desembolso.data
            prestamo.estado = AdelantoEstado.APROBADO
            prestamo.aprobado_por = current_user.usuario

            # Calculate installment amount based on amortization method
            if prestamo.cuotas_pactadas and prestamo.cuotas_pactadas > 0:
                from coati_payroll.interes_engine import calcular_cuota_frances

                tasa_interes = prestamo.tasa_interes or Decimal("0.0000")
                metodo = prestamo.metodo_amortizacion or "frances"

                # For French method, calculate constant payment
                # For German method, payment varies so we store the first payment
                if metodo == "frances":
                    prestamo.monto_por_cuota = calcular_cuota_frances(
                        prestamo.monto_aprobado, tasa_interes, prestamo.cuotas_pactadas
                    )
                else:
                    # For German method, store average payment for reference
                    # Actual payment will be calculated per installment
                    prestamo.monto_por_cuota = prestamo.monto_aprobado / prestamo.cuotas_pactadas
            else:
                prestamo.monto_por_cuota = prestamo.monto_aprobado

            # Set pending balance and initialize interest tracking
            prestamo.saldo_pendiente = prestamo.monto_aprobado
            prestamo.interes_acumulado = Decimal("0.00")
            prestamo.fecha_ultimo_calculo_interes = prestamo.fecha_aprobacion or date.today()
            prestamo.modificado_por = current_user.usuario

            db.session.commit()
            flash(_("Préstamo aprobado exitosamente."), "success")

        elif form.rechazar.data:
            # Reject the loan
            prestamo.estado = AdelantoEstado.RECHAZADO
            prestamo.rechazado_por = current_user.usuario
            prestamo.motivo_rechazo = form.motivo_rechazo.data
            prestamo.modificado_por = current_user.usuario

            db.session.commit()
            flash(_("Préstamo rechazado."), "info")

        return redirect(url_for("prestamo.detail", prestamo_id=prestamo_id))

    # Pre-populate form with loan data
    if request.method == "GET":
        form.monto_aprobado.data = prestamo.monto_solicitado
        form.fecha_aprobacion.data = date.today()

    return render_template("modules/prestamo/approve.html", form=form, prestamo=prestamo)


@prestamo_bp.route("/<prestamo_id>/cancel", methods=["POST"])
@require_write_access()
def cancel(prestamo_id):
    """Cancel a loan."""
    prestamo = db.session.get(Adelanto, prestamo_id)
    if not prestamo:
        flash(_("Préstamo no encontrado."), "danger")
        return redirect(url_for("prestamo.index"))

    if prestamo.estado in [AdelantoEstado.PAGADO, AdelantoEstado.CANCELADO]:
        flash(_("Este préstamo ya está finalizado."), "warning")
        return redirect(url_for("prestamo.detail", prestamo_id=prestamo_id))

    prestamo.estado = AdelantoEstado.CANCELADO
    prestamo.modificado_por = current_user.usuario
    db.session.commit()

    flash(_("Préstamo cancelado."), "info")
    return redirect(url_for("prestamo.detail", prestamo_id=prestamo_id))


@prestamo_bp.route("/<prestamo_id>/pago-extraordinario", methods=["GET", "POST"])
@require_write_access()
def pago_extraordinario(prestamo_id):
    """Register an extraordinary/manual payment on a loan."""
    prestamo = db.session.get(Adelanto, prestamo_id)
    if not prestamo:
        flash("Préstamo no encontrado.", "danger")
        return redirect(url_for("prestamo.index"))

    # Touch relationship to ensure it is loaded before rendering
    prestamo.empleado

    # Only allow payments on approved/active loans
    if prestamo.estado not in [AdelantoEstado.APROBADO, AdelantoEstado.APLICADO]:
        flash(
            _("Solo se pueden registrar pagos en préstamos aprobados o aplicados."),
            "warning",
        )
        return redirect(url_for("prestamo.detail", prestamo_id=prestamo_id))

    if prestamo.saldo_pendiente <= 0:
        flash(_("Este préstamo ya está totalmente pagado."), "info")
        return redirect(url_for("prestamo.detail", prestamo_id=prestamo_id))

    form = PagoExtraordinarioForm()

    if form.validate_on_submit():
        monto_abonado = form.monto_abonado.data

        # Validate payment amount
        if monto_abonado > prestamo.saldo_pendiente:
            flash(
                _("El monto del pago ({monto}) excede el saldo pendiente ({saldo}).").format(
                    monto=monto_abonado, saldo=prestamo.saldo_pendiente
                ),
                "warning",
            )
            return render_template(
                "modules/prestamo/pago_extraordinario.html",
                form=form,
                prestamo=prestamo,
            )

        # Record the payment
        abono = AdelantoAbono()
        abono.adelanto_id = prestamo.id
        abono.fecha_abono = form.fecha_abono.data
        abono.monto_abonado = monto_abonado
        abono.saldo_anterior = prestamo.saldo_pendiente
        abono.saldo_posterior = prestamo.saldo_pendiente - monto_abonado
        abono.tipo_abono = "manual"
        abono.observaciones = form.observaciones.data or "Pago extraordinario"
        # Audit trail information
        abono.tipo_comprobante = form.tipo_comprobante.data
        abono.numero_comprobante = form.numero_comprobante.data
        abono.referencia_bancaria = form.referencia_bancaria.data
        # Optional accounting entries
        abono.cuenta_debe = form.cuenta_debe.data
        abono.cuenta_haber = form.cuenta_haber.data
        abono.creado_por = current_user.usuario

        # Update loan balance
        prestamo.saldo_pendiente = abono.saldo_posterior
        prestamo.modificado_por = current_user.usuario

        # Apply payment according to selected method
        tipo_aplicacion = form.tipo_aplicacion.data

        # Calculate remaining installments (those not yet paid)
        total_abonado_previo = sum(a.monto_abonado for a in prestamo.abonos if a.tipo_abono in ["nomina", "manual"])
        cuotas_pagadas = 0
        if prestamo.monto_por_cuota and prestamo.monto_por_cuota > 0:
            cuotas_pagadas = int(total_abonado_previo / prestamo.monto_por_cuota)

        cuotas_pendientes = prestamo.cuotas_pactadas - cuotas_pagadas

        if tipo_aplicacion == "reducir_cuotas":
            # Option 1: Reduce number of installments, keep installment amount
            if prestamo.monto_por_cuota and prestamo.monto_por_cuota > 0:
                cuotas_a_eliminar = int(monto_abonado / prestamo.monto_por_cuota)
                # Store original values for the observation
                cuotas_originales = prestamo.cuotas_pactadas
                # Adjust total installments
                nueva_cuotas_pactadas = max(cuotas_pagadas, prestamo.cuotas_pactadas - cuotas_a_eliminar)
                prestamo.cuotas_pactadas = nueva_cuotas_pactadas

                abono.observaciones = (
                    f"{abono.observaciones or 'Pago extraordinario'} - "
                    f"Cuotas reducidas de {cuotas_originales} a {nueva_cuotas_pactadas}"
                )

        elif tipo_aplicacion == "reducir_monto":
            # Option 2: Reduce installment amount, keep number of installments
            if cuotas_pendientes > 0:
                # Recalculate installment amount based on remaining balance
                nueva_cuota = prestamo.saldo_pendiente / cuotas_pendientes
                monto_original = prestamo.monto_por_cuota
                prestamo.monto_por_cuota = nueva_cuota

                abono.observaciones = (
                    f"{abono.observaciones or 'Pago extraordinario'} - "
                    f"Monto por cuota reducido de {monto_original:.2f} a {nueva_cuota:.2f}"
                )

        # Check if loan is fully paid
        if prestamo.saldo_pendiente <= Decimal("0.01"):  # Allow for rounding
            prestamo.saldo_pendiente = Decimal("0.00")
            prestamo.estado = AdelantoEstado.PAGADO

        db.session.add(abono)
        db.session.commit()

        flash(
            _("Pago extraordinario registrado exitosamente."),
            "success",
        )
        return redirect(url_for("prestamo.detail", prestamo_id=prestamo_id))

    # Pre-populate form
    if request.method == "GET":
        form.fecha_abono.data = date.today()
        # Default to reducing installment amount (usually more beneficial for employee)
        form.tipo_aplicacion.data = "reducir_monto"

    return render_template(
        "modules/prestamo/pago_extraordinario.html",
        form=form,
        prestamo=prestamo,
    )


@prestamo_bp.route("/<prestamo_id>/condonacion", methods=["GET", "POST"])
@require_write_access()
def condonacion(prestamo_id):
    """Record a loan forgiveness/write-off (condonación de deuda)."""
    prestamo = db.session.get(Adelanto, prestamo_id)
    if not prestamo:
        flash(_("Préstamo no encontrado."), "danger")
        return redirect(url_for("prestamo.index"))

    # Only allow forgiveness on approved/active loans
    if prestamo.estado not in [AdelantoEstado.APROBADO, AdelantoEstado.APLICADO]:
        flash(
            _("Solo se pueden condonar préstamos aprobados o aplicados."),
            "warning",
        )
        return redirect(url_for("prestamo.detail", prestamo_id=prestamo_id))

    if prestamo.saldo_pendiente <= 0:
        flash(_("Este préstamo ya está totalmente pagado."), "info")
        return redirect(url_for("prestamo.detail", prestamo_id=prestamo_id))

    form = CondonacionForm()

    if form.validate_on_submit():
        monto_condonado = form.monto_condonado.data

        # Validate forgiveness amount
        if monto_condonado > prestamo.saldo_pendiente:
            flash(
                _("El monto a condonar ({monto}) excede el saldo pendiente ({saldo}).").format(
                    monto=monto_condonado, saldo=prestamo.saldo_pendiente
                ),
                "warning",
            )
            return render_template(
                "modules/prestamo/condonacion.html",
                form=form,
                prestamo=prestamo,
            )

        # Record the forgiveness as a special type of payment
        abono = AdelantoAbono()
        abono.adelanto_id = prestamo.id
        abono.fecha_abono = form.fecha_condonacion.data
        abono.monto_abonado = monto_condonado
        abono.saldo_anterior = prestamo.saldo_pendiente
        abono.saldo_posterior = prestamo.saldo_pendiente - monto_condonado
        abono.tipo_abono = "condonacion"

        # Store complete audit trail
        abono.autorizado_por = form.autorizado_por.data
        abono.documento_soporte = form.documento_soporte.data
        abono.referencia_documento = form.referencia_documento.data
        abono.justificacion = form.justificacion.data
        # Optional accounting entries
        abono.cuenta_debe = form.cuenta_debe.data
        abono.cuenta_haber = form.cuenta_haber.data

        # Build observation summary
        porcentaje = ""
        if form.porcentaje_condonado.data:
            porcentaje = f" ({form.porcentaje_condonado.data}%)"

        abono.observaciones = (
            f"CONDONACIÓN DE DEUDA{porcentaje} - "
            f"Autorizado por: {form.autorizado_por.data}. "
            f"Documento: {form.documento_soporte.data} - {form.referencia_documento.data}"
        )
        abono.creado_por = current_user.usuario

        # Update loan balance
        prestamo.saldo_pendiente = abono.saldo_posterior
        prestamo.modificado_por = current_user.usuario

        # If loan is fully forgiven/paid, mark as paid
        if prestamo.saldo_pendiente <= Decimal("0.01"):  # Allow for rounding
            prestamo.saldo_pendiente = Decimal("0.00")
            prestamo.estado = AdelantoEstado.PAGADO

        db.session.add(abono)
        db.session.commit()

        flash(
            _("Condonación de deuda registrada exitosamente. Monto condonado: {monto}").format(monto=monto_condonado),
            "success",
        )
        return redirect(url_for("prestamo.detail", prestamo_id=prestamo_id))

    # Pre-populate form
    if request.method == "GET":
        form.fecha_condonacion.data = date.today()
        form.monto_condonado.data = prestamo.saldo_pendiente

    return render_template(
        "modules/prestamo/condonacion.html",
        form=form,
        prestamo=prestamo,
    )


@prestamo_bp.route("/<prestamo_id>/tabla-pago/excel")
@require_read_access()
def export_excel(prestamo_id):
    """Export payment schedule to Excel."""
    prestamo = db.session.get(Adelanto, prestamo_id)
    if not prestamo:
        flash(_("Préstamo no encontrado."), "danger")
        return redirect(url_for("prestamo.index"))

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        flash(
            _("Excel export no disponible. Instale openpyxl."),
            "warning",
        )
        return redirect(url_for("prestamo.detail", prestamo_id=prestamo_id))

    # Generate payment schedule
    tabla_pago = generar_tabla_pago(prestamo)

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Tabla de Pagos"

    # Header
    ws["A1"] = "TABLA DE PAGOS - PRÉSTAMO/ADELANTO"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:E1")

    # Loan details
    row = 3
    ws[f"A{row}"] = "Empleado:"
    ws[f"B{row}"] = f"{prestamo.empleado.primer_nombre} {prestamo.empleado.primer_apellido}"
    row += 1
    ws[f"A{row}"] = "Tipo:"
    ws[f"B{row}"] = prestamo.tipo
    row += 1
    ws[f"A{row}"] = "Monto:"
    ws[f"B{row}"] = float(prestamo.monto_aprobado or prestamo.monto_solicitado)
    row += 1
    ws[f"A{row}"] = "Cuotas:"
    ws[f"B{row}"] = prestamo.cuotas_pactadas
    row += 2

    # Table header
    headers = ["#", "Fecha Estimada", "Cuota", "Interés", "Capital", "Saldo"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")

    # Table data
    for item in tabla_pago:
        row += 1
        ws.cell(row=row, column=1, value=item["numero"])
        ws.cell(row=row, column=2, value=item["fecha_estimada"])
        ws.cell(row=row, column=3, value=float(item["cuota"]))
        ws.cell(row=row, column=4, value=float(item["interes"]))
        ws.cell(row=row, column=5, value=float(item["capital"]))
        ws.cell(row=row, column=6, value=float(item["saldo"]))

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Use first 8 chars of ULID (26 chars total) for short filename
    filename = f"tabla_pago_{prestamo.empleado.codigo_empleado}_{prestamo.id[:8]}.xlsx"

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@prestamo_bp.route("/<prestamo_id>/tabla-pago/pdf")
@require_read_access()
def export_pdf(prestamo_id):
    """Export payment schedule to PDF."""
    prestamo = db.session.get(Adelanto, prestamo_id)
    if not prestamo:
        flash(_("Préstamo no encontrado."), "danger")
        return redirect(url_for("prestamo.index"))

    # Generate payment schedule
    tabla_pago = generar_tabla_pago(prestamo)

    # Render HTML template for PDF
    html = render_template(
        "modules/prestamo/tabla_pago_pdf.html",
        prestamo=prestamo,
        tabla_pago=tabla_pago,
    )

    try:
        from flask_weasyprint import HTML, render_pdf

        pdf = render_pdf(HTML(string=html))
        # Use first 8 chars of ULID (26 chars total) for short filename
        filename = f"tabla_pago_{prestamo.empleado.codigo_empleado}_{prestamo.id[:8]}.pdf"

        return Response(
            pdf,
            mimetype="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    except ImportError:
        flash(
            _("PDF export no disponible. Instale WeasyPrint."),
            "warning",
        )
        return redirect(url_for("prestamo.detail", prestamo_id=prestamo_id))


def generar_tabla_pago(prestamo: Adelanto) -> list[dict]:
    """Generate payment schedule for a loan.

    Args:
        prestamo: Loan object

    Returns:
        List of payment schedule items with fields:
        - numero: Payment number
        - fecha_estimada: Estimated payment date
        - cuota: Total payment amount
        - interes: Interest portion
        - capital: Principal portion
        - saldo: Remaining balance
    """
    if not prestamo.cuotas_pactadas or prestamo.cuotas_pactadas <= 0:
        return []

    monto_base = prestamo.monto_aprobado or prestamo.monto_solicitado
    if not monto_base or monto_base <= 0:
        return []

    # Determine start date
    fecha_inicio = prestamo.fecha_aprobacion or prestamo.fecha_solicitud or date.today()

    # Import interest engine
    from coati_payroll.interes_engine import generar_tabla_amortizacion

    # Get interest rate and type
    tasa_interes = prestamo.tasa_interes or Decimal("0.0000")
    tipo_interes = prestamo.tipo_interes or "ninguno"
    metodo_amortizacion = prestamo.metodo_amortizacion or "frances"

    # Generate amortization schedule using the interest engine
    cuotas = generar_tabla_amortizacion(
        principal=monto_base,
        tasa_anual=tasa_interes,
        num_cuotas=prestamo.cuotas_pactadas,
        fecha_inicio=fecha_inicio,
        metodo=cast(Any, metodo_amortizacion),
        tipo_interes=cast(Any, tipo_interes),
    )

    # Convert to dict format for template
    tabla = []
    for cuota in cuotas:
        tabla.append(
            {
                "numero": cuota.numero,
                "fecha_estimada": cuota.fecha_estimada,
                "cuota": cuota.cuota_total,
                "interes": cuota.interes,
                "capital": cuota.capital,
                "saldo": cuota.saldo,
            }
        )

    return tabla
