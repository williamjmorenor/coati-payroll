# Copyright 2025 BMO Soluciones, S.A.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#    http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
"""Report export functionality for Excel format.

Provides utilities to export report results to Excel files with proper
formatting and metadata.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from coati_payroll.config import DIRECTORIO_APP
from coati_payroll.log import log


class ReportExporter:
    """Handles exporting report results to various formats."""

    def __init__(self, report_name: str, results: List[Dict[str, Any]]):
        """Initialize exporter.

        Args:
            report_name: Name of the report
            results: List of result dictionaries
        """
        self.report_name = report_name
        self.results = results

    def to_excel(self, output_path: Optional[str] = None) -> str:
        """Export results to Excel format.

        Args:
            output_path: Optional output file path. If not provided, generates one.

        Returns:
            Path to exported file

        Raises:
            ImportError: If openpyxl is not installed
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export. " "Install it with: pip install openpyxl")

        # Generate output path if not provided
        if not output_path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            safe_name = "".join(c for c in self.report_name if c.isalnum() or c in (" ", "_", "-")).strip()
            filename = f"{safe_name}_{timestamp}.xlsx"

            # Create exports directory if it doesn't exist
            exports_dir = Path(DIRECTORIO_APP) / "exports" / "reports"
            exports_dir.mkdir(parents=True, exist_ok=True)

            output_path = str(exports_dir / filename)

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = self.report_name[:31]  # Excel sheet name limit

        # Add metadata
        ws["A1"] = "Report:"
        ws["B1"] = self.report_name
        ws["A2"] = "Generated:"
        ws["B2"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws["A3"] = "Total Records:"
        ws["B3"] = len(self.results)

        # Style metadata
        for cell in ["A1", "A2", "A3"]:
            ws[cell].font = Font(bold=True)

        # Add blank row
        start_row = 5

        if self.results:
            # Add headers
            headers = list(self.results[0].keys())
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=start_row, column=col_idx, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Add data
            for row_idx, row_data in enumerate(self.results, start=start_row + 1):
                for col_idx, header in enumerate(headers, start=1):
                    value = row_data.get(header)
                    ws.cell(row=row_idx, column=col_idx, value=value)

            # Auto-adjust column widths
            for col_idx, header in enumerate(headers, start=1):
                column_letter = get_column_letter(col_idx)
                max_length = len(str(header))

                for row_data in self.results:
                    value = row_data.get(header)
                    if value is not None:
                        max_length = max(max_length, len(str(value)))

                # Set width with some padding
                adjusted_width = min(max_length + 2, 50)  # Max 50 chars
                ws.column_dimensions[column_letter].width = adjusted_width

        # Save workbook
        wb.save(output_path)
        log.info(f"Report exported to: {output_path}")

        return output_path

    def to_csv(self, output_path: Optional[str] = None) -> str:
        """Export results to CSV format.

        Args:
            output_path: Optional output file path

        Returns:
            Path to exported file
        """
        import csv

        # Generate output path if not provided
        if not output_path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            safe_name = "".join(c for c in self.report_name if c.isalnum() or c in (" ", "_", "-")).strip()
            filename = f"{safe_name}_{timestamp}.csv"

            exports_dir = Path(DIRECTORIO_APP) / "exports" / "reports"
            exports_dir.mkdir(parents=True, exist_ok=True)

            output_path = str(exports_dir / filename)

        # Write CSV
        if self.results:
            headers = list(self.results[0].keys())

            with open(output_path, "w", newline="", encoding="utf-8") as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=headers)
                writer.writeheader()
                writer.writerows(self.results)

        log.info(f"Report exported to: {output_path}")
        return output_path


def export_report_to_excel(report_name: str, results: List[Dict[str, Any]], output_path: Optional[str] = None) -> str:
    """Convenience function to export report to Excel.

    Args:
        report_name: Name of the report
        results: Report results
        output_path: Optional output path

    Returns:
        Path to exported file
    """
    exporter = ReportExporter(report_name, results)
    return exporter.to_excel(output_path)


def export_report_to_csv(report_name: str, results: List[Dict[str, Any]], output_path: Optional[str] = None) -> str:
    """Convenience function to export report to CSV.

    Args:
        report_name: Name of the report
        results: Report results
        output_path: Optional output path

    Returns:
        Path to exported file
    """
    exporter = ReportExporter(report_name, results)
    return exporter.to_csv(output_path)
